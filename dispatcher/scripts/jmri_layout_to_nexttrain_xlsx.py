#!/usr/bin/env python3
"""
Read JMRI panel tables.xml LayoutEditor geometry and refresh Segments + ControlPoints
in NextTrainDispatcherApp.xlsx.

By default exports the **middle horizontal third** (midpoint-Y of each segment) and
rebases coordinates. `--crop-method cluster` restores k-means with `--middle-along`.
Use `--whole-layout` for the full panel.

Labels on the schematic (positionablelabel → station in older exports) are skipped.

Usage:
  python3 dispatcher/scripts/jmri_layout_to_nexttrain_xlsx.py \\
    --tables dispatcher/inputs/tables.xml \\
    --workbook dispatcher/exports/NextTrainDispatcherApp.xlsx
"""

from __future__ import annotations

import argparse
import json
import math
import os
import re
import shutil
import sys
import unicodedata
import xml.etree.ElementTree as ET
from pathlib import Path

import openpyxl

JMRI_ROOT = Path(__file__).resolve().parents[2] / "jmri"
sys.path.insert(0, str(JMRI_ROOT))
from layout_paths import layout_paths  # noqa: E402

DEFAULT_COORD_SCALE = 1.0
DEFAULT_SEGMENT_SCALE = 2.0
DEFAULT_CONTROL_POINT_SCALE = 2.0
DEFAULT_OFFSET_X = 48.0
DEFAULT_OFFSET_Y = 36.0
# Sheet-coordinate gap above which we add connector lines (cols K–T)
CONNECTION_SNAP_TOLERANCE = 6.0

CONNECTION_FIELD_KEYS = (
    "connectsToSegmentId",
    "connectionStartX",
    "connectionStartY",
    "connectionEndX",
    "connectionEndY",
    "connectsToSegmentId2",
    "connection2StartX",
    "connection2StartY",
    "connection2EndX",
    "connection2EndY",
)


def nfkc(s: str) -> str:
    return unicodedata.normalize("NFKC", s or "")


def slug(s: str) -> str:
    s = nfkc(s).lower().strip()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-") or "pt"


def fnum(v: str | None) -> float | None:
    if v is None:
        return None
    try:
        return float(v)
    except ValueError:
        return None


LEG_MAP = {
    "TURNOUT_A": ("xa", "ya"),
    "TURNOUT_B": ("xb", "yb"),
    "TURNOUT_C": ("xc", "yc"),
    "TURNOUT_D": ("xd", "yd"),
}


def load_layout(paths: Path) -> ET.Element:
    tree = ET.parse(paths)
    root = tree.getroot()
    layouts = root.findall(".//LayoutEditor")
    if not layouts:
        raise SystemExit(f"No LayoutEditor found in {paths}")
    return layouts[0]


def collect_turnouts(layout: ET.Element) -> dict[str, dict]:
    turnouts: dict[str, dict] = {}
    for el in layout.findall(".//layoutturnout"):
        ident = el.get("ident")
        if not ident:
            continue
        d: dict[str, float | str] = {}
        for k in ("xa", "ya", "xb", "yb", "xc", "yc", "xd", "yd", "xcen", "ycen"):
            fv = fnum(el.get(k))
            if fv is not None:
                d[k] = fv
        tn = nfkc(el.get("turnoutname") or "").strip()
        d["display_name"] = tn or ident
        turnouts[ident] = d
    return turnouts


def collect_positionable_points(layout: ET.Element) -> dict[str, tuple[float, float]]:
    pts: dict[str, tuple[float, float]] = {}
    for el in layout.findall(".//positionablepoint"):
        ident = el.get("ident")
        if not ident:
            continue
        x = fnum(el.get("x"))
        y = fnum(el.get("y"))
        if x is None or y is None:
            continue
        pts[ident] = (x, y)
    return pts


def resolve_endpoint(
    name: str,
    kind: str,
    points: dict[str, tuple[float, float]],
    turnouts: dict[str, dict],
) -> tuple[float, float] | None:
    if kind == "POS_POINT":
        return points.get(name)
    if kind in LEG_MAP:
        t = turnouts.get(name)
        if not t:
            return None
        xk, yk = LEG_MAP[kind]
        xv = t.get(xk)
        yv = t.get(yk)
        if xv is None or yv is None:
            # AnyRail exports often omit xa/ya; use turnout center for throat (A).
            xc = t.get("xcen")
            yc = t.get("ycen")
            if xc is not None and yc is not None:
                return (float(xc), float(yc))
            return None
        return (float(xv), float(yv))
    return None


def segment_type(hidden: bool, dashed: str | None, mainline: str | None) -> str:
    if hidden:
        return "industrial"
    if dashed == "yes":
        return "siding"
    if mainline == "yes":
        return "mainline"
    return "yard"


def miles_placeholder(dx: float, dy: float) -> float:
    dist = math.hypot(dx, dy)
    return round(max(0.01, dist / 850.0), 3)


def segment_display_name(block: str, ident: str) -> str:
    b = nfkc(block).strip()
    return b if b else ident


def kmeans3_labels(
    xy: list[tuple[float, float]], max_iter: int = 40
) -> tuple[list[int], list[tuple[float, float]]]:
    """Lloyd k-means, k=3. Returns (label per point, final centroids)."""
    n = len(xy)
    k = 3
    if n < k:
        return [0] * n, [xy[0]] * k if n else []
    xs = [p[0] for p in xy]
    ys = [p[1] for p in xy]
    rx = max(xs) - min(xs)
    ry = max(ys) - min(ys)
    primary = 0 if rx >= ry else 1
    order = sorted(range(n), key=lambda i: xy[i][primary])
    idx0 = order[n // 6]
    idx1 = order[n // 2]
    idx2 = order[(5 * n) // 6]
    centroids = [xy[idx0], xy[idx1], xy[idx2]]

    labels = [0] * n
    for _ in range(max_iter):
        prev_l = labels[:]
        for i, p in enumerate(xy):
            best = min(
                range(k),
                key=lambda j: (p[0] - centroids[j][0]) ** 2 + (p[1] - centroids[j][1]) ** 2,
            )
            labels[i] = best
        new_c: list[tuple[float, float]] = []
        for j in range(k):
            members = [xy[i] for i in range(n) if labels[i] == j]
            if not members:
                far = max(
                    range(n),
                    key=lambda i: min(
                        (xy[i][0] - centroids[m][0]) ** 2 + (xy[i][1] - centroids[m][1]) ** 2
                        for m in range(k)
                    ),
                )
                new_c.append(xy[far])
            else:
                mx = sum(q[0] for q in members) / len(members)
                my = sum(q[1] for q in members) / len(members)
                new_c.append((mx, my))
        centroids = new_c
        if labels == prev_l:
            break
    return labels, centroids


def dominant_axis(xy: list[tuple[float, float]]) -> int:
    if not xy:
        return 0
    xs = [p[0] for p in xy]
    ys = [p[1] for p in xy]
    return 0 if (max(xs) - min(xs)) >= (max(ys) - min(ys)) else 1


def centroid_axis_spreads(centroids: list[tuple[float, float]]) -> tuple[float, float]:
    """How far apart cluster centroids lie on X vs Y (for auto axis)."""
    xs = [c[0] for c in centroids]
    ys = [c[1] for c in centroids]
    return max(xs) - min(xs), max(ys) - min(ys)


def resolve_section_axis(mode: str, centroids: list[tuple[float, float]]) -> int:
    """
    Axis used to decide which cluster is geometrically ``middle`:
    - 0 = median along X (center column among three blobs)
    - 1 = median along Y (center horizontal band among three blobs)
    """
    if mode == "x":
        return 0
    if mode == "y":
        return 1
    sx, sy = centroid_axis_spreads(centroids)
    # Prefer whichever axis separates the three k-means centroids more clearly.
    # Tie favors Y — matches typical ``north / mains / staging'' striping plus user feedback.
    return 1 if sy >= sx else 0


def resolve_thirds_axis(mode: str, mids: list[tuple[float, float]]) -> int:
    """Fallback when fewer midpoints — use midpoint cloud extent if axis is ``auto''."""
    if mode == "x":
        return 0
    if mode == "y":
        return 1
    return dominant_axis(mids)


def middle_cluster_median_on_axis(centroids: list[tuple[float, float]], axis: int) -> int:
    order = sorted(range(len(centroids)), key=lambda j: centroids[j][axis])
    return order[len(centroids) // 2]


def thirds_middle_mask(midpoints: list[tuple[float, float]], axis: int) -> list[bool]:
    """Middle third along given axis (0=X, 1=Y)."""
    if not midpoints:
        return []
    vals = sorted(p[axis] for p in midpoints)
    lo, hi = vals[0], vals[-1]
    span = hi - lo or 1.0
    a = lo + span / 3.0
    b = lo + 2.0 * span / 3.0
    return [a <= p[axis] <= b for p in midpoints]


def bbox_endpoints(segs: list[dict]) -> tuple[float, float, float, float]:
    xs: list[float] = []
    ys: list[float] = []
    for s in segs:
        xs += [s["sx"], s["ex"]]
        ys += [s["sy"], s["ey"]]
    return min(xs), min(ys), max(xs), max(ys)


def inflate(bb: tuple[float, float, float, float], m: float) -> tuple[float, float, float, float]:
    x0, y0, x1, y1 = bb
    return x0 - m, y0 - m, x1 + m, y1 + m


def in_bbox(x: float, y: float, bb: tuple[float, float, float, float]) -> bool:
    x0, y0, x1, y1 = bb
    return x0 <= x <= x1 and y0 <= y <= y1


def junction_key(name: str, kind: str) -> str | None:
    """Group segment ends that meet at the same anchor or turnout."""
    if kind == "POS_POINT":
        return f"PT:{name}"
    if kind in LEG_MAP:
        return f"TO:{name}"
    return None


def closest_point_on_segment(
    px: float, py: float, ax: float, ay: float, bx: float, by: float
) -> tuple[float, float]:
    dx, dy = bx - ax, by - ay
    len2 = dx * dx + dy * dy
    if len2 < 1e-12:
        return ax, ay
    t = max(0.0, min(1.0, ((px - ax) * dx + (py - ay) * dy) / len2))
    return ax + t * dx, ay + t * dy


def infer_segment_connections(
    kept: list[dict],
    sheet_coords: dict[str, tuple[float, float, float, float]],
    snap_tol: float = CONNECTION_SNAP_TOLERANCE,
) -> dict[str, list[dict[str, object]]]:
    """
    Infer optional connector lines (spreadsheet columns K–T) for ends that share a
    JMRI junction but do not coincide after export/scaling (typical at turnouts).
    """
    attachments: dict[str, list[dict]] = {}
    for s in kept:
        ident = s["ident"]
        sc = sheet_coords[ident]
        for end_key, node_name, node_kind in (
            ("s", s["n1"], s["t1"]),
            ("e", s["n2"], s["t2"]),
        ):
            jk = junction_key(node_name, node_kind)
            if not jk:
                continue
            x = sc[0] if end_key == "s" else sc[2]
            y = sc[1] if end_key == "s" else sc[3]
            attachments.setdefault(jk, []).append(
                {"seg": ident, "x": x, "y": y}
            )

    proposals: list[tuple[str, str, float, float, float, float, float]] = []
    for group in attachments.values():
        if len(group) < 2:
            continue
        for i, a in enumerate(group):
            for b in group[i + 1 :]:
                if a["seg"] == b["seg"]:
                    continue
                dist = math.hypot(a["x"] - b["x"], a["y"] - b["y"])
                if dist <= snap_tol:
                    continue
                other = sheet_coords[b["seg"]]
                ex, ey = closest_point_on_segment(
                    a["x"], a["y"], other[0], other[1], other[2], other[3]
                )
                # One direction per pair (avoid duplicate overlay lines)
                if a["seg"] < b["seg"]:
                    proposals.append(
                        (a["seg"], b["seg"], a["x"], a["y"], ex, ey, dist)
                    )
                else:
                    ex2, ey2 = closest_point_on_segment(
                        b["x"], b["y"], sheet_coords[a["seg"]][0],
                        sheet_coords[a["seg"]][1],
                        sheet_coords[a["seg"]][2],
                        sheet_coords[a["seg"]][3],
                    )
                    proposals.append(
                        (b["seg"], a["seg"], b["x"], b["y"], ex2, ey2, dist)
                    )

    proposals.sort(key=lambda p: p[6])
    by_seg: dict[str, list[dict[str, object]]] = {}
    for from_seg, to_seg, sx, sy, ex, ey, _dist in proposals:
        bucket = by_seg.setdefault(from_seg, [])
        if len(bucket) >= 2:
            continue
        if any(c["connectsToSegmentId"] == to_seg for c in bucket):
            continue
        bucket.append(
            {
                "connectsToSegmentId": to_seg,
                "connectionStartX": round(sx, 3),
                "connectionStartY": round(sy, 3),
                "connectionEndX": round(ex, 3),
                "connectionEndY": round(ey, 3),
            }
        )

    return by_seg


def connection_fields_for_segment(
    seg_id: str, inferred: dict[str, list[dict[str, object]]]
) -> list[object]:
    """Ten values for columns K–T (or None)."""
    conns = inferred.get(seg_id, [])
    out: list[object] = []
    for slot in range(2):
        if slot < len(conns):
            c = conns[slot]
            out.extend(
                [
                    c["connectsToSegmentId"],
                    c["connectionStartX"],
                    c["connectionStartY"],
                    c["connectionEndX"],
                    c["connectionEndY"],
                ]
            )
        else:
            out.extend([None, None, None, None, None])
    return out


def transform_coord(
    x: float,
    y: float,
    origin: tuple[float, float],
    scale: float,
    offset: tuple[float, float],
) -> tuple[float, float]:
    ox, oy = origin
    dx, dy = offset
    return (x - ox) * scale + dx, (y - oy) * scale + dy


def load_export_options(path: Path) -> dict[str, float]:
    if not path.is_file():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        return {k: float(v) for k, v in data.items() if isinstance(v, (int, float))}
    except (json.JSONDecodeError, TypeError, ValueError):
        return {}


def ensure_workbook(workbook: Path, template: Path) -> None:
    workbook.parent.mkdir(parents=True, exist_ok=True)
    if workbook.is_file():
        return
    if not template.is_file():
        raise SystemExit(
            f"Workbook missing at {workbook} and no template at {template}"
        )
    shutil.copy2(template, workbook)
    print(f"  Created workbook from template: {workbook}")


def resolve_layout_paths() -> dict[str, str]:
    layout = os.environ.get("JMRI_LAYOUT", "").strip()
    if not layout:
        return {}
    return layout_paths(layout)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--tables",
        type=Path,
        default=Path(__file__).resolve().parent.parent / "inputs" / "tables.xml",
    )
    ap.add_argument(
        "--workbook",
        type=Path,
        default=Path(__file__).resolve().parent.parent / "exports" / "NextTrainDispatcherApp.xlsx",
    )
    ap.add_argument(
        "--include-hidden",
        action="store_true",
        help='Include LayoutEditor segments with hidden="yes" (normally skipped).',
    )
    ap.add_argument(
        "--whole-layout",
        action="store_true",
        help="Export full panel (no middle-section crop).",
    )
    ap.add_argument(
        "--bbox-margin",
        type=float,
        default=12.0,
        help="Extra layout units around segment bbox for control-point filtering.",
    )
    ap.add_argument(
        "--middle-along",
        choices=("auto", "x", "y"),
        default="y",
        help="For k-means / x-thirds cropping: median cluster along this axis.",
    )
    ap.add_argument(
        "--crop-method",
        choices=("y-thirds", "x-thirds", "cluster"),
        default="y-thirds",
        help='How to pick sub-panel: ``y-thirds'' = middle horizontal band by layout Y '
        '(best for stacked sections); ``cluster'' = k-means 3 regions + median on --middle-along.',
    )
    ap.add_argument(
        "--coord-scale",
        type=float,
        default=None,
        help="Scale for control points (default from export_options or 1.0).",
    )
    ap.add_argument(
        "--segment-scale",
        type=float,
        default=None,
        help=f"Scale segment endpoints (default {DEFAULT_SEGMENT_SCALE} = +100%%).",
    )
    ap.add_argument(
        "--control-point-scale",
        type=float,
        default=None,
        help=f"Scale control-point coordinates (default {DEFAULT_CONTROL_POINT_SCALE}, same as segments).",
    )
    ap.add_argument(
        "--offset-x",
        type=float,
        default=None,
        help=f"Left margin in spreadsheet units after scale (default {DEFAULT_OFFSET_X}).",
    )
    ap.add_argument(
        "--offset-y",
        type=float,
        default=None,
        help=f"Top margin in spreadsheet units after scale (default {DEFAULT_OFFSET_Y}).",
    )
    ap.add_argument(
        "--export-options",
        type=Path,
        default=None,
        help="JSON with coord_scale, offset_x, offset_y (overrides defaults).",
    )
    ap.add_argument(
        "--no-infer-connections",
        action="store_true",
        help="Leave columns K–T empty (do not infer turnout/anchor connectors).",
    )
    ap.add_argument(
        "--connection-snap",
        type=float,
        default=None,
        help=f"Max gap (sheet units) before adding a connector line (default {CONNECTION_SNAP_TOLERANCE}).",
    )
    args = ap.parse_args()

    lp = resolve_layout_paths()
    if lp:
        if args.tables == Path(__file__).resolve().parent.parent / "inputs" / "tables.xml":
            if os.path.isfile(lp["dispatcher_tables"]):
                args.tables = Path(lp["dispatcher_tables"])
            elif os.path.isfile(lp["output"]):
                args.tables = Path(lp["output"])
        default_wb = Path(__file__).resolve().parent.parent / "exports" / "NextTrainDispatcherApp.xlsx"
        if args.workbook == default_wb and lp.get("dispatcher_workbook"):
            args.workbook = Path(lp["dispatcher_workbook"])
        opts_path = args.export_options
        if opts_path is None and lp.get("dispatcher_dir"):
            opts_path = Path(lp["dispatcher_dir"]) / "export_options.json"
        if opts_path and opts_path.is_file():
            file_opts = load_export_options(opts_path)
            if args.coord_scale is None and "coord_scale" in file_opts:
                args.coord_scale = file_opts["coord_scale"]
            if args.segment_scale is None and "segment_scale" in file_opts:
                args.segment_scale = file_opts["segment_scale"]
            if args.control_point_scale is None and "control_point_scale" in file_opts:
                args.control_point_scale = file_opts["control_point_scale"]
            if args.offset_x is None and "offset_x" in file_opts:
                args.offset_x = file_opts["offset_x"]
            if args.offset_y is None and "offset_y" in file_opts:
                args.offset_y = file_opts["offset_y"]
            if args.connection_snap is None and "connection_snap_tolerance" in file_opts:
                args.connection_snap = file_opts["connection_snap_tolerance"]
        template = Path(lp.get("dispatcher_template_workbook", ""))
        ensure_workbook(args.workbook, template)

    control_point_scale = (
        args.control_point_scale
        if args.control_point_scale is not None
        else (args.coord_scale if args.coord_scale is not None else DEFAULT_CONTROL_POINT_SCALE)
    )
    segment_scale = (
        args.segment_scale if args.segment_scale is not None else DEFAULT_SEGMENT_SCALE
    )
    offset_x = args.offset_x if args.offset_x is not None else DEFAULT_OFFSET_X
    offset_y = args.offset_y if args.offset_y is not None else DEFAULT_OFFSET_Y
    connection_snap = (
        args.connection_snap
        if args.connection_snap is not None
        else CONNECTION_SNAP_TOLERANCE
    )

    layout = load_layout(args.tables)
    points = collect_positionable_points(layout)
    turnouts = collect_turnouts(layout)

    raw: list[dict] = []
    skipped = 0

    for el in layout.findall("tracksegment"):
        ident = el.get("ident")
        block = el.get("blockname") or ""
        if not ident:
            continue
        hidden = el.get("hidden") == "yes"
        if hidden and not args.include_hidden:
            continue
        n1 = el.get("connect1name") or ""
        t1 = el.get("type1") or ""
        n2 = el.get("connect2name") or ""
        t2 = el.get("type2") or ""
        if not (n1 and t1 and n2 and t2):
            skipped += 1
            continue
        p1 = resolve_endpoint(n1, t1, points, turnouts)
        p2 = resolve_endpoint(n2, t2, points, turnouts)
        if not p1 or not p2:
            skipped += 1
            continue
        sx, sy = p1[0], p1[1]
        ex, ey = p2[0], p2[1]
        raw.append(
            {
                "ident": ident,
                "block": block,
                "sx": sx,
                "sy": sy,
                "ex": ex,
                "ey": ey,
                "n1": n1,
                "t1": t1,
                "n2": n2,
                "t2": t2,
                "hidden": hidden,
                "dashed": el.get("dashed"),
                "mainline": el.get("mainline"),
            }
        )

    mids = [((s["sx"] + s["ex"]) / 2, (s["sy"] + s["ey"]) / 2) for s in raw]

    section_axis_detail = ""

    if args.whole_layout or not raw:
        keep_mask = [True] * len(raw)
        section_axis_detail = "whole layout" if args.whole_layout else "empty (no crop)"
    else:
        cm = args.crop_method
        if cm == "y-thirds":
            keep_mask = thirds_middle_mask(mids, axis=1)
            section_axis_detail = (
                "middle third of midpoint Y (--crop-method y-thirds, panel stacked by height)"
            )
        elif cm == "x-thirds":
            keep_mask = thirds_middle_mask(mids, axis=0)
            section_axis_detail = "middle third of midpoint X (--crop-method x-thirds)"
        elif len(mids) >= 12:
            labels, centroids = kmeans3_labels(mids)
            sax = resolve_section_axis(args.middle_along, centroids)
            mid_lab = middle_cluster_median_on_axis(centroids, sax)
            keep_mask = [labels[i] == mid_lab for i in range(len(raw))]
            how = "X (median cluster)" if sax == 0 else "Y (median cluster)"
            section_axis_detail = (
                f"k-means + {how}; --middle-along={args.middle_along} (--crop-method cluster)"
            )
        else:
            sax = resolve_thirds_axis(args.middle_along, mids)
            keep_mask = thirds_middle_mask(mids, sax)
            how = "X" if sax == 0 else "Y"
            section_axis_detail = (
                f"middle third on {how}, --middle-along={args.middle_along} (few segments)"
            )
        if not any(keep_mask):
            keep_mask = [True] * len(raw)
            section_axis_detail += " → fell back to all segments"

    kept = [s for s, k in zip(raw, keep_mask) if k]
    bb = bbox_endpoints(kept)
    span = max(bb[2] - bb[0], bb[3] - bb[1]) or 1.0
    margin = max(args.bbox_margin, span * 0.02)
    bb_filter = inflate(bb, margin)
    ox, oy = bb[0], bb[1]
    origin = (ox, oy)
    offset = (offset_x, offset_y)

    def tx_segment(x: float, y: float) -> tuple[float, float]:
        return transform_coord(x, y, origin, segment_scale, offset)

    def tx_control(x: float, y: float) -> tuple[float, float]:
        return transform_coord(x, y, origin, control_point_scale, offset)

    sheet_coords: dict[str, tuple[float, float, float, float]] = {}
    for s in kept:
        sx, sy = tx_segment(s["sx"], s["sy"])
        ex, ey = tx_segment(s["ex"], s["ey"])
        sx, sy, ex, ey = round(sx, 3), round(sy, 3), round(ex, 3), round(ey, 3)
        sheet_coords[s["ident"]] = (sx, sy, ex, ey)

    inferred: dict[str, list[dict[str, object]]] = {}
    if not args.no_infer_connections:
        inferred = infer_segment_connections(kept, sheet_coords, snap_tol=connection_snap)

    segments: list[list[object]] = []
    conn_count = 0
    for s in kept:
        sx, sy, ex, ey = sheet_coords[s["ident"]]
        dx = ex - sx
        dy = ey - sy
        conn_cols = connection_fields_for_segment(s["ident"], inferred)
        conn_count += sum(1 for c in conn_cols[::5] if c is not None)
        segments.append(
            [
                s["ident"],
                1,
                segment_display_name(s["block"], s["ident"]),
                sx,
                sy,
                ex,
                ey,
                miles_placeholder(dx, dy),
                79,
                segment_type(s["hidden"], s["dashed"], s["mainline"]),
                *conn_cols,
            ]
        )

    ctrl: list[list[object]] = []
    nid = 0

    def add_cp(cid: str, name: str, x: float, y: float, typ: str) -> None:
        if not in_bbox(x, y, bb_filter):
            return
        cx, cy = tx_control(x, y)
        ctrl.append([cid, nfkc(name)[:80], round(cx, 3), round(cy, 3), typ])

    # No positionablelabel → station (reduces clutter)

    for el in layout.findall("signalmasticon"):
        name = nfkc(el.get("signalmast") or "").strip() or "Signal"
        x = fnum(el.get("x"))
        y = fnum(el.get("y"))
        if x is None or y is None:
            continue
        nid += 1
        add_cp(f"sig-{nid}", name, x, y, "signal")

    for ident, t in turnouts.items():
        xc = t.get("xcen")
        yc = t.get("ycen")
        if xc is None or yc is None:
            continue
        nid += 1
        add_cp(f"sw-{slug(ident)}-{nid}", ident, float(xc), float(yc), "switch")

    for pid, (x, y) in points.items():
        pel = next(
            (e for e in layout.findall("positionablepoint") if e.get("ident") == pid),
            None,
        )
        if pel is None or pel.get("type") != "END_BUMPER":
            continue
        nid += 1
        add_cp(f"bump-{pid}", pid, x, y, "staging")

    wb = openpyxl.load_workbook(args.workbook)
    ws = wb["Segments"]
    hdr_row = next(ws.iter_rows(min_row=1, max_row=1))
    hdr_len = max((c.column for c in hdr_row), default=20)

    prev_max = ws.max_row or 1
    if prev_max > 1:
        ws.delete_rows(2, prev_max - 1)
    for row in segments:
        filled = row + [None] * max(0, hdr_len - len(row))
        ws.append(filled[:hdr_len])

    wsc = wb["ControlPoints"]
    prev_c = wsc.max_row or 1
    if prev_c > 1:
        wsc.delete_rows(2, prev_c - 1)
    for row in ctrl:
        wsc.append(row)

    wb.save(args.workbook)
    print(f"Saved {args.workbook}")
    mode = "whole layout" if args.whole_layout else args.crop_method
    print(f"Mode: {mode}")
    print(f"Cropping: {section_axis_detail}")
    print(f"Segments: {len(segments)} rows (skipped unresolved: {skipped})")
    print(f"ControlPoints: {len(ctrl)} rows (no schematic text labels)")
    if not args.no_infer_connections:
        print(f"Connector lines (cols K–T): {conn_count} slot(s) filled")
    print(
        f"Transform: segment_scale={segment_scale}, control_point_scale={control_point_scale}, "
        f"offset=({offset_x}, {offset_y}) (rebased from layout min {origin[0]:.1f}, {origin[1]:.1f})"
    )
    if skipped:
        print("Note: unresolved segments omitted (missing refs / geometry)")


if __name__ == "__main__":
    main()
