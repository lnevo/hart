#!/usr/bin/env python3
"""Build industry routing review HTML — same portal format as HART Device Map."""
from __future__ import annotations

import html
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

SCRIPTS = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPTS))

from lib.review_portal_html import callout, filter_script, page, stat_grid  # noqa: E402

CON = Path(__file__).resolve().parents[1]
XLSX = CON / "external/hart-ops/industries/HART_Industry_Routing_Matrix.xlsx"
OUT = CON / "html/review/industry-matrix.html"

GRAMMAR_ROWS = [
    ["IN", "Brick Yard", "Coal", "Penn Coal Co.", "Local", "—", "—", "Hopper"],
    ["OUT", "Plane Scale", "Steel coils", "US Steel", "Local", "—", "—", "Gondola"],
    ["IN", "Interchange", "Lumber", "Regional Lumber", "Foreign", "B&O", "McKees Rocks", "Boxcar"],
]

SHEETS = [
    ("Industry Routing Matrix", "hart-lanes", "lane"),
    ("Interchange_Matrix", "interchange", "interchange"),
]


def sheet_rows(wb, name: str) -> tuple[list[str], list[list[str]]]:
    if name not in wb.sheetnames:
        return [], []
    ws = wb[name]
    raw = list(ws.iter_rows(values_only=True))
    if not raw:
        return [], []
    headers = [str(c or "").strip() for c in raw[0]]
    data: list[list[str]] = []
    for row in raw[1:]:
        if not any(row):
            continue
        cells = [str(c or "").strip() for c in row]
        while len(cells) < len(headers):
            cells.append("")
        data.append(cells[: len(headers)])
    return headers, data


def render_section(
    title: str,
    table_id: str,
    kind_attr: str,
    headers: list[str],
    rows: list[list[str]],
) -> str:
    head = "".join(f"<th>{html.escape(h)}</th>" for h in headers)
    body_rows = []
    for cells in rows:
        kind_val = cells[1] if len(cells) > 1 and kind_attr == "data-flow" else kind_attr.replace("data-", "")
        if kind_attr == "data-sheet":
            kind_val = table_id
        search = " ".join(cells).lower()
        attrs = f"{kind_attr}='{html.escape(kind_val)}' data-search='{html.escape(search)}'"
        if kind_attr == "data-flow" and cells:
            attrs += f" data-industry='{html.escape(cells[0].lower())}'"
        tds = "".join(f"<td>{html.escape(c) if c else '—'}</td>" for c in cells)
        body_rows.append(f"<tr {attrs}>{tds}</tr>")
    return f"""
<h2>{html.escape(title)}</h2>
<p class="subhead section-count" data-table="{table_id}">{len(rows)} rows</p>
<div class="table-wrap">
<table class="review-table" id="{table_id}">
  <thead><tr>{head}</tr></thead>
  <tbody>{''.join(body_rows)}</tbody>
</table></div>
"""


def main() -> int:
    try:
        import openpyxl
    except ImportError:
        print("openpyxl required", file=sys.stderr)
        return 1
    if not XLSX.is_file():
        print(f"MISSING {XLSX}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(XLSX, read_only=True)
    sections = []
    all_lane_rows: list[list[str]] = []
    flows: Counter[str] = Counter()
    industries: set[str] = set()

    for sheet_name, table_id, _ in SHEETS:
        headers, data = sheet_rows(wb, sheet_name)
        if not data:
            continue
        kind_attr = "data-flow" if sheet_name == "Industry Routing Matrix" else "data-sheet"
        sections.append(render_section(sheet_name, table_id, kind_attr, headers, data))
        if sheet_name == "Industry Routing Matrix":
            all_lane_rows = data
            for row in data:
                if row:
                    industries.add(row[0])
                    if len(row) > 1 and row[1]:
                        flows[row[1]] += 1
    wb.close()

    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    stats = [
        (str(len(all_lane_rows)), "HART lanes"),
        (str(flows.get("IN", 0)), "IN flows"),
        (str(flows.get("OUT", 0)), "OUT flows"),
        (str(len(industries)), "Industries"),
    ]

    grammar_head = "".join(f"<th>{html.escape(h)}</th>" for h in [
        "Industry", "Flow", "Commodity / Product", "Supplier / Customer",
        "Region", "Foreign Railroad", "Interchange", "Car Type",
    ])
    grammar_body = "".join(
        "<tr>" + "".join(f"<td>{html.escape(c)}</td>" for c in row) + "</tr>"
        for row in GRAMMAR_ROWS
    )

    body = f"""
<h1>Industry routing matrix — review</h1>
<p class="review-lead">
  From <code>HART_Industry_Routing_Matrix.xlsx</code> in hart-ops. Cross-check lane changes with
  <code>data/HART_Spot_Waybills.csv</code> and spot assignments. Generated {ts}.
</p>
{callout(
    "Lane grammar",
    "Each row is one supplier/customer lane: Industry + IN/OUT flow + commodity + party + region. "
    "Interchange rows map foreign railroad handoffs. When a lane changes, regenerate spot waybills "
    "and validate STS seed consistency.",
)}
{stat_grid(stats)}
<h2>Grammar</h2>
<table class="review-table">
<thead><tr>{grammar_head}</tr></thead>
<tbody>{grammar_body}</tbody>
</table>
<h2>All lanes</h2>
<p class="subhead" id="row-count">{len(all_lane_rows)} of {len(all_lane_rows)} shown</p>
<div class="review-toolbar">
  <select id="kind-filter">
    <option value="all">All tables</option>
    <option value="lane">HART lanes only</option>
    <option value="IN">IN flows</option>
    <option value="OUT">OUT flows</option>
    <option value="interchange">Interchange matrix</option>
  </select>
  <input type="search" id="q" placeholder="Filter industry, commodity, car type…" />
</div>
{''.join(sections)}
<script>
(function() {{
  const kindSel = document.getElementById('kind-filter');
  const q = document.getElementById('q');
  const countEl = document.getElementById('row-count');
  const tables = ['hart-lanes', 'interchange'];
  function allRows() {{
    return tables.flatMap(id => [...document.querySelectorAll('#' + id + ' tbody tr')]);
  }}
  function apply() {{
    const kind = kindSel.value;
    const term = (q.value || '').toLowerCase();
    let shown = 0;
    let total = 0;
    allRows().forEach(tr => {{
      total++;
      const flow = tr.getAttribute('data-flow') || '';
      const sheet = tr.closest('table').id === 'interchange' ? 'interchange' : 'lane';
      let okKind = kind === 'all'
        || (kind === 'lane' && sheet === 'lane')
        || (kind === 'interchange' && sheet === 'interchange')
        || (kind === flow);
      const hay = (tr.getAttribute('data-search') || '').toLowerCase();
      const ok = okKind && (!term || hay.includes(term));
      tr.style.display = ok ? '' : 'none';
      if (ok) shown++;
    }});
    countEl.textContent = shown + ' of ' + total + ' shown';
  }}
  kindSel.addEventListener('change', apply);
  q.addEventListener('input', apply);
  apply();
}})();
</script>
"""
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(page("Industry routing — review", "industry", body), encoding="utf-8")
    print(f"Wrote {OUT} ({len(all_lane_rows)} lanes)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
