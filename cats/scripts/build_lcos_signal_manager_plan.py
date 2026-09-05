#!/usr/bin/env python3
"""Refresh LCOS Signal Manager plan: Stop/Clear ports + Main/Diverged + OR.

Reads cats/data/signal_wiring.csv (R/Y/G pins) and lcos_signal_logic.csv.
Keeps plant/block notes already in the workbook; rewrites Port A/B/C from
the live wiring (Stop = R pin, Approach = Y, Clear = G).
"""
from __future__ import annotations

import csv
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parents[2]
DATA = ROOT / "cats/data"
PLAN = DATA / "lcos_signal_manager_config_plan.xlsx"
PORT_RE = re.compile(r"^C(\d+)-OU(\d+)-(\d+)$")


def port_index(port_id: str) -> int:
    m = PORT_RE.match(port_id)
    if not m:
        raise SystemExit(f"bad port_id {port_id}")
    ou, pin = int(m.group(2)), int(m.group(3))
    return (ou - 1) * 8 + (pin - 1)


def load_heads() -> dict[tuple[str, str], dict]:
    by: dict[tuple[str, int], dict[str, str]] = defaultdict(dict)
    meta: dict[tuple[str, int], dict[str, str]] = {}
    with (DATA / "signal_wiring.csv").open(newline="") as f:
        for row in csv.DictReader(f):
            mast = row["mast_user_name"]
            packed = int(row["packed"])
            by[(mast, packed)][row["lamp_color"]] = row["port_id"]
            meta[(mast, packed)] = row
    out = {}
    for (mast, packed), lamps in by.items():
        row = meta[(mast, packed)]
        disc = row["disc_role"]
        out[(mast, disc)] = {
            "stop": lamps["R"],
            "approach": lamps["Y"],
            "clear": lamps["G"],
            "signal_index": row["signal_index"],
            "uid": row["uid"],
            "packed": row["packed"],
            "node": row["parent_node_id"],
            "mqtt": row["mqtt_node"],
            "location": row["board_location"],
        }
    return out


def load_logic() -> dict[tuple[str, str], dict[str, str]]:
    out = {}
    with (DATA / "lcos_signal_logic.csv").open(newline="") as f:
        for row in csv.DictReader(f):
            out[(row["mast"], row["head"])] = row
    return out


def main() -> int:
    heads = load_heads()
    logic = load_logic()
    wb = load_workbook(PLAN)
    legend = wb["Legend"]
    legend["A10"] = (
        "  Aspect map: Port A = STOP (R), Port B = APPROACH (Y), Port C = CLEAR (G). "
        "Field wiring is Stop, Approach, Clear on consecutive increasing ports "
        "(R then Y then G). Do not use the old G-first docs."
    )
    if not str(legend["A11"].value or "").startswith("  Turnout polarity"):
        legend.insert_rows(11, 2)
    legend["A11"] = (
        "  Turnout polarity: Main = Normal (closed); Diverged = Reverse (thrown). "
        "Intermediates 2035/2036 have no plant turnout (—)."
    )
    legend["A12"] = (
        "  Stock Signal Manager sentence: CLEAR if aligned {Main|Diverged} "
        "or set aspect {Stop|Approach}. Occupied blocks still force STOP. "
        "2-head top = Main/Stop; bottom = Diverged/Stop (unused disc stays red). "
        "Diverge dwarfs = Diverged/Stop. Balloon 2035/2036 = — / Approach."
    )

    ws = wb["By Signal"]
    headers = [c.value for c in ws[1]]
    stop_i = headers.index("Port STOP (R)") + 1
    stop_ou_i = headers.index("Port STOP OU") + 1
    app_i = headers.index("Port APPROACH (Y)") + 1
    app_ou_i = headers.index("Port APPROACH OU") + 1
    clr_i = headers.index("Port CLEAR (G)") + 1
    clr_ou_i = headers.index("Port CLEAR OU") + 1
    mast_i = headers.index("Mast") + 1
    head_i = headers.index("Head") + 1
    role_i = headers.index("Signal role") + 1

    if "Turnout polarity" not in headers:
        ws.insert_cols(role_i + 1, 2)
        ws.cell(1, role_i + 1, "Turnout polarity")
        ws.cell(1, role_i + 2, "OR aspect")
        ws.cell(1, role_i + 1).font = Font(bold=True)
        ws.cell(1, role_i + 2).font = Font(bold=True)
        headers = [c.value for c in ws[1]]
        role_i = headers.index("Signal role") + 1
        stop_i = headers.index("Port STOP (R)") + 1
        stop_ou_i = headers.index("Port STOP OU") + 1
        app_i = headers.index("Port APPROACH (Y)") + 1
        app_ou_i = headers.index("Port APPROACH OU") + 1
        clr_i = headers.index("Port CLEAR (G)") + 1
        clr_ou_i = headers.index("Port CLEAR OU") + 1
        mast_i = headers.index("Mast") + 1
        head_i = headers.index("Head") + 1
        role_i = headers.index("Signal role") + 1

    pol_i = headers.index("Turnout polarity") + 1
    or_i = headers.index("OR aspect") + 1

    missing = []
    for row in range(2, ws.max_row + 1):
        mast = ws.cell(row, mast_i).value
        head = ws.cell(row, head_i).value
        if not mast:
            continue
        key = (str(mast), str(head))
        h = heads.get(key)
        if not h:
            missing.append(key)
            continue
        ws.cell(row, stop_i, port_index(h["stop"]))
        ws.cell(row, stop_ou_i, h["stop"])
        ws.cell(row, app_i, port_index(h["approach"]))
        ws.cell(row, app_ou_i, h["approach"])
        ws.cell(row, clr_i, port_index(h["clear"]))
        ws.cell(row, clr_ou_i, h["clear"])
        lg = logic.get(key)
        if lg:
            ws.cell(row, pol_i, lg["turnout_polarity"])
            ws.cell(row, or_i, lg["or_aspect"])
        else:
            missing.append(key)
    if missing:
        raise SystemExit(f"missing head/logic rows: {missing}")
    wb.save(PLAN)
    print(f"wrote {PLAN.relative_to(ROOT)} ({len(heads)} heads)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
