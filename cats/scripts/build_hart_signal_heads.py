#!/usr/bin/env python3
"""Build HART Digicon→JMRI virtual signal heads (LCOS packed MQTT numbers).

MQTT packed node == RF24 radio Address. Enclosure ID is C{Address} (D{Address}
for the helix DCC client). Never assign heads to D5 (radio 5, DCC).

Each physical searchlight disc is one 3-pin LCOS object (STOP/APPROACH/CLEAR =
G/Y/R). Two-head masts are two objects (T + B) on the **same DNOU8** (6 pins)
so a module can sit next to its mast. A neighboring dwarf uses the leftover 2
pins; the 3rd dwarf pin spills to the adjacent cluster board when the plant
needs 9 pins. New 5V **OU4** on C1 / C4 / C13 (C12 and C11 use existing 5V OUs).

  radio 1  C1  Princess west (36 + 38). Packed 1xx.
  radio 2  C2  East End west overflow (24). Packed 2xx.
  radio 4  C4  Brick + Plane (motors 100–102/116 on OU1). Packed 4xx.
  radio 11 C11 Princess east (40) + balloon 2035/2036. Packed 11xx.
  radio 12 C12 East End 34 + turnout motors 107–112. Packed 12xx.
  radio 13 C13 Barn. Packed 13xx.
  radio 3  C3  Motors 103–106 only — no Digicon heads.
  radio 5  D5  DCC only — no signal ports

`--wiring-only` writes CSVs only (no tables.xml / publisher).
"""

from __future__ import annotations

import csv
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
CATS_MASTS = ROOT / "cats/resources/signals/cats-masts"
DATA = ROOT / "cats/data"

SKIP_HEAD: set[str] = set()
SKIP_MAST: set[str] = set()

# One 3-pin STOP/APPROACH/CLEAR object per physical disc. A 2-head mast owns
# 6 consecutive pins on one board (T then B). Packed = radio*100 + UID.
# disc: T top, B bottom, S single.
# mast, node, parent, location, packed, g, y, r, disc, v84_was
HEADS_3PIN: list[tuple[str, int, str, str, int, str, str, str, str, str]] = [
    # C4 radio 4 — Brick + Plane (motors 100–102, 116 on OU1). Packed 4xx.
    ("Mast 6LB", 4, "C4", "West - Lower (Plane)", 432, "C4-OU2-1", "C4-OU2-2", "C4-OU2-3", "T", "S3-6 G/Y/R"),
    ("Mast 6LB", 4, "C4", "West - Lower (Plane)", 433, "C4-OU2-4", "C4-OU2-5", "C4-OU2-6", "B", "with 6LB T on OU2"),
    ("Mast 6LA", 4, "C4", "West - Lower (Plane)", 434, "C4-OU2-7", "C4-OU3-8", "C4-OU3-7", "S", "G on Plane OU2; Y/R on 2L OU3 (OU2-8 is BS cal)"),
    ("Mast 2L", 4, "C4", "West - Lower (Brick east)", 438, "C4-OU3-1", "C4-OU3-2", "C4-OU3-3", "T", "S3-8 G/Y/R"),
    ("Mast 2L", 4, "C4", "West - Lower (Brick east)", 439, "C4-OU3-4", "C4-OU3-5", "C4-OU3-6", "B", "with 2L T on OU3"),
    ("Mast 4RA", 4, "C4", "West - Lower (Brick west)", 436, "C4-OU4-1", "C4-OU4-2", "C4-OU4-3", "S", "Brick west OU4 with 4RB"),
    ("Mast 4RB", 4, "C4", "West - Lower (Brick west)", 437, "C4-OU4-4", "C4-OU4-5", "C4-OU4-6", "S", "Brick west OU4 with 4RA"),
    # C13 radio 13 — Barn.
    ("Mast 8RA", 13, "C13", "West - Lower (Barn left)", 1332, "C13-OU1-1", "C13-OU1-2", "C13-OU1-3", "T", "S3-10 G/Y/R"),
    ("Mast 8RA", 13, "C13", "West - Lower (Barn left)", 1333, "C13-OU1-4", "C13-OU1-5", "C13-OU1-6", "B", "with 8RA T on OU1"),
    ("Mast 8RB", 13, "C13", "West - Lower (Barn left)", 1335, "C13-OU4-1", "C13-OU4-2", "C13-OU4-3", "T", "S3-12; sequential G/Y/R"),
    ("Mast 8RB", 13, "C13", "West - Lower (Barn left)", 1336, "C13-OU4-4", "C13-OU4-5", "C13-OU4-6", "B", "with 8RB T on OU4"),
    ("Mast 8LA", 13, "C13", "West - Lower (Barn right)", 1337, "C13-OU2-1", "C13-OU2-2", "C13-OU2-3", "T", "S3-5 G/Y/R"),
    ("Mast 8LA", 13, "C13", "West - Lower (Barn right)", 1338, "C13-OU2-4", "C13-OU2-5", "C13-OU2-6", "B", "with 8LA T on OU2"),
    ("Mast 8LB", 13, "C13", "West - Lower (Barn right)", 1334, "C13-OU2-7", "C13-OU2-8", "C13-OU4-7", "S", "G/Y with 8LA; R spills to 8RB OU4"),
    # C2 radio 2 — East End west overflow (24). C2 sits on the west end of East End.
    ("Mast 24RA", 2, "C2", "North - Lower (East End west 24)", 232, "C2-OU1-1", "C2-OU1-2", "C2-OU1-3", "T", "C2-OU1 5V"),
    ("Mast 24RA", 2, "C2", "North - Lower (East End west 24)", 238, "C2-OU1-4", "C2-OU1-5", "C2-OU1-6", "B", "with 24RA T on OU1"),
    ("Mast 24RB", 2, "C2", "North - Lower (East End west 24)", 234, "C2-OU1-7", "C2-OU3-8", "C2-OU2-7", "S", "G with 24RA; Y on OU3 (OU1-8 is BS cal); R with 24L"),
    ("Mast 24L", 2, "C2", "North - Lower (East End west 24)", 233, "C2-OU2-1", "C2-OU2-2", "C2-OU2-3", "T", "with 24RA cluster"),
    ("Mast 24L", 2, "C2", "North - Lower (East End west 24)", 239, "C2-OU2-4", "C2-OU2-5", "C2-OU2-6", "B", "with 24L T on OU2"),
    # C12 radio 12 — East End 34 + turnout motors 107–112. Packed 12xx.
    ("Mast 34L", 12, "C12", "North - Lower (East End 34)", 1232, "C12-OU2-1", "C12-OU2-2", "C12-OU2-3", "T", "was C2 235"),
    ("Mast 34L", 12, "C12", "North - Lower (East End 34)", 1233, "C12-OU2-4", "C12-OU2-5", "C12-OU2-6", "B", "with 34L T on OU2"),
    ("Mast 32R", 12, "C12", "North - Lower (East End 34)", 1234, "C12-OU2-7", "C12-OU3-8", "C12-OU3-7", "S", "G with 34L; Y/R on 34R OU3 (OU2-8 is BS cal)"),
    ("Mast 34R", 12, "C12", "North - Lower (East End 34)", 1235, "C12-OU3-1", "C12-OU3-2", "C12-OU3-3", "T", "was C2 237"),
    ("Mast 34R", 12, "C12", "North - Lower (East End 34)", 1236, "C12-OU3-4", "C12-OU3-5", "C12-OU3-6", "B", "with 34R T on OU3"),
    # C1 radio 1 — Princess west (36 @ SW35, 38 @ SW37). 40 and balloon on C11.
    ("Mast 36RA", 1, "C1", "Helix - Lower (Princess SW35)", 135, "C1-OU2-1", "C1-OU2-2", "C1-OU2-3", "T", "S1-2 G/Y/R"),
    ("Mast 36RA", 1, "C1", "Helix - Lower (Princess SW35)", 136, "C1-OU2-4", "C1-OU2-5", "C1-OU2-6", "B", "with 36RA T on OU2"),
    ("Mast 36RB", 1, "C1", "Helix - Lower (Princess SW35)", 132, "C1-OU3-1", "C1-OU3-2", "C1-OU3-3", "T", "was C11 1132"),
    ("Mast 36RB", 1, "C1", "Helix - Lower (Princess SW35)", 133, "C1-OU3-4", "C1-OU3-5", "C1-OU3-6", "B", "with 36RB T on OU3"),
    ("Mast 38LB", 1, "C1", "Helix - Lower (Princess SW37)", 139, "C1-OU4-1", "C1-OU4-2", "C1-OU4-3", "T", "S1-3 G/Y/R"),
    ("Mast 38LB", 1, "C1", "Helix - Lower (Princess SW37)", 140, "C1-OU4-4", "C1-OU4-5", "C1-OU4-6", "B", "with 38LB T on OU4"),
    ("Mast 38LA", 1, "C1", "Helix - Lower (Princess SW37)", 143, "C1-OU4-7", "C1-OU4-8", "C1-OU2-7", "S", "G/Y with 38LB; R spills to 36RA OU2"),
    # C11 radio 11 — Princess east (40 @ SW39) + balloon.
    ("Mast 40LB", 11, "C11", "Helix (Princess east SW39)", 1132, "C11-OU2-1", "C11-OU2-2", "C11-OU2-3", "T", "was C1 132"),
    ("Mast 40LB", 11, "C11", "Helix (Princess east SW39)", 1135, "C11-OU2-4", "C11-OU2-5", "C11-OU2-6", "B", "with 40LB T on OU2"),
    ("Mast 40LA", 11, "C11", "Helix (Princess east SW39)", 1136, "C11-OU2-7", "C11-OU3-8", "C11-OU3-7", "S", "G with 40LB; Y/R on balloon OU3 (OU2-8 is BS cal)"),
    ("Mast 2036", 11, "C11", "Helix (balloon 114)", 1133, "C11-OU3-1", "C11-OU3-2", "C11-OU3-3", "S", "S1-6; sequential G/Y/R"),
    ("Mast 2035", 11, "C11", "Helix (balloon 115)", 1134, "C11-OU3-4", "C11-OU3-5", "C11-OU3-6", "S", "S4-6; sequential G/Y/R"),
]

APPEAR = {1: "SL-1-low", 2: "SL-2-digicon", 3: "SL-3-high"}
SYSTEM_BY_HEADS = {1: "AAR-1946", 2: "hart-aar", 3: "AAR-1946"}
DISC_SORT = {"T": 0, "S": 1, "B": 2}

PIN_ORDER = (("G", "g_port"), ("Y", "y_port"), ("R", "r_port"))

# Pin 8 of the first 5V DNOU8 on every node that has block sensors.
# Occupancy-detector calibration current; never a lamp or relay.
BLOCK_CAL_PORTS = {
    "C1-OU2-8",
    "C2-OU1-8",
    "C3-OU2-8",
    "C4-OU2-8",
    "C11-OU2-8",
    "C12-OU2-8",
    "C13-OU1-8",
    "C14-OU2-8",
    "C21-OU2-8",
    "C22-OU2-8",
    "C23-OU2-8",
    "C24-OU2-8",
}


def packed(node: int, signal_index: int) -> int:
    return node * 100 + 32 + signal_index


def discs_of(mast: str, rows: list[dict]) -> list[dict]:
    by_packed: dict[int, dict] = {}
    for r in rows:
        if r["mast_user_name"] == mast:
            by_packed.setdefault(r["packed"], r)
    discs = list(by_packed.values())
    discs.sort(key=lambda r: DISC_SORT.get(r["disc_role"], 9))
    return discs


def build_rows() -> list[dict]:
    rows: list[dict] = []
    seen_ports: dict[str, str] = {}
    for mast, node, parent, loc, pid, g, y, r, disc, was in HEADS_3PIN:
        uid = pid - node * 100
        idx = uid - 32
        short = mast[5:] if mast.startswith("Mast ") else mast
        ports = {"G": g, "Y": y, "R": r}
        label = f"{short} {disc}" if disc in ("T", "B") else short
        for color, _key in PIN_ORDER:
            port = ports[color]
            if port in BLOCK_CAL_PORTS:
                raise SystemExit(
                    f"{port} is block-sensor calibration on the first 5V OU; "
                    f"cannot assign {mast} {disc} {color}"
                )
            if port in seen_ports:
                raise SystemExit(f"duplicate port {port}: {seen_ports[port]} and {mast}")
            seen_ports[port] = f"{mast} {disc} {color}"
            rows.append(
                {
                    "mqtt_node": node,
                    "parent_node_id": parent,
                    "board_location": loc,
                    "signal_index": idx,
                    "uid": uid,
                    "packed": pid,
                    "system_name": f"IH{pid}",
                    "user_name": f"Head {label} {color}",
                    "mast_user_name": mast,
                    "disc_role": disc,
                    "head_role": disc,
                    "lamp_color": color,
                    "port_id": port,
                    "g_port": g,
                    "y_port": y,
                    "r_port": r,
                    "v84_was": was,
                    "lcos_recipe": "STOP/APPROACH/CLEAR",
                    "topic": f"track/signalhead/{pid}",
                }
            )
    return rows


def mast_system_name(mast: str, rows: list[dict]) -> str:
    discs = discs_of(mast, rows)
    n = len(discs)
    ihs = "".join(f"({d['system_name']})" for d in discs)
    return f"IF$shsm:{SYSTEM_BY_HEADS[n]}:{APPEAR[n]}{ihs}"


def write_csvs(rows: list[dict]) -> None:
    DATA.mkdir(parents=True, exist_ok=True)
    wiring_fields = [
        "port_id",
        "parent_node_id",
        "mqtt_node",
        "board_location",
        "signal_index",
        "uid",
        "packed",
        "system_name",
        "user_name",
        "mast_user_name",
        "disc_role",
        "head_role",
        "lamp_color",
        "topic",
        "device_type",
        "voltage_rail",
        "notes",
    ]
    with (DATA / "signal_wiring.csv").open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=wiring_fields)
        w.writeheader()
        for r in rows:
            w.writerow(
                {
                    **{k: r[k] for k in wiring_fields if k in r},
                    "device_type": "Searchlight Signal Head",
                    "voltage_rail": "5V",
                    "notes": (
                        f"LCOS Signal {r['signal_index']} UID {r['uid']}; "
                        f"{r['lcos_recipe']}; lamp {r['lamp_color']}; "
                        f"was {r['v84_was']}"
                    ),
                }
            )

    plan_fields = [
        "mast_user_name",
        "mqtt_node",
        "parent_node_id",
        "heads",
        "appearance",
        "physignal",
        "mast_system_name",
        "head_system_names",
        "packed_ids",
        "port_ids",
    ]
    masts = []
    seen = set()
    for r in rows:
        m = r["mast_user_name"]
        if m in SKIP_MAST or m in seen:
            continue
        seen.add(m)
        hs = [x for x in rows if x["mast_user_name"] == m]
        discs = discs_of(m, rows)
        n = len(discs)
        masts.append(
            {
                "mast_user_name": m,
                "mqtt_node": r["mqtt_node"],
                "parent_node_id": r["parent_node_id"],
                "heads": n,
                "appearance": APPEAR[n],
                "physignal": "double" if n == 2 else "single",
                "mast_system_name": mast_system_name(m, rows),
                "head_system_names": " ".join(d["system_name"] for d in discs),
                "packed_ids": " ".join(str(d["packed"]) for d in discs),
                "port_ids": " ".join(x["port_id"] for x in hs),
            }
        )
    with (DATA / "signal_head_plan.csv").open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=plan_fields)
        w.writeheader()
        w.writerows(masts)

    # Enrich Digicon plan CSV with mqtt/packed/port columns (keep panel coords).
    plan_src = DATA / "signal_mast_plan.csv"
    if plan_src.exists():
        with plan_src.open(newline="") as f:
            old = list(csv.DictReader(f))
        by_name = {m["mast_user_name"]: m for m in masts}
        out_fields = list(old[0].keys())
        for col in (
            "heads",
            "mqtt_node",
            "parent_node_id",
            "packed_ids",
            "port_ids",
            "mast_system_name",
            "jmri_binding",
        ):
            if col not in out_fields:
                out_fields.append(col)
        enriched = []
        for row in old:
            name = row["proposed_mast_name"]
            if name in by_name:
                m = by_name[name]
                row = {
                    **row,
                    "heads": str(m["heads"]),
                    "mqtt_node": str(m["mqtt_node"]),
                    "parent_node_id": m["parent_node_id"],
                    "packed_ids": m["packed_ids"],
                    "port_ids": m["port_ids"],
                    "mast_system_name": m["mast_system_name"],
                    "jmri_binding": "virtual-heads+shsm",
                }
            enriched.append(row)
        with plan_src.open("w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=out_fields)
            w.writeheader()
            w.writerows(enriched)
        print(f"updated {plan_src}")

    print(f"wrote {DATA / 'signal_wiring.csv'} ({len(rows)} pin rows, {len(masts)} masts)")
    print(f"wrote {DATA / 'signal_head_plan.csv'} ({len(masts)} masts)")


def update_lcos_inventory(rows: list[dict]) -> None:
    """Rewrite Digicon signal ports on docs/wiring LCOS inventory DNOU8 sheet."""
    inv = ROOT / "docs/wiring/LCOS_Layout_Inventory_v85.xlsx"
    if not inv.exists():
        print(f"skip inventory update (missing {inv})")
        return
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("skip inventory update (openpyxl not installed)")
        return
    wb = load_workbook(inv)
    ws = wb["DNOU8"]
    # Map existing rows by Output Port ID (col A)
    by_port: dict[str, int] = {}
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        pid = row[0].value
        if pid:
            by_port[str(pid)] = i
    assigned = {r["port_id"] for r in rows}
    # Clear old dwarf labels on ports we are reclaiming (leave unrelated alone)
    for r in rows:
        port = r["port_id"]
        channel = int(port.rsplit("-", 1)[-1])
        vals = (
            port,
            r["parent_node_id"],
            r["board_location"],
            channel,
            r["user_name"],
            "Searchlight Signal Head",
            "5V",
            (
                f"HART Digicon; MQTT {r['topic']}; "
                f"packed {r['packed']} (node {r['mqtt_node']} sig {r['signal_index']})"
            ),
        )
        if port in by_port:
            rr = by_port[port]
            for col, v in enumerate(vals, start=1):
                ws.cell(rr, col, v)
        else:
            ws.append(vals)
    out = DATA / "LCOS_Layout_Inventory_v85_signal_ports.xlsx"
    wb.save(out)
    wb.save(inv)
    print(f"updated inventory DNOU8 → {inv} and {out}")


def write_cats_virtual_3() -> None:
    """3-head appearance from Digicon triple ASPECTMAP + AAR SL-3 imagelinks."""
    # Exact Digicon SIGNALTEMPLATE triple ASPECTMAP (HART_Master / West_Yard2).
    amap = {
        "R281": "green|red|red",
        "R281B": "green|red|red",
        "R281C": "green|red|red",
        "C412": "green|red|red",
        "R285": "yellow|red|red",
        "R281D": "yellow|red|red",
        "R282": "yellow|yellow|red",
        "R284": "yellow|yellow|red",
        "ADV_NORM": "yellow|yellow|red",
        "ADV_LIM": "yellow|yellow|red",
        "C413": "yellow|yellow|red",
        "C414": "yellow|yellow|red",
        "R283": "red|green|red",
        "R283A": "red|green|red",
        "R283B": "red|green|red",
        "C417": "red|green|red",
        "ADV_MED": "red|green|red",
        "R286": "red|yellow|red",
        "R287": "red|red|green",
        "C422": "red|red|green",
        "C423": "red|red|green",
        "C424": "red|red|green",
        "ADV_SLO": "red|red|green",
        "R288": "red|red|yellow",
        "RES_NORM": "red|red|red",
        "RES_LIM": "red|red|red",
        "RES_MED": "red|red|red",
        "RES_SLO": "red|red|red",
        "R291": "red|red|red",
        "R292": "red|red|red",
    }
    # Prefer matching rule gif when present
    icon = {
        "R281": "rule-281.gif",
        "R281B": "rule-281B.gif",
        "R281C": "rule-281C.gif",
        "R281D": "rule-285.gif",
        "R282": "rule-282.gif",
        "R284": "rule-284.gif",
        "R285": "rule-285.gif",
        "R283": "rule-283.gif",
        "R283A": "rule-283A.gif",
        "R283B": "rule-283B.gif",
        "R286": "rule-286.gif",
        "R287": "rule-287.gif",
        "R288": "rule-288.gif",
        "R291": "rule-292.gif",
        "R292": "rule-292.gif",
    }
    base = "../../../resources/icons/smallschematics/aspects/AAR-1946/SL-3-high/"
    lines = [
        '<?xml version="1.0" encoding="utf-8"?>',
        '<?xml-stylesheet href="../../XSLT/appearancetable.xsl" type="text/xsl"?>',
        "<!-- Three-head cats-virtual; Digicon triple ASPECTMAP + SL-3 imagelinks. -->",
        '<appearancetable xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"',
        '    xsi:noNamespaceSchemaLocation="http://jmri.org/xml/schema/appearancetable.xsd">',
        "  <aspecttable>CATS Virtual Signals</aspecttable>",
        "  <name>cats-virtual-3</name>",
        "  <reference>CATS Digicon triple-lamp PHYSIGNAL aspect keys</reference>",
        "  <description>Three searchlight heads driven by Digicon R/C/RES/ADV indication names</description>",
        "  <appearances>",
    ]
    for aspect, colors in amap.items():
        shows = "".join(f"<show>{c}</show>" for c in colors.split("|"))
        gif = icon.get(aspect, "rule-292.gif")
        lines.append(
            f"    <appearance><aspectname>{aspect}</aspectname>{shows}"
            f'<imagelink type="default">{base}{gif}</imagelink></appearance>'
        )
    stop = f"{base}rule-292.gif"
    lines += [
        "  </appearances>",
        "  <specificappearances>",
        "    <danger><aspect>R292</aspect></danger>",
        "    <held>",
        "      <aspect>R292</aspect>",
        f'      <imagelink type="default">{stop}</imagelink>',
        "    </held>",
        "    <dark>",
        "      <aspect>R292</aspect>",
        f'      <imagelink type="default">{stop}</imagelink>',
        "    </dark>",
        "    <permissive><aspect>R285</aspect></permissive>",
        "  </specificappearances>",
        "  <aspectMappings/>",
        "</appearancetable>",
        "",
    ]
    (CATS_MASTS / "appearance-cats-virtual-3.xml").write_text("\n".join(lines))
    # Ensure single-head file name element is cats-virtual (file stem match)
    single = CATS_MASTS / "appearance-cats-virtual.xml"
    t = single.read_text()
    t = re.sub(r"<name>Default</name>", "<name>cats-virtual</name>", t)
    single.write_text(t)
    # aspects.xml appearancefiles
    aspects = CATS_MASTS / "aspects.xml"
    at = aspects.read_text()
    if "appearance-cats-virtual-3.xml" not in at:
        at = at.replace(
            '    <appearancefile href="appearance-cats-virtual-2.xml" />\n  </appearancefiles>',
            '    <appearancefile href="appearance-cats-virtual-2.xml" />\n'
            '    <appearancefile href="appearance-cats-virtual-3.xml" />\n'
            "  </appearancefiles>",
        )
    if "appearance-cats-virtual-dwarf.xml" not in at:
        at = at.replace(
            '    <appearancefile href="appearance-cats-virtual-3.xml" />\n  </appearancefiles>',
            '    <appearancefile href="appearance-cats-virtual-3.xml" />\n'
            '    <appearancefile href="appearance-cats-virtual-dwarf.xml" />\n'
            "  </appearancefiles>",
        )
    aspects.write_text(at)
    print("wrote appearance-cats-virtual-3.xml; updated aspects + cats-virtual name")


def _heads_xml(rows: list[dict]) -> str:
    parts = [
        '  <signalheads class="jmri.managers.configurexml.AbstractSignalHeadManagerXml">',
    ]
    for r in rows:
        parts.append(
            '    <signalhead class="jmri.implementation.configurexml.VirtualSignalHeadXml">'
        )
        parts.append(f"      <systemName>{r['system_name']}</systemName>")
        parts.append(f"      <userName>{r['user_name']}</userName>")
        parts.append("    </signalhead>")
    parts.append("  </signalheads>")
    return "\n".join(parts)


def _masts_xml(rows: list[dict]) -> str:
    parts = [
        '  <signalmasts class="jmri.managers.configurexml.DefaultSignalMastManagerXml">',
    ]
    seen = set()
    for r in rows:
        m = r["mast_user_name"]
        if m in SKIP_MAST or m in seen:
            continue
        seen.add(m)
        sysname = mast_system_name(m, rows)
        parts.append(
            '    <signalmast class="jmri.implementation.configurexml.SignalHeadSignalMastXml">'
        )
        parts.append(f"      <systemName>{sysname}</systemName>")
        parts.append(f"      <userName>{m}</userName>")
        parts.append('      <unlit allowed="no" />')
        # hart-aar SL-2-digicon only defines displayable aspects; no disables needed.
        parts.append("    </signalmast>")
    parts.append("  </signalmasts>")
    return "\n".join(parts)


def patch_tables(rows: list[dict]) -> None:
    heads = _heads_xml(rows)
    masts = _masts_xml(rows)
    block = heads + "\n" + masts
    files = [
        ROOT / "tables/new_tables.xml",
        ROOT / "jmri/layouts/hart/output/tables.xml",
        ROOT / "jmri/layouts/hart/output/hart_prod.xml",
    ]
    # Match from <signalheads ...> through </signalmasts>
    pat = re.compile(
        r"  <signalheads class=\"jmri\.managers\.configurexml\.AbstractSignalHeadManagerXml\">.*?"
        r"  </signalmasts>",
        re.S,
    )
    for path in files:
        text = path.read_text()
        if not pat.search(text):
            print(f"SKIP (no signalheads/masts block): {path}", file=sys.stderr)
            continue
        # Preserve Layout Editor signalmasticons (see add_digicon_le_signal_icons.py).
        text2, n = pat.subn(block, text, count=1)
        path.write_text(text2)
        print(f"patched {path.relative_to(ROOT)} (masts block replaced, n={n})")


def write_publisher(rows: list[dict]) -> None:
    """Refresh HEAD_NAMES in mqtt_signalhead_publisher.py; do not rewrite the script."""
    names = [r["system_name"] for r in rows]
    path = ROOT / "jmri/scripts/mqtt_signalhead_publisher.py"
    names_lit = ",\n    ".join(repr(n) for n in names)
    begin = "# HEAD_NAMES_BEGIN"
    end_mark = "# HEAD_NAMES_END"
    block = (
        f"{begin}\nHEAD_NAMES = [\n    {names_lit},\n]\n{end_mark}"
    )
    src = path.read_text()
    pat = re.compile(re.escape(begin) + r".*?" + re.escape(end_mark), re.S)
    if not pat.search(src):
        raise SystemExit(f"{path}: missing {begin} / {end_mark} markers")
    path.write_text(pat.sub(block, src, count=1))
    print(f"updated HEAD_NAMES in {path.relative_to(ROOT)} ({len(names)} heads)")



def main() -> int:
    wiring_only = "--wiring-only" in sys.argv
    rows = build_rows()
    write_csvs(rows)
    if wiring_only:
        print("wiring-only: skipped inventory xlsx, tables.xml, publisher")
    else:
        update_lcos_inventory(rows)
        write_cats_virtual_3()
        patch_tables(rows)
        write_publisher(rows)
    by: dict[int, list] = {}
    for r in rows:
        by.setdefault(r["mqtt_node"], []).append(r)
    for node in sorted(by):
        pins = by[node]
        masts = {p["mast_user_name"] for p in pins}
        print(
            f"node {node} {pins[0]['parent_node_id']}: {len(masts)} heads / {len(pins)} pins "
            f"packed {sorted({p['packed'] for p in pins})}"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
