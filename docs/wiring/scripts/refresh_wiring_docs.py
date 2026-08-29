#!/usr/bin/env python3
"""Refresh LCOS inventory + Digicon as-built workbooks from hart CSVs.

Reads:
  cats/data/occupancy_bindings.csv
  cats/data/signal_wiring.csv
  cats/data/signal_head_plan.csv
  cats/data/signal_mast_plan.csv
  docs/wiring/imported/LCOS_Layout_Inventory_v84.xlsx
  docs/wiring/imported/signals_asbuilt_abs_v1.xlsx
  docs/wiring/imported/signals_split_v8.xlsx

Writes:
  docs/wiring/LCOS_Layout_Inventory_v85.xlsx
  docs/wiring/signals_asbuilt_abs_v2.xlsx
  docs/wiring/signals_split_v8.xlsx  (historical RGB plan + README sheet)
"""
from __future__ import annotations

import csv
import re
import shutil
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parents[3]
WIRING = ROOT / "docs/wiring"
IMPORTED = WIRING / "imported"
CATS = ROOT / "cats/data"
PUBLIC_MAP = ROOT / "jmri/layouts/hart/data/public_name_map.csv"

BLOCK_IN_NOTES = re.compile(r"Block\s+(\d+-\d+)")


def load_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="") as f:
        return list(csv.DictReader(f))


def load_proposed() -> dict[str, str]:
    """Live public string → proposed (device-map grammar)."""
    out: dict[str, str] = {}
    for row in load_csv(PUBLIC_MAP):
        current = (row.get("current") or "").strip()
        proposed = (row.get("proposed") or "").strip()
        if current and proposed and current != proposed:
            out[current] = proposed
    return out


_PROPOSED = load_proposed()


def proposed(name: str | None) -> str | None:
    if not name:
        return name
    mapped = _PROPOSED.get(name)
    if mapped:
        return mapped
    if " / " in name:
        return " / ".join(proposed(part.strip()) or part.strip() for part in name.split("/"))
    return name


def apply_proposed_to_wiring(rows: list[dict[str, str]]) -> None:
    for row in rows:
        if row.get("user_name"):
            row["user_name"] = proposed(row["user_name"]) or row["user_name"]
        if row.get("mast_user_name"):
            row["mast_user_name"] = proposed(row["mast_user_name"]) or row["mast_user_name"]


def occupancy_by_hw() -> dict[str, str]:
    """MQTT occupancy userName (Block n-n) → public block name."""
    groups: dict[str, list[str]] = defaultdict(list)
    for row in load_csv(CATS / "occupancy_bindings.csv"):
        hw = (row.get("occupancy_sensor_user_name") or "").strip()
        name = (row.get("block_user_name") or "").strip()
        if hw and name and name not in groups[hw]:
            groups[hw].append(name)

    picked: dict[str, str] = {}
    for hw, names in groups.items():
        raw_os = [n for n in names if n.startswith("OS ")]
        raw_other = [n for n in names if n not in raw_os]
        os_names = [proposed(n) or n for n in raw_os]
        other = [proposed(n) or n for n in raw_other]
        if os_names and other:
            picked[hw] = f"{os_names[0]} / {other[0]}"
        else:
            picked[hw] = os_names[0] if os_names else other[0]
    return picked


def replace_in_cell(value, replacements: list[tuple[str, str]]):
    if not isinstance(value, str) or not value:
        return value
    out = value
    for old, new in replacements:
        if old in out:
            out = out.replace(old, new)
    return out


def walk_replace(ws: Worksheet, replacements: list[tuple[str, str]]) -> int:
    n = 0
    for row in ws.iter_rows():
        for cell in row:
            new = replace_in_cell(cell.value, replacements)
            if new != cell.value:
                cell.value = new
                n += 1
    return n


NAME_REPLACEMENTS: list[tuple[str, str]] = [
    # Long JMRI mast userNames → Digicon (longest first)
    ("West Yard West East Main Ext", proposed("117RB") or "117RB"),
    ("West Yard East OS 117b", proposed("117LA") or "117LA"),
    ("West Yard East Yard T6", proposed("117LB") or "117LB"),
    ("West Yard West OS 117", proposed("117RA") or "117RA"),
    ("East End West Yard Track 1", proposed("111RB") or "111RB"),
    ("East End West Main West", proposed("111RA") or "111RA"),
    ("East End South OS 112", proposed("112R") or "112R"),
    ("East End South OS 110", proposed("110R") or "110R"),
    ("East End East OS 111a", proposed("111L") or "111L"),
    ("East End East Lead", proposed("112L") or "112L"),
    ("Princess North McKees Rocks", proposed("115LB") or "115LB"),
    ("Princess South McKeesport", proposed("114LB") or "114LB"),
    ("Princess West OS 113a", proposed("113RB") or "113RB"),
    ("Princess West OS 113b", proposed("113RA") or "113RA"),
    ("Princess East McKeesport", proposed("120R") or "120R"),
    ("Princess East McKees Rocks", proposed("120L") or "120L"),
    ("Princess East K-1", proposed("115LA") or "115LA"),
    ("Princess East K-2", proposed("114LA") or "114LA"),
    ("Plane East East Main Ext", proposed("102LB") or "102LB"),
    ("Plane East OS 102", proposed("102LA") or "102LA"),
    ("Brick East Main West", proposed("100L") or "100L"),
    ("Brick West Yard 1", proposed("101RA") or "101RA"),
    ("Brick West Yard 2", proposed("101RB") or "101RB"),
    ("Brick W-1 West Stub", proposed("101LA") or "101LA"),
    ("Brick W-2 West Stub", proposed("101LB") or "101LB"),
    # OS public names (ADR-005: CP lives in comment, not the name)
    ("OS 117b (West Yard)", proposed("OS 117b") or "OS 117b"),
    ("OS 119 (West Yard)", proposed("OS 119") or "OS 119"),
    ("OS 118 (West Yard)", proposed("OS 118") or "OS 118"),
    ("OS 117 (West Yard)", proposed("OS 117") or "OS 117"),
    ("OS 116 (West Yard)", proposed("OS 116") or "OS 116"),
    ("OS 115 (Princess)", proposed("OS 115") or "OS 115"),
    ("OS 114 (Princess)", proposed("OS 114") or "OS 114"),
    ("OS 113b (Princess)", proposed("OS 113b") or "OS 113b"),
    ("OS 113a (Princess)", proposed("OS 113a") or "OS 113a"),
    ("OS 112 (East End)", proposed("OS 112") or "OS 112"),
    ("OS 111b (East End)", proposed("OS 111b") or "OS 111b"),
    ("OS 111a (East End)", proposed("OS 111a") or "OS 111a"),
    ("OS 110 (East End)", proposed("OS 110") or "OS 110"),
    ("OS 109 (East End)", proposed("OS 109") or "OS 109"),
    ("OS 108 (East End)", proposed("OS 108") or "OS 108"),
    ("OS 107 (East End)", proposed("OS 107") or "OS 107"),
    ("OS 106 (South Yard)", proposed("OS 106") or "OS 106"),
    ("OS 105 (South Yard)", proposed("OS 105") or "OS 105"),
    ("OS 104 (South Yard)", proposed("OS 104") or "OS 104"),
    ("OS 103 (South Yard)", proposed("OS 103") or "OS 103"),
    ("OS 102 (Plane)", proposed("OS 102") or "OS 102"),
    ("OS 101 (Brick)", proposed("OS 101") or "OS 101"),
    ("OS 100 (Brick)", proposed("OS 100") or "OS 100"),
    ("Main West Brick–Plane", proposed("Brick-Plane") or "Brick-Plane"),
    ("Main West Brick-Plane", proposed("Brick-Plane") or "Brick-Plane"),
    # Yard plates (longest first so T11 before T1)
    ("Yard Track 5", proposed("S-5") or "S-5"),
    ("Yard Track 4", proposed("S-4") or "S-4"),
    ("Yard Track 3", proposed("S-3") or "S-3"),
    ("Yard Track 2", proposed("S-2") or "S-2"),
    ("Yard Track 1", proposed("S-1") or "S-1"),
    ("West Yard Track 2", proposed("W-2") or "W-2"),
    ("West Yard Track 1", proposed("W-1") or "W-1"),
    ("West Yard 2", proposed("W-2") or "W-2"),
    ("West Yard 1", proposed("W-1") or "W-1"),
    ("Yard T6", proposed("Barn") or "Barn"),
    ("Yard T1", proposed("Scale") or "Scale"),
    ("track/signalmast/464", "track/signalhead/IH438 / IH439"),
]


# v84 sequential enclosure IDs → C{radio Address} (D{Address} for helix DCC).
# Imported v84 snapshot keeps the old labels; refresh remaps every cell.
LEGACY_NODE_IDS: dict[str, str] = {
    "C1": "C1",
    "C2": "C12",
    "C3": "C2",
    "C4": "C3",
    "C5": "C13",
    "C6": "C4",
    "C7": "C11",
    "C8": "C21",
    "C9": "C22",
    "C10": "C32",
    "C11": "C23",
    "C12": "C14",
    "C13": "C24",
    "D1": "D5",
}
NODE_ID_TOKEN = re.compile(r"(?<![A-Za-z0-9])([CD]\d+)(?![0-9])")
DISC_SORT = {"T": 0, "S": 1, "B": 2}

# v84 "renamed for linear6" concatenated radio/100 onto upper-deck SW1xx plant
# IDs (SW127 → Switch 1127, SW138 → Switch 10038). Restore SW*. Do not write
# "Switch 127" — that is CTC East End Switch 27.
MANGLED_SWITCH_FIXES: list[tuple[str, str]] = [
    ("Switch 1125,Switch 1126,139,Switch 10040", "SW125, SW126, SW139, SW140"),
    ("Switch 1121,Switch 1123", "SW121, SW123"),
    ("SW,120,Switch 1122", "SW120, SW122"),
    ("Switch 10041,Switch 10042", "SW141, SW142"),
    ("Switch 10050", "SW150"),
    ("Switch 10049", "SW149"),
    ("Switch 10048", "SW148"),
    ("Switch 10047", "SW147"),
    ("Switch 10046", "SW146"),
    ("Switch 10045", "SW145"),
    ("Switch 10044", "SW144"),
    ("Switch 10043", "SW143"),
    ("Switch 10042", "SW142"),
    ("Switch 10041", "SW141"),
    ("Switch 10040", "SW140"),
    ("Switch 10038", "SW138"),
    ("Switch 1129", "SW129"),
    ("Switch 1127", "SW127"),
    ("Switch 1126", "SW126"),
    ("Switch 1125", "SW125"),
    ("Switch 1124", "SW124"),
    ("Switch 1123", "SW123"),
    ("Switch 1122", "SW122"),
    ("Switch 1121", "SW121"),
]

# v84 DCC Switch 100–119 → live CTC numbers (ADR-005). Longest first so
# Switch 113a becomes Switch 35a. Do not emit "Switch 127" (that is SW127).
DCC_SWITCH_TO_CTC: list[tuple[str, str]] = [
    ("Switch 119", "Switch 9"),
    ("Switch 118", "Switch 11"),
    ("Switch 117", "Switch 7"),
    ("Switch 116", "Switch 13"),
    ("Switch 115", "Switch 39"),
    ("Switch 114", "Switch 37"),
    ("Switch 113", "Switch 35"),
    ("Switch 112", "Switch 33"),
    ("Switch 111", "Switch 23"),
    ("Switch 110", "Switch 31"),
    ("Switch 109", "Switch 29"),
    ("Switch 108", "Switch 27"),
    ("Switch 107", "Switch 25"),
    ("Switch 106", "Switch 21"),
    ("Switch 105", "Switch 19"),
    ("Switch 104", "Switch 17"),
    ("Switch 103", "Switch 15"),
    ("Switch 102", "Switch 5"),
    ("Switch 101", "Switch 3"),
    ("Switch 100", "Switch 1"),
]

# Pre-linear6 SW1–SW18 / SCX* tokens still on the Nodes summary columns.
# Digit-bounded so SW12 does not eat SW124.
_SW_TOKEN_CTC: list[tuple[re.Pattern[str], str]] = [
    (re.compile(r"(?<![A-Za-z0-9])SCXB(?![A-Za-z0-9])"), "Switch 35"),
    (re.compile(r"(?<![A-Za-z0-9])SCXA(?![A-Za-z0-9])"), "Switch 23"),
    (re.compile(r"(?<![A-Za-z0-9])SW18(?![0-9])"), "Switch 39"),
    (re.compile(r"(?<![A-Za-z0-9])SW17(?![0-9])"), "Switch 37"),
    (re.compile(r"(?<![A-Za-z0-9])SW13(?![0-9])"), "Switch 13"),
    (re.compile(r"(?<![A-Za-z0-9])SW12(?![0-9])"), "Switch 23"),
    (re.compile(r"(?<![A-Za-z0-9])SW11(?![0-9])"), "Switch 31"),
    (re.compile(r"(?<![A-Za-z0-9])SW10(?![0-9])"), "Switch 29"),
    (re.compile(r"(?<![A-Za-z0-9])SW9(?![0-9])"), "Switch 27"),
    (re.compile(r"(?<![A-Za-z0-9])SW8(?![0-9])"), "Switch 25"),
    (re.compile(r"(?<![A-Za-z0-9])SW7(?![0-9])"), "Switch 21"),
    (re.compile(r"(?<![A-Za-z0-9])SW6(?![0-9])"), "Switch 19"),
    (re.compile(r"(?<![A-Za-z0-9])SW5(?![0-9])"), "Switch 17"),
    (re.compile(r"(?<![A-Za-z0-9])SW4(?![0-9])"), "Switch 15"),
    (re.compile(r"(?<![A-Za-z0-9])SW3(?![0-9])"), "Switch 5"),
    (re.compile(r"(?<![A-Za-z0-9])SW2(?![0-9])"), "Switch 3"),
    (re.compile(r"(?<![A-Za-z0-9])SW1(?![0-9])"), "Switch 1"),
]
HEAD_MAST_RE = re.compile(r"^Head (\S+)")
LED_NAME_RE = re.compile(r"^S\d+-\d+")
RGB_LED_RE = re.compile(r"^(S([4-6])-(\d+))\s+([GYR])$")

# Upper deck continues the lower-deck odd-switch / even-signal pattern from 61/62.
# CP4 (helix) first, then CP5 (north), then CP6 (peninsula/west). v8 column order
# inside each CP. Compound plants keep a/b/c/d on the physical machines.
UPPER_SWITCH_NUM: dict[str, str] = {
    "NIX": "Switch 61",
    "SW127": "Switch 63",
    "SW138": "Switch 65",
    "SW129": "Switch 67",
    "DJE": "Switch 69",
    "DJW": "Switch 71",
    "SW124": "Switch 73",
    "CBX": "Switch 75",
    "SW143": "Switch 77",
    "SW144": "Switch 79",
    "SW145": "Switch 81",
    "SW146": "Switch 83",
    "SW147": "Switch 85",
    "SW148": "Switch 87",
    "SW149": "Switch 89",
    "SW150": "Switch 91",
    "SW125": "Switch 61a",
    "SW126": "Switch 61b",
    "SW139": "Switch 61c",
    "SW140": "Switch 61d",
    "SW121": "Switch 69a",
    "SW123": "Switch 69b",
    "SW120": "Switch 71a",
    "SW122": "Switch 71b",
    "SW141": "Switch 75a",
    "SW142": "Switch 75b",
}

# Planned RGB head → even mast id (switch+1, L/R/A/B like 24L / 36RA).
UPPER_HEAD_NAME: dict[str, str] = {
    "S4-6": "62L",
    "S4-3": "64L",
    "S4-1": "64R",
    "S4-4": "66RA",
    "S4-5": "66RB",
    "S4-2": "68L",
    "S4-7": "68R",
    "S5-1": "70L",
    "S5-5": "70RA",
    "S5-6": "70RB",
    "S5-2": "72L",
    "S5-3": "74L",
    "S5-4": "74RA",
    "S5-7": "74RB",
    "S6-1": "76L",
    "S6-2": "76R",
    "S6-3": "78L",
    "S6-6": "78R",
    "S6-4": "78RA",
    "S6-5": "78RB",
    "S6-10": "82L",
    "S6-11": "84RA",
    "S6-12": "84RB",
    "S6-13": "86L",
    "S6-7": "88L",
    "S6-8": "88R",
    "S6-9": "88RA",
    "S6-14": "90L",
    "S6-15": "92L",
}

# v8 heads that are clearly extra discs on an existing plant (not a new mast).
V8_MAST_HINT: dict[str, str] = {
    "S4-4": "Switch 65 (with 66RB)",
    "S4-5": "Switch 65 (with 66RA)",
    "S5-6": "Switch 69 reverse",
    "S5-7": "Switch 73 reverse",
}

UPPER_TURNOUT_CP: dict[str, str] = {
    "SW127": "CP4 (Helix - Upper)",
    "SW129": "CP4 (Helix - Upper)",
    "SW138": "CP4 (Helix - Upper)",
    "NIX": "CP4 (Helix - Upper)",
    "SW124": "CP5 (North - Upper)",
    "DJE": "CP5 (North - Upper)",
    "DJW": "CP5 (North - Upper)",
    "CBX": "CP6 (Peninsula - Upper)",
    "SW143": "CP6 (West - Upper)",
    "SW144": "CP6 (West - Upper)",
    "SW145": "CP6 (West - Upper)",
    "SW146": "CP6 (West - Upper)",
    "SW147": "CP6 (West - Upper)",
    "SW148": "CP6 (Peninsula - Upper)",
    "SW149": "CP6 (Peninsula - Upper)",
    "SW150": "CP6 (Peninsula - Upper)",
}


def _build_upper_name_tokens() -> list[tuple[re.Pattern[str], str]]:
    """Longest-first regexes: S6-15 before S6-1, SW150 before SW124, NIX last."""
    out: list[tuple[re.Pattern[str], str]] = []
    for old, new in sorted(
        UPPER_HEAD_NAME.items(), key=lambda kv: (-len(kv[0]), kv[0])
    ):
        out.append(
            (re.compile(rf"(?<!was )(?<![A-Za-z0-9]){re.escape(old)}(?![0-9])"), new)
        )
    for old, new in sorted(
        UPPER_SWITCH_NUM.items(), key=lambda kv: (-len(kv[0]), kv[0])
    ):
        out.append(
            (re.compile(rf"(?<![A-Za-z0-9]){re.escape(old)}(?![A-Za-z0-9])"), new)
        )
    return out


_UPPER_NAME_TOKENS = _build_upper_name_tokens()

RETIRED_S3_TURNOUTS = ("Switch 15", "Switch 17", "Switch 19", "Switch 21")
RETIRED_S2_TURNOUTS = ("Switch 25", "Switch 27", "Switch 29")
TURNOUT_SIGNAL_COLS = (
    "Entry Signal",
    "Entry Signal R Port",
    "Entry Signal Y Port",
    "Entry Signal G Port",
    "Normal Exit Signal",
    "Normal Exit R Port",
    "Normal Exit Y Port",
    "Normal Exit G Port",
    "Reverse Exit Signal",
    "Reverse Exit R Port",
    "Reverse Exit Y Port",
    "Reverse Exit G Port",
)


def rewrite_legacy_node_ids(value):
    """Rewrite C12/C3/… tokens; longest match is the full number (C12 ≠ C1)."""
    if not isinstance(value, str) or not value:
        return value

    def repl(match: re.Match[str]) -> str:
        tok = match.group(1)
        return LEGACY_NODE_IDS.get(tok, tok)

    return NODE_ID_TOKEN.sub(repl, value)


def fix_mangled_switch_names(wb) -> int:
    """Undo v84 10038/1127 concatenations on every sheet."""
    n = 0
    for ws in wb.worksheets:
        n += walk_replace(ws, MANGLED_SWITCH_FIXES)
    return n


def apply_ctc_switch_names(wb) -> int:
    """Switch 113 → Switch 35 (and the rest of DCC 100–119)."""
    n = 0
    for ws in wb.worksheets:
        n += walk_replace(ws, DCC_SWITCH_TO_CTC)
    return n


def apply_legacy_sw_tokens(wb) -> int:
    """SCXB / SW17 / … on Nodes summaries (digit-bounded)."""
    n = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str) or not cell.value:
                    continue
                out = cell.value
                for pat, repl in _SW_TOKEN_CTC:
                    out = pat.sub(repl, out)
                if out != cell.value:
                    cell.value = out
                    n += 1
    return n


def apply_upper_deck_names(wb) -> int:
    """SW127/NIX/S4-1 → Switch 63 / Switch 61 / 64R (upper CTC 61+)."""
    n = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str) or not cell.value:
                    continue
                out = cell.value
                for pat, repl in _UPPER_NAME_TOKENS:
                    out = pat.sub(repl, out)
                if out != cell.value:
                    cell.value = out
                    n += 1
    return n


def _unique_keep_order(items: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for item in items:
        if item and item not in seen:
            seen.add(item)
            out.append(item)
    return out


def summarize_node_labels(nodes: Worksheet, dnou: Worksheet, dnin: Worksheet) -> None:
    """Rebuild Turnout / button / FB / signal summary columns from live ports."""
    motors: dict[str, list[str]] = defaultdict(list)
    signals: dict[str, list[str]] = defaultdict(list)
    buttons: dict[str, list[str]] = defaultdict(list)
    fbs: dict[str, list[str]] = defaultdict(list)
    for row in range(2, dnou.max_row + 1):
        nid = dnou.cell(row, 2).value
        dev = dnou.cell(row, 5).value
        dtype = str(dnou.cell(row, 6).value or "")
        if not nid or not isinstance(dev, str):
            continue
        nid = str(nid)
        if dtype == "Turnout Motor":
            name = re.sub(r" [NR]$", "", dev).strip()
            motors[nid].append(name)
        elif dtype == "Searchlight Signal Head":
            m = HEAD_MAST_RE.match(dev)
            signals[nid].append(m.group(1) if m else dev)
        elif dtype in ("Signal LED", "Dwarf Upgrade"):
            m = HEAD_MAST_RE.match(dev)
            if m:
                signals[nid].append(m.group(1))
            elif LED_NAME_RE.match(dev.split()[0] if dev else ""):
                signals[nid].append(dev.split()[0])
    for row in range(2, dnin.max_row + 1):
        nid = dnin.cell(row, 2).value
        dev = dnin.cell(row, 5).value
        if not nid or not isinstance(dev, str):
            continue
        nid = str(nid)
        if dev.endswith(" BTN"):
            buttons[nid].append(dev)
        elif " FB" in dev:
            fbs[nid].append(dev)
    idx = header_index(nodes)
    id_col = idx["Node ID"]
    for row in range(2, nodes.max_row + 1):
        nid = nodes.cell(row, id_col).value
        if not nid:
            continue
        nid = str(nid)
        if "Turnout Names" in idx:
            names = _unique_keep_order(motors.get(nid, []))
            nodes.cell(row, idx["Turnout Names"]).value = ", ".join(names) or None
        if "Button Devices" in idx:
            names = _unique_keep_order(buttons.get(nid, []))
            nodes.cell(row, idx["Button Devices"]).value = ", ".join(names) or None
        if "Turnout Feedback Devices" in idx:
            names = _unique_keep_order(fbs.get(nid, []))
            nodes.cell(row, idx["Turnout Feedback Devices"]).value = ", ".join(names) or None
        if "Signal Names" in idx:
            names = _unique_keep_order(signals.get(nid, []))
            nodes.cell(row, idx["Signal Names"]).value = ", ".join(names) or None


def remap_workbook_node_ids(wb) -> dict[int, str]:
    """Rewrite every C#/D1 token. Returns Nodes-sheet row → legacy Node ID."""
    nodes = wb["Nodes"]
    idx = header_index(nodes)
    id_col = idx["Node ID"]
    legacy_by_row: dict[int, str] = {}
    for row in range(2, nodes.max_row + 1):
        nid = nodes.cell(row, id_col).value
        if nid:
            legacy_by_row[row] = str(nid)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                new = rewrite_legacy_node_ids(cell.value)
                if new != cell.value:
                    cell.value = new
    return legacy_by_row


def add_legacy_node_column(ws: Worksheet, legacy_by_row: dict[int, str]) -> None:
    if "Legacy Node ID" in header_index(ws):
        return
    ws.insert_cols(2)
    ws.cell(1, 2).value = "Legacy Node ID"
    ws.cell(1, 2).font = Font(bold=True)
    for row, old in legacy_by_row.items():
        ws.cell(row, 2).value = old


def sort_nodes_by_address(ws: Worksheet) -> None:
    idx = header_index(ws)
    addr_col = idx["Address"]
    rows: list[tuple[int, list]] = []
    max_col = ws.max_column
    for row in range(2, ws.max_row + 1):
        vals = [ws.cell(row, c).value for c in range(1, max_col + 1)]
        if not any(v is not None for v in vals):
            continue
        addr = ws.cell(row, addr_col).value
        try:
            key = int(addr)
        except (TypeError, ValueError):
            key = 999
        rows.append((key, vals))
    rows.sort(key=lambda item: item[0])
    for i, (_, vals) in enumerate(rows, start=2):
        for c, v in enumerate(vals, start=1):
            ws.cell(i, c).value = v
    extra = ws.max_row - (1 + len(rows))
    if extra > 0:
        ws.delete_rows(2 + len(rows), extra)


# TurnoutSummary: one 3-pin head per face. R/Y/G columns are lamp colors again.
# Ports use C{radio Address} (Plane/Brick C4, East End 24 on C2 / 34 on C12).
TURNOUT_DIGICON: dict[str, dict[str, object]] = {
    "Switch 100": {
        "entry": "2L",
        "entry_ports": ("C4-OU3-3", "C4-OU3-2", "C4-OU3-1"),
        "normal": None,
        "reverse": None,
    },
    "Switch 101": {
        "entry": None,
        "normal": "4RA",
        "normal_ports": ("C4-OU4-3", "C4-OU4-2", "C4-OU4-1"),
        "reverse": "4RB",
        "reverse_ports": ("C4-OU4-6", "C4-OU4-5", "C4-OU4-4"),
    },
    "Switch 102": {
        "entry": "6LA",
        "entry_ports": ("C4-OU3-7", "C4-OU3-8", "C4-OU2-7"),
        "normal": None,
        "reverse": "6LB",
        "reverse_ports": ("C4-OU2-3", "C4-OU2-2", "C4-OU2-1"),
    },
    "Switch 111": {
        "entry": "24L",
        "entry_ports": ("C2-OU2-3", "C2-OU2-2", "C2-OU2-1"),
        "normal": "24RA",
        "normal_ports": ("C2-OU1-3", "C2-OU1-2", "C2-OU1-1"),
        "reverse": "24RB",
        "reverse_ports": ("C2-OU3-8", "C2-OU3-7", "C2-OU3-6"),
    },
    "Switch 110": {
        "entry": "32R",
        "entry_ports": ("C12-OU3-7", "C12-OU3-8", "C12-OU2-7"),
        "normal": None,
        "reverse": None,
    },
    "Switch 112": {
        "entry": "34L",
        "entry_ports": ("C12-OU2-3", "C12-OU2-2", "C12-OU2-1"),
        "normal": None,
        "reverse": "34R",
        "reverse_ports": ("C12-OU3-3", "C12-OU3-2", "C12-OU3-1"),
    },
    "Switch 113": {
        "entry": "36RA",
        "entry_ports": ("C1-OU2-3", "C1-OU2-2", "C1-OU2-1"),
        "normal": None,
        "reverse": "36RB",
        "reverse_ports": ("C1-OU3-3", "C1-OU3-2", "C1-OU3-1"),
    },
    "Switch 114": {
        "entry": "2036",
        "entry_ports": ("C11-OU3-6", "C11-OU3-5", "C11-OU3-4"),
        "normal": "38LB",
        "normal_ports": ("C1-OU4-3", "C1-OU4-2", "C1-OU4-1"),
        "reverse": "38LA",
        "reverse_ports": ("C1-OU2-7", "C1-OU3-8", "C1-OU3-7"),
    },
    "Switch 115": {
        "entry": "2035",
        "entry_ports": ("C11-OU2-7", "C11-OU3-8", "C11-OU3-7"),
        "normal": "40LB",
        "normal_ports": ("C11-OU2-3", "C11-OU2-2", "C11-OU2-1"),
        "reverse": "40LA",
        "reverse_ports": ("C11-OU3-3", "C11-OU3-2", "C11-OU3-1"),
    },
    "Switch 117": {
        "entry": "8RA",
        "entry_ports": ("C13-OU1-3", "C13-OU1-2", "C13-OU1-1"),
        "normal": "8LA",
        "normal_ports": ("C13-OU2-3", "C13-OU2-2", "C13-OU2-1"),
        "reverse": "8LB / 8RB",
        "reverse_ports": ("C13-OU4-7", "C13-OU2-8", "C13-OU2-7"),
    },
}


def header_index(ws: Worksheet) -> dict[str, int]:
    return {str(c.value): i for i, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=1) if c.value}


def refresh_block_sensors(ws: Worksheet, occ: dict[str, str]) -> list[str]:
    log: list[str] = []
    idx = header_index(ws)
    name_col = idx["Block Section Name"]
    notes_col = idx["Notes"]
    port_col = idx["Port ID"]
    for row in range(2, ws.max_row + 1):
        notes = ws.cell(row, notes_col).value
        if not isinstance(notes, str):
            continue
        m = BLOCK_IN_NOTES.search(notes)
        if not m:
            continue
        hw = f"Block {m.group(1)}"
        public = occ.get(hw)
        if not public:
            continue
        old = ws.cell(row, name_col).value
        if old != public:
            ws.cell(row, name_col).value = public
            log.append(f"{ws.cell(row, port_col).value}: {old!r} → {public!r} ({hw})")
    return log


# New 5V DNOU8: C1 Princess 38, C4 Brick west 4R, C13 Barn 8RB.
# C2 west-24 and C11/C12 use existing 5V boards (no new OU4).
NEW_5V_OU4 = ("C1", "C4", "C13")


def ensure_ou4_boards(ws: Worksheet) -> list[str]:
    """Append empty C*-OU4-1..8 5V rows so overlay / the schematic see the new boards."""
    log: list[str] = []
    have: set[str] = set()
    loc_by_node: dict[str, str] = {}
    for row in range(2, ws.max_row + 1):
        pid = ws.cell(row, 1).value
        parent = ws.cell(row, 2).value
        loc = ws.cell(row, 3).value
        if pid:
            have.add(str(pid))
        if parent and loc and str(parent) not in loc_by_node:
            loc_by_node[str(parent)] = str(loc)
    for nid in NEW_5V_OU4:
        loc = loc_by_node.get(nid, "")
        for ch in range(1, 9):
            port = f"{nid}-OU4-{ch}"
            if port in have:
                continue
            ws.append(
                (
                    port,
                    nid,
                    loc,
                    ch,
                    None,
                    None,
                    "5V",
                    "NEW 5V DNOU8 (OU4) for second searchlight disc",
                )
            )
            log.append(f"add {port}")
    return log


def overlay_dnou8(ws: Worksheet, wiring: list[dict[str, str]]) -> list[str]:
    log: list[str] = []
    by_port: dict[str, int] = {}
    for row in range(2, ws.max_row + 1):
        pid = ws.cell(row, 1).value
        if pid:
            by_port[str(pid)] = row
    for r in wiring:
        port = r["port_id"]
        channel = int(port.rsplit("-", 1)[-1])
        prev = None
        if port in by_port:
            prev = ws.cell(by_port[port], 5).value
        note = (
            f"HART Digicon {r.get('lcos_recipe', 'STOP/APPROACH/CLEAR')}; "
            f"lamp {r.get('lamp_color') or r.get('head_role')}; "
            f"MQTT {r['topic']}; packed {r['packed']} "
            f"(node {r['mqtt_node']} sig {r['signal_index']})"
        )
        if prev and str(prev) != r["user_name"]:
            note += f"; was {prev}"
        vals = (
            port,
            r["parent_node_id"],
            r["board_location"],
            channel,
            r["user_name"],
            "Searchlight Signal Head",
            "5V",
            note,
        )
        if port in by_port:
            rr = by_port[port]
            for col, v in enumerate(vals, start=1):
                ws.cell(rr, col, v)
            log.append(f"overlay {port}: {prev!r} → {r['user_name']!r}")
        else:
            ws.append(vals)
            log.append(f"append {port}: {r['user_name']}")
    return log


def first_5v_cal_ports(dnou_ws: Worksheet, bs_ws: Worksheet) -> dict[str, str]:
    """Pin 8 of the lowest-numbered 5V OU on each node that has block sensors."""
    nodes: set[str] = set()
    nidx = header_index(bs_ws)
    ncol = nidx.get("Node ID")
    if ncol:
        for row in range(2, bs_ws.max_row + 1):
            nid = bs_ws.cell(row, ncol).value
            if nid:
                nodes.add(str(nid))
    fives: dict[str, set[int]] = defaultdict(set)
    for row in range(2, dnou_ws.max_row + 1):
        port = dnou_ws.cell(row, 1).value
        parent = dnou_ws.cell(row, 2).value
        rail = dnou_ws.cell(row, 7).value
        if not port or str(parent) not in nodes or str(rail) != "5V":
            continue
        parts = str(port).split("-")
        if len(parts) < 2 or not parts[1].startswith("OU"):
            continue
        try:
            fives[str(parent)].add(int(parts[1][2:]))
        except ValueError:
            continue
    return {nid: f"{nid}-OU{min(ous)}-8" for nid, ous in fives.items() if ous}


def stamp_block_sensor_cal(
    ws: Worksheet, cal_ports: dict[str, str]
) -> list[str]:
    """Reserve first-5V-OU pin 8 for occupancy detector calibration current."""
    log: list[str] = []
    by_port: dict[str, int] = {}
    loc_by_node: dict[str, str] = {}
    for row in range(2, ws.max_row + 1):
        pid = ws.cell(row, 1).value
        parent = ws.cell(row, 2).value
        loc = ws.cell(row, 3).value
        if pid:
            by_port[str(pid)] = row
        if parent and loc and str(parent) not in loc_by_node:
            loc_by_node[str(parent)] = str(loc)
    for nid, port in sorted(cal_ports.items()):
        prev = None
        channel = 8
        loc = loc_by_node.get(nid, "")
        note = "Block sensor calibration current (first 5V OU pin 8); not a lamp"
        if port in by_port:
            prev = ws.cell(by_port[port], 5).value
            if prev:
                note += f"; was {prev}"
            loc = ws.cell(by_port[port], 3).value or loc
        vals = (
            port,
            nid,
            loc,
            channel,
            "Block sensor calibration",
            "Block Sensor Cal",
            "5V",
            note,
        )
        if port in by_port:
            rr = by_port[port]
            for col, v in enumerate(vals, start=1):
                ws.cell(rr, col, v)
            log.append(f"cal {port}: {prev!r} → Block sensor calibration")
        else:
            ws.append(vals)
            log.append(f"cal add {port}")
    return log


def load_v8_signal_info() -> dict[str, dict[str, str]]:
    """Route / block text from the frozen RGB workbook (imported snapshot)."""
    src = IMPORTED / "signals_split_v8.xlsx"
    wb = load_workbook(src, data_only=True)
    out: dict[str, dict[str, str]] = {}
    for name in wb.sheetnames:
        if not name.startswith("control_point_"):
            continue
        cp = "CP" + name.rsplit("_", 1)[-1]
        ws = wb[name]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = dict(zip(headers, row))
            sig = d.get("Signal")
            if not sig or str(sig) in out:
                continue
            out[str(sig)] = {
                "cp": cp,
                "from": str(d.get("Route_From") or ""),
                "to": str(d.get("Route_To") or ""),
                "block": str(d.get("Signal_Block") or ""),
                "location": str(d.get("Location") or ""),
            }
    wb.close()
    return out


def clear_retired_s3_rgb(ws: Worksheet) -> list[str]:
    """Drop leftover CP3 RGB on C3, stray S3-14 on C13, leftover S2 on C2. Remove C3-OU3."""
    log: list[str] = []
    drop_rows: list[int] = []
    for row in range(2, ws.max_row + 1):
        pid = str(ws.cell(row, 1).value or "")
        if pid.startswith("C3-OU3-"):
            drop_rows.append(row)
            continue
        prev = ws.cell(row, 5).value
        if pid.startswith("C3-OU2-") and pid != "C3-OU2-8" and prev:
            ws.cell(row, 5).value = None
            ws.cell(row, 6).value = None
            ws.cell(row, 8).value = f"spare (CP3 leftover RGB removed; was {prev})"
            log.append(f"clear {pid}: {prev!r}")
        elif pid == "C13-OU1-7" and prev and str(prev).startswith("S3-"):
            ws.cell(row, 5).value = None
            ws.cell(row, 6).value = None
            ws.cell(row, 8).value = f"spare (CP3 leftover RGB removed; was {prev})"
            log.append(f"clear {pid}: {prev!r}")
        elif pid in ("C2-OU2-7",) or pid.startswith("C2-OU3-"):
            ch = int(pid.rsplit("-", 1)[-1])
            # 24RB lives on C2-OU3-6/7/8 — keep that dwarf; drop leftover S2 only.
            if pid.startswith("C2-OU3-") and ch >= 6:
                continue
            if prev and str(prev).startswith("S2-"):
                ws.cell(row, 5).value = None
                ws.cell(row, 6).value = None
                ws.cell(row, 8).value = f"spare (CP2 leftover RGB removed; was {prev})"
                log.append(f"clear {pid}: {prev!r}")
    for row in reversed(drop_rows):
        pid = ws.cell(row, 1).value
        ws.delete_rows(row)
        log.append(f"drop {pid}")
    return log


def annotate_upper_rgb_heads(ws: Worksheet) -> list[str]:
    """Label remaining S4/S5/S6 LEDs as defined heads on CP4/CP5/CP6."""
    info = load_v8_signal_info()
    log: list[str] = []
    for row in range(2, ws.max_row + 1):
        dev = ws.cell(row, 5).value
        if not isinstance(dev, str):
            continue
        m = RGB_LED_RE.match(dev.strip())
        if not m:
            continue
        sig, color = m.group(1), m.group(4)
        meta = info.get(sig, {})
        cp = meta.get("cp") or f"CP{m.group(2)}"
        mast_id = UPPER_HEAD_NAME.get(sig, sig)
        new_name = f"Head {mast_id} {color}"
        bits = [cp]
        if meta.get("from") or meta.get("to"):
            bits.append(f"{meta.get('from') or '?'} → {meta.get('to') or '?'}")
        if meta.get("block"):
            bits.append(meta["block"])
        mast = V8_MAST_HINT.get(sig)
        if mast:
            bits.append(f"mast {mast}")
        else:
            bits.append("defined head")
        bits.append(f"was {sig}")
        note = " | ".join(bits)
        ws.cell(row, 5).value = new_name
        prev_note = ws.cell(row, 8).value
        if prev_note and str(prev_note) not in note:
            note = f"{note}; was {prev_note}"
        ws.cell(row, 8).value = note
        log.append(f"{ws.cell(row, 1).value}: {dev!r} → {new_name!r}")
    return log


def clear_s3_turnout_signals(ws: Worksheet) -> list[str]:
    """Switch 15–21 / 25–29 no longer have leftover S3/S2 RGB faces."""
    log: list[str] = []
    idx = header_index(ws)
    tcol = idx["Turnout"]
    retired = set(RETIRED_S3_TURNOUTS + RETIRED_S2_TURNOUTS)
    for row in range(2, ws.max_row + 1):
        name = str(ws.cell(row, tcol).value or "")
        if name not in retired:
            continue
        changed = False
        for col_name in TURNOUT_SIGNAL_COLS:
            col = idx.get(col_name)
            if not col:
                continue
            if ws.cell(row, col).value:
                ws.cell(row, col).value = None
                changed = True
        if changed:
            log.append(f"{name}: cleared leftover S2/S3 faces")
    return log


def tag_upper_turnouts_cp(ws: Worksheet) -> list[str]:
    """S4/S5/S6 plant switches get CP4/CP5/CP6 location labels."""
    log: list[str] = []
    idx = header_index(ws)
    tcol = idx["Turnout"]
    loc_col = idx.get("Location/Area")
    if not loc_col:
        return log
    for row in range(2, ws.max_row + 1):
        name = str(ws.cell(row, tcol).value or "")
        loc = UPPER_TURNOUT_CP.get(name)
        if not loc:
            continue
        prev = ws.cell(row, loc_col).value
        if prev != loc:
            ws.cell(row, loc_col).value = loc
            log.append(f"{name}: {prev!r} → {loc!r}")
    return log


def refresh_nodes(ws: Worksheet) -> None:
    idx = header_index(ws)
    loc_col = idx.get("Location")
    boards_5v_col = idx.get("5V Boards")
    num_5v_col = idx.get("Num 5V")
    leds_col = idx.get("Num Signal LEDs")
    id_col = idx["Node ID"]
    boards_12 = idx.get("12V Boards")
    num_12 = idx.get("Num 12V")
    for row in range(2, ws.max_row + 1):
        nid = ws.cell(row, id_col).value
        # D5 is helix DCC (radio 5). Do not invent 5V signal boards there.
        if nid == "D5":
            if loc_col:
                ws.cell(row, loc_col).value = "Helix DCC (radio 5; no Digicon DNOU8)"
            if boards_5v_col:
                ws.cell(row, boards_5v_col).value = 0
            if num_5v_col:
                ws.cell(row, num_5v_col).value = 0
            if leds_col:
                ws.cell(row, leds_col).value = 0
        if nid == "C1":
            if loc_col:
                ws.cell(row, loc_col).value = "Helix - Lower (Princess)"
            if boards_5v_col:
                ws.cell(row, boards_5v_col).value = 3
            if num_5v_col:
                ws.cell(row, num_5v_col).value = 24
            dnou = idx.get("DNOU8")
            if dnou:
                ws.cell(row, dnou).value = "OU1, OU2, OU3, OU4"
        if nid == "C2":
            if loc_col:
                ws.cell(row, loc_col).value = "North - Lower (East End west 24)"
            if boards_12:
                ws.cell(row, boards_12).value = 0
            if num_12:
                ws.cell(row, num_12).value = 0
            if boards_5v_col:
                ws.cell(row, boards_5v_col).value = 3
            if num_5v_col:
                ws.cell(row, num_5v_col).value = 24
            dnou = idx.get("DNOU8")
            if dnou:
                ws.cell(row, dnou).value = "OU1, OU2, OU3"
        if nid == "C3":
            if loc_col:
                ws.cell(row, loc_col).value = "West - Lower (Switch 15–21 motors)"
            if boards_5v_col:
                ws.cell(row, boards_5v_col).value = 1
            if num_5v_col:
                ws.cell(row, num_5v_col).value = 8
            used_col = idx.get("5V Used")
            free_col = idx.get("5V Free")
            leds_col_n = idx.get("Num Signal LEDs")
            led_dev_col = idx.get("Signal LED Devices")
            if used_col:
                ws.cell(row, used_col).value = 1
            if free_col:
                ws.cell(row, free_col).value = 7
            if leds_col_n:
                ws.cell(row, leds_col_n).value = 0
            if led_dev_col:
                ws.cell(row, led_dev_col).value = 0
            dnou = idx.get("DNOU8")
            if dnou:
                ws.cell(row, dnou).value = "OU1, OU2"
        if nid == "C4":
            if loc_col:
                ws.cell(row, loc_col).value = "West - Lower (Brick / Plane)"
            if boards_5v_col:
                ws.cell(row, boards_5v_col).value = 3
            if num_5v_col:
                ws.cell(row, num_5v_col).value = 24
            dnou = idx.get("DNOU8")
            if dnou:
                ws.cell(row, dnou).value = "OU1, OU2, OU3, OU4"
        if nid == "C11" and loc_col:
            ws.cell(row, loc_col).value = "Helix (Princess east 40 + balloon)"
        if nid == "C12" and loc_col:
            ws.cell(row, loc_col).value = "North - Lower (East End 34 + Switch 25–33)"
        if nid == "C13":
            if loc_col:
                ws.cell(row, loc_col).value = "West - Lower (Barn)"
            if boards_5v_col:
                ws.cell(row, boards_5v_col).value = 3
            if num_5v_col:
                ws.cell(row, num_5v_col).value = 24
            dnou = idx.get("DNOU8")
            if dnou:
                ws.cell(row, dnou).value = "OU1, OU2, OU3, OU4"


def refresh_turnout_summary(ws: Worksheet) -> list[str]:
    log: list[str] = []
    idx = header_index(ws)
    tcol = idx["Turnout"]
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, tcol).value
        spec = TURNOUT_DIGICON.get(str(name) if name else "")
        if not spec:
            continue
        ws.cell(row, tcol).value = proposed(str(name)) or name

        def set_group(prefix: str, signal_key: str, ports_key: str) -> None:
            sig = proposed(spec.get(signal_key)) or spec.get(signal_key)
            r_key, y_key, g_key = (
                f"{prefix} R Port",
                f"{prefix} Y Port",
                f"{prefix} G Port",
            )
            if not sig:
                ws.cell(row, idx[prefix]).value = None
                for key in (r_key, y_key, g_key):
                    if key in idx:
                        ws.cell(row, idx[key]).value = None
                return
            ws.cell(row, idx[prefix]).value = sig
            ports = spec.get(ports_key) or (None, None, None)
            # R/Y/G columns are lamp colors of one 3-pin head.
            if r_key in idx:
                ws.cell(row, idx[r_key]).value = ports[0]
            if y_key in idx:
                ws.cell(row, idx[y_key]).value = ports[1]
            if g_key in idx:
                ws.cell(row, idx[g_key]).value = ports[2]

        set_group("Entry Signal", "entry", "entry_ports")
        set_group("Normal Exit Signal", "normal", "normal_ports")
        set_group("Reverse Exit Signal", "reverse", "reverse_ports")
        log.append(f"{name}: Digicon faces {spec.get('entry')} / {spec.get('normal')} / {spec.get('reverse')}")
    return log


def add_digicon_sheet(wb, wiring: list[dict[str, str]], heads: list[dict[str, str]]) -> None:
    if "DigiconSignals" in wb.sheetnames:
        del wb["DigiconSignals"]
    ws = wb.create_sheet("DigiconSignals", 0)
    ws["A1"] = "HART Digicon searchlight heads (from cats/data/signal_wiring.csv)"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:L1")
    headers = [
        "port_id",
        "parent_node_id",
        "mqtt_node",
        "board_location",
        "signal_index",
        "packed",
        "system_name",
        "user_name",
        "mast_user_name",
        "disc_role",
        "head_role",
        "lamp_color",
        "topic",
        "notes",
    ]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(2, i, h)
        cell.font = Font(bold=True)
    for r_i, row in enumerate(wiring, start=3):
        for c_i, key in enumerate(headers, start=1):
            ws.cell(r_i, c_i, row.get(key))
    start = 3 + len(wiring) + 2
    ws.cell(start, 1, "Masts (cats/data/signal_head_plan.csv)")
    ws.cell(start, 1).font = Font(bold=True, size=12)
    m_headers = list(heads[0].keys()) if heads else []
    for i, h in enumerate(m_headers, start=1):
        cell = ws.cell(start + 1, i, h)
        cell.font = Font(bold=True)
    for r_i, row in enumerate(heads, start=start + 2):
        for c_i, key in enumerate(m_headers, start=1):
            ws.cell(r_i, c_i, row.get(key))
    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 22


def rebuild_asbuilt_inventory(ws: Worksheet, wiring: list[dict[str, str]], masts: list[dict[str, str]]) -> None:
    heads_by_mast: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in wiring:
        heads_by_mast[row["mast_user_name"]].append(row)
    # Keep header
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    plan_by_name = {r["proposed_mast_name"]: r for r in masts}
    for mast, pin_rows in heads_by_mast.items():
        pin_rows = sorted(
            pin_rows,
            key=lambda r: (
                DISC_SORT.get(r.get("disc_role") or r.get("head_role") or "S", 9),
                r.get("lamp_color") or "",
            ),
        )
        plan = plan_by_name.get(mast, {})
        packed_seen: list[str] = []
        topics_seen: list[str] = []
        for r in pin_rows:
            p = str(r["packed"])
            if p not in packed_seen:
                packed_seen.append(p)
                topics_seen.append(r["topic"])
        n = len(packed_seen)
        ports = " ".join(r["port_id"] for r in pin_rows)
        aspect = (
            "hart-aar SL-2-digicon (two 3-pin STOP/APPROACH/CLEAR discs)"
            if n == 2
            else "AAR-1946 SL-1-low (3-pin STOP/APPROACH/CLEAR)"
        )
        ws.append(
            [
                plan.get("cp") or pin_rows[0].get("board_location"),
                mast,
                n,
                plan.get("direction") or "",
                f"({plan.get('panel_x')},{plan.get('panel_y')}) {plan.get('edge')}" if plan else "",
                plan.get("protects_switch") or "",
                pin_rows[0]["mqtt_node"],
                pin_rows[0]["parent_node_id"],
                " ".join(packed_seen),
                ports,
                " ".join(topics_seen),
                plan.get("mast_system_name") or "",
                "double" if n == 2 else "single",
                aspect,
                pin_rows[0].get("notes") or "",
            ]
        )


def add_split_readme(path: Path) -> None:
    wb = load_workbook(path)
    if "README" in wb.sheetnames:
        del wb["README"]
    ws = wb.create_sheet("README", 0)
    lines = [
        ("signals_split_v8 — frozen planned RGB matrix", True),
        ("", False),
        ("This workbook is the Nov 2025 RGB LED plan (S1-1 … S6-15).", False),
        ("Lower-deck switch columns use live CTC names (Switch 1, Switch 35, …).", False),
        ("Upper-deck plants are Switch 61, 63, 65…; heads are 62L, 64R, 66RA….", False),
        ("S4-* = CP4, S5-* = CP5, S6-* = CP6 (historical IDs kept in Notes as was S4-1).", False),
        ("Lower-deck Digicon searchlights (6LB, 8RA, 36RA, …) are NOT in this file.", False),
        ("", False),
        ("Live lower-deck signal matrix:", False),
        ("  docs/wiring/signals_asbuilt_abs_v2.xlsx", False),
        ("Port / MQTT / LCOS overlay:", False),
        ("  cats/data/signal_wiring.csv", False),
        ("  docs/wiring/LCOS_Layout_Inventory_v85.xlsx  (DigiconSignals + DNOU8)", False),
        ("", False),
        ("Upper-deck CP4–CP6 (Helix/North/Peninsula Upper) still use this RGB plan.", False),
        ("Do not rename S1-* here to Digicon 11x names — those are a different plant.", False),
    ]
    for i, (text, bold) in enumerate(lines, start=1):
        ws.cell(i, 1, text).font = Font(bold=bold, size=14 if bold else 11)
        ws.row_dimensions[i].height = 18
    ws.column_dimensions["A"].width = 100
    ws["A1"].alignment = Alignment(wrap_text=True)
    wb.save(path)


def add_v8_mast_column(path: Path) -> int:
    """Fill Mast only when v8 routes make a grouping obvious."""
    wb = load_workbook(path)
    n = 0
    for name in wb.sheetnames:
        if not name.startswith("control_point_"):
            continue
        ws = wb[name]
        idx = header_index(ws)
        if "Mast" not in idx:
            sig_col = idx["Signal"]
            ws.insert_cols(sig_col + 1)
            ws.cell(1, sig_col + 1).value = "Mast"
            ws.cell(1, sig_col + 1).font = Font(bold=True)
            idx = header_index(ws)
        sig_col = idx["Signal"]
        mast_col = idx["Mast"]
        for row in range(2, ws.max_row + 1):
            sig = ws.cell(row, sig_col).value
            hint = V8_MAST_HINT.get(str(sig) if sig else "")
            if hint:
                ws.cell(row, mast_col).value = hint
                n += 1
    sw_n = 0
    for ws in wb.worksheets:
        if ws.title == "README":
            continue
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str) or not cell.value:
                    continue
                out = cell.value
                for pat, repl in _SW_TOKEN_CTC:
                    out = pat.sub(repl, out)
                if out != cell.value:
                    cell.value = out
                    sw_n += 1
    upper_n = apply_upper_deck_names(wb)
    wb.save(path)
    print(f"  v8 Mast cells: {n}; CTC switch tokens: {sw_n}; upper 61+ names: {upper_n}")
    return n


def refresh_inventory() -> Path:
    src = IMPORTED / "LCOS_Layout_Inventory_v84.xlsx"
    dest = WIRING / "LCOS_Layout_Inventory_v85.xlsx"
    shutil.copy2(src, dest)
    wb = load_workbook(dest)
    legacy_by_row = remap_workbook_node_ids(wb)
    mangled_n = fix_mangled_switch_names(wb)
    add_legacy_node_column(wb["Nodes"], legacy_by_row)
    sort_nodes_by_address(wb["Nodes"])
    occ = occupancy_by_hw()
    wiring = load_csv(CATS / "signal_wiring.csv")
    apply_proposed_to_wiring(wiring)
    heads = load_csv(CATS / "signal_head_plan.csv")
    apply_proposed_to_wiring(heads)

    bs_log = refresh_block_sensors(wb["BlockSensors"], occ)
    ou4_log = ensure_ou4_boards(wb["DNOU8"])
    dnou_log = overlay_dnou8(wb["DNOU8"], wiring)
    cal_ports = first_5v_cal_ports(wb["DNOU8"], wb["BlockSensors"])
    used = {r["port_id"] for r in wiring}
    clash = sorted(p for p in cal_ports.values() if p in used)
    if clash:
        raise SystemExit(
            "Digicon heads on block-sensor calibration pins: " + ", ".join(clash)
        )
    cal_log = stamp_block_sensor_cal(wb["DNOU8"], cal_ports)
    s3_log = clear_retired_s3_rgb(wb["DNOU8"])
    rgb_log = annotate_upper_rgb_heads(wb["DNOU8"])
    ts_log = refresh_turnout_summary(wb["TurnoutSummary"])
    ctc_n = apply_ctc_switch_names(wb)
    sw_n = apply_legacy_sw_tokens(wb)
    s3_ts_log = clear_s3_turnout_signals(wb["TurnoutSummary"])
    cp_ts_log = tag_upper_turnouts_cp(wb["TurnoutSummary"])
    upper_n = apply_upper_deck_names(wb)
    refresh_nodes(wb["Nodes"])
    summarize_node_labels(wb["Nodes"], wb["DNOU8"], wb["DNIN8"])
    add_digicon_sheet(wb, wiring, heads)
    mismatches = []
    nidx = header_index(wb["Nodes"])
    for row in range(2, wb["Nodes"].max_row + 1):
        nid = wb["Nodes"].cell(row, nidx["Node ID"]).value
        addr = wb["Nodes"].cell(row, nidx["Address"]).value
        if not nid or str(nid)[0] not in ("C", "D") or addr is None:
            continue
        expect = f"{str(nid)[0]}{int(addr)}"
        if str(nid) != expect:
            mismatches.append(f"{nid} address {addr} (expected {expect})")
    if mismatches:
        raise SystemExit("Node ID != C/D{Address}: " + "; ".join(mismatches))
    wb.save(dest)

    print(f"wrote {dest}")
    print(f"  Mangled SW1xx restorations (cells): {mangled_n}")
    print(f"  DCC Switch 100–119 → CTC (cells): {ctc_n}")
    print(f"  Legacy SW/SCX tokens (cells): {sw_n}")
    print(f"  Upper-deck 61+ names (cells): {upper_n}")
    print(f"  BlockSensors renames: {len(bs_log)}")
    for line in bs_log:
        print(f"    {line}")
    print(f"  DNOU8 OU4 added: {len(ou4_log)}")
    for line in ou4_log:
        print(f"    {line}")
    print(f"  DNOU8 overlays: {len(dnou_log)}")
    for line in dnou_log:
        print(f"    {line}")
    print(f"  DNOU8 block-sensor cal: {len(cal_log)}")
    for line in cal_log:
        print(f"    {line}")
    print(f"  Retired C3 S3 RGB / drop OU3: {len(s3_log)}")
    for line in s3_log:
        print(f"    {line}")
    print(f"  Upper-deck RGB → CP heads: {len(rgb_log)}")
    print(f"  TurnoutSummary Digicon: {len(ts_log)}")
    for line in ts_log:
        print(f"    {line}")
    print(f"  TurnoutSummary cleared S3 faces: {len(s3_ts_log)}")
    for line in s3_ts_log:
        print(f"    {line}")
    print(f"  TurnoutSummary CP4/5/6 labels: {len(cp_ts_log)}")
    for line in cp_ts_log:
        print(f"    {line}")
    return dest


# Princess sheet / all_logic column order (asbuilt v1).
PRINCESS_LOGIC_HEADERS = [
    "Control_Point",
    "Location",
    "Signal",
    "Signal_Block",
    "Route_From",
    "Route_To",
    "Block1",
    "Block2",
    "Next_Signal",
    "SW100",
    "SW101",
    "SW102",
    "SW117",
    "SW111",
    "SW110",
    "SW112",
    "SW113",
    "SW114",
    "SW115",
    "B1",
    "B2",
    "BO_Rule",
    "BO_Aspect",
    "MQTT_Out",
    "Notes",
]

# Clear / Approach / Stop occupancy flags (B1 = stop, B2 = approach).
_DWARF_ASPECTS = (
    (None, None, "R281", "Clear", "Green"),
    (None, "X", "R285", "Approach", "Yellow"),
    ("X", None, "R292", "Stop", "Red"),
)


def _princess_dwarf_rows() -> list[dict[str, object]]:
    """ABS rows for 114LA / 115LA / 120R / 120L (SML PAIRS + SIGNAL_FACING)."""
    loc = "Princess / Helix DCC"
    routes = [
        # K-1 dwarf, westbound. 115 Closed = K-1; dest 111L (113 N) or 112L (113 R).
        dict(
            Signal="115LA",
            Signal_Block="K-1",
            Route_From="K-1",
            Route_To="West Main Ext",
            Block1="OS 115",
            Block2="West Main Ext",
            Next_Signal="111L",
            SW113="N",
            SW115="N",
            ih="IH142",
            Notes="K-1 dwarf (SL-1-low). SW115 Closed = K-1. Dest 111L when 113 Normal.",
        ),
        dict(
            Signal="115LA",
            Signal_Block="K-1",
            Route_From="K-1",
            Route_To="East Lead",
            Block1="OS 115",
            Block2="East Lead",
            Next_Signal="112L",
            SW113="R",
            SW115="N",
            ih="IH142",
            Notes="K-1 dwarf. Dest 112L when 113 Reverse (East Lead).",
        ),
        # K-2 dwarf, westbound. 114 Closed = K-2.
        dict(
            Signal="114LA",
            Signal_Block="K-2",
            Route_From="K-2",
            Route_To="West Main Ext",
            Block1="OS 114",
            Block2="West Main Ext",
            Next_Signal="111L",
            SW113="N",
            SW114="N",
            ih="IH143",
            Notes="K-2 dwarf (SL-1-low). SW114 Closed = K-2. Dest 111L when 113 Normal.",
        ),
        dict(
            Signal="114LA",
            Signal_Block="K-2",
            Route_From="K-2",
            Route_To="East Lead",
            Block1="OS 114",
            Block2="East Lead",
            Next_Signal="112L",
            SW113="R",
            SW114="N",
            ih="IH143",
            Notes="K-2 dwarf. Dest 112L when 113 Reverse (East Lead).",
        ),
        # Balloon A48: south-track 120R (IH134) protects Rocks / OS 115; dest 120L.
        dict(
            Signal="120R",
            Signal_Block="McKees Rocks",
            Route_From="McKeesport",
            Route_To="McKees Rocks",
            Block1="McKees Rocks",
            Block2="OS 115",
            Next_Signal="120L",
            SW114="R",
            ih="IH134",
            Notes=(
                "Balloon connector (SL-1-low). A48 east ends join: 120R/IH134 on McKeesport "
                "track protects Rocks / OS 115. SW114 Thrown = McKeesport."
            ),
        ),
        # North-track 120L (IH141) protects McKeesport / OS 114; dest 120R.
        dict(
            Signal="120L",
            Signal_Block="McKeesport",
            Route_From="McKees Rocks",
            Route_To="McKeesport",
            Block1="McKeesport",
            Block2="OS 114",
            Next_Signal="120R",
            SW115="R",
            ih="IH141",
            Notes=(
                "Balloon connector (SL-1-low). 120L/IH141 on Rocks track protects McKeesport / "
                "OS 114. SML also Stop if McKees Rocks occupied. SW115 Thrown = Rocks."
            ),
        ),
    ]
    out: list[dict[str, object]] = []
    for route in routes:
        ih = route.pop("ih")
        for b1x, b2x, rule, aspect, color in _DWARF_ASPECTS:
            row = {h: None for h in PRINCESS_LOGIC_HEADERS}
            row.update(
                Control_Point="Princess",
                Location=loc,
                B1=b1x,
                B2=b2x,
                BO_Rule=rule,
                BO_Aspect=aspect,
                MQTT_Out=f"{ih}={color}",
                **route,
            )
            out.append(row)
    public_keys = (
        "Signal",
        "Signal_Block",
        "Route_From",
        "Route_To",
        "Block1",
        "Block2",
        "Next_Signal",
    )
    for row in out:
        for key in public_keys:
            value = row.get(key)
            if isinstance(value, str):
                row[key] = proposed(value) or value
    return out


def _logic_signals(ws: Worksheet) -> set[str]:
    idx = header_index(ws)
    col = idx.get("Signal")
    if not col:
        return set()
    found: set[str] = set()
    for row in range(2, ws.max_row + 1):
        v = ws.cell(row, col).value
        if v:
            found.add(str(v))
    return found


def append_princess_dwarf_logic(wb) -> int:
    """Add 114LA/115LA/120R/120L ABS rows; strip leftover 3-head MQTT on 114LB/115LB."""
    added = 0
    new_rows = _princess_dwarf_rows()
    for sheet_name in ("Princess", "all_logic"):
        ws = wb[sheet_name]
        have = _logic_signals(ws)
        idx = header_index(ws)
        for spec in new_rows:
            if spec["Signal"] in have:
                continue
            ws.append([spec.get(h) for h in PRINCESS_LOGIC_HEADERS])
            added += 1
        # 114LB / 115LB were authored as triples; IH134/IH141 are now 120R/120L.
        mqtt_col = idx.get("MQTT_Out")
        notes_col = idx.get("Notes")
        sig_col = idx.get("Signal")
        if mqtt_col and sig_col:
            for row in range(2, ws.max_row + 1):
                sig = ws.cell(row, sig_col).value
                mqtt = ws.cell(row, mqtt_col).value
                if not isinstance(mqtt, str):
                    continue
                if sig == "115LB" and "IH134" in mqtt:
                    ws.cell(row, mqtt_col).value = mqtt.replace(" IH134=Red", "").replace(" IH134=Green", "").replace(
                        " IH134=Yellow", ""
                    )
                if sig == "114LB" and "IH141" in mqtt:
                    ws.cell(row, mqtt_col).value = mqtt.replace(" IH141=Red", "").replace(" IH141=Green", "").replace(
                        " IH141=Yellow", ""
                    )
                if notes_col and sig in ("114LB", "115LB"):
                    notes = ws.cell(row, notes_col).value
                    if isinstance(notes, str) and "Triple" in notes:
                        ws.cell(row, notes_col).value = notes.replace("Triple exit. ", "2-head exit. ")
    return added


def refresh_asbuilt() -> Path:
    src = IMPORTED / "signals_asbuilt_abs_v1.xlsx"
    dest = WIRING / "signals_asbuilt_abs_v2.xlsx"
    shutil.copy2(src, dest)
    wb = load_workbook(dest)
    wiring = load_csv(CATS / "signal_wiring.csv")
    apply_proposed_to_wiring(wiring)
    masts = load_csv(CATS / "signal_mast_plan.csv")
    for row in masts:
        if row.get("proposed_mast_name"):
            row["proposed_mast_name"] = (
                proposed(row["proposed_mast_name"]) or row["proposed_mast_name"]
            )
    rebuild_asbuilt_inventory(wb["inventory"], wiring, masts)
    n = 0
    for name in wb.sheetnames:
        if name == "inventory":
            continue
        n += walk_replace(wb[name], NAME_REPLACEMENTS)
    occ = occupancy_by_hw()
    mqtt = wb["mqtt_topics"]
    idx = header_index(mqtt)
    name_col = idx.get("Name")
    jmri_col = idx.get("JMRI")
    if name_col and jmri_col:
        for row in range(2, mqtt.max_row + 1):
            jmri = mqtt.cell(row, jmri_col).value
            if not isinstance(jmri, str):
                continue
            m = BLOCK_IN_NOTES.search(jmri)
            if not m:
                continue
            public = occ.get(f"Block {m.group(1)}")
            if public:
                mqtt.cell(row, name_col).value = public.split(" / ", 1)[0]
                if "Digicon_Block_or_Role" in idx:
                    mqtt.cell(row, idx["Digicon_Block_or_Role"]).value = public
    dwarf_n = append_princess_dwarf_logic(wb)
    # README header
    readme = wb["README_ABS"]
    readme["A1"] = "HART Digicon as-built signals — ABS field logic (v2)"
    readme["A2"] = (
        "Refreshed 2026-08-22 from signal_head_plan / signal_wiring / public names "
        "(Scale, Barn, S-1…S-5, W-1/W-2, EH-*, Digicon 100L/117LA/114LA). "
        "Companion to frozen signals_split_v8.xlsx (planned RGB). "
        "100L is IH438/IH439 searchlights (not MQTT mast 464). "
        "Princess dwarfs 114LA / 115LA / 120R / 120L have ABS rows on Princess + all_logic."
    )
    wb.save(dest)
    print(f"wrote {dest} ({n} cell replacements on logic/mqtt sheets, +{dwarf_n} Princess dwarf rows)")
    cats_copy = CATS / "signals_asbuilt_abs_v2.xlsx"
    shutil.copy2(dest, cats_copy)
    old = CATS / "signals_asbuilt_abs_v1.xlsx"
    if old.exists():
        old.unlink()
        print(f"replaced {old} → {cats_copy}")
    return dest


def main() -> None:
    refresh_inventory()
    refresh_asbuilt()
    split_dest = WIRING / "signals_split_v8.xlsx"
    shutil.copy2(IMPORTED / "signals_split_v8.xlsx", split_dest)
    add_v8_mast_column(split_dest)
    add_split_readme(split_dest)
    print(f"wrote {split_dest} (CTC switch names + Mast column + README)")
    # Ports extract for cats/data
    inv = load_workbook(WIRING / "LCOS_Layout_Inventory_v85.xlsx")
    extract = CATS / "LCOS_Layout_Inventory_v85_signal_ports.xlsx"
    # Keep a thin copy of the full v85 next to the old v48 extract name.
    shutil.copy2(WIRING / "LCOS_Layout_Inventory_v85.xlsx", extract)
    old_extract = CATS / "LCOS_Layout_Inventory_v48_signal_ports.xlsx"
    if old_extract.exists():
        old_extract.unlink()
    print(f"wrote {extract}")
    inv.close()


if __name__ == "__main__":
    main()
