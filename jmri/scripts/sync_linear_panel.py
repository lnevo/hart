#!/usr/bin/env python3
"""
Align linear.xml with mac_jmri2.xml: snap shared point/turnout coordinates to the reference panel,
infer segment->block for new/changed geometry from reference + graph propagation, double on-canvas
draw scale (xscale/yscale, line widths, turnout sizes) while leaving window/panel pixel size unchanged.

Usage:
  python3 sync_linear_panel.py [linear.xml] [reference_panel.xml] [layout_blocks.xlsx]
"""
import copy
import os
import sys
import xml.etree.ElementTree as ET

import openpyxl

JMRI_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, JMRI_ROOT)
from layout_paths import layout_paths

_PATHS = layout_paths()
SCRIPT_DIR = _PATHS["working"]
COORD_ATTRS = frozenset(
    {"x", "y", "xcen", "ycen", "xa", "ya", "xb", "yb", "xc", "yc", "xd", "yd"}
)
DRAW_SCALE = 2.0
WINDOW_ATTRS = frozenset(
    {"windowheight", "windowwidth", "panelheight", "panelwidth", "height", "width"}
)


def _local_tag(elem):
    tag = elem.tag or ""
    if "}" in tag:
        tag = tag.split("}", 1)[1]
    return tag.lower()


def _find_layout(root):
    layout = root.find(".//LayoutEditor")
    if layout is None:
        layout = root.find(".//{*}LayoutEditor")
    if layout is None:
        for elem in root.iter():
            if elem.tag and "LayoutEditor" in elem.tag:
                return elem
    return layout


def _index_ref_geometry(ref_layout):
    """ident -> element for positionable points and layout turnouts (and similar)."""
    by_ident = {}
    for elem in ref_layout:
        t = _local_tag(elem)
        if "positionablepoint" in t or "layoutturnout" in t:
            ident = elem.get("ident")
            if ident:
                by_ident[ident] = elem
    return by_ident


def _index_ref_segments(ref_layout):
    """ident -> (connect1, connect2, blockname or None)."""
    out = {}
    for elem in ref_layout:
        if _local_tag(elem) != "tracksegment":
            continue
        ident = elem.get("ident")
        if not ident:
            continue
        c1, c2 = elem.get("connect1name"), elem.get("connect2name")
        bn = elem.get("blockname")
        out[ident] = (c1, c2, bn)
    return out


def _snap_coords(target_elem, ref_elem):
    for a in COORD_ATTRS:
        v = ref_elem.get(a)
        if v is not None:
            target_elem.set(a, v)


def _sync_bezier_controlpoints(target_seg, ref_seg):
    ref_cp = None
    for ch in ref_seg:
        if _local_tag(ch) == "controlpoints":
            ref_cp = ch
            break
    if ref_cp is None:
        return
    to_remove = [ch for ch in list(target_seg) if _local_tag(ch) == "controlpoints"]
    for ch in to_remove:
        target_seg.remove(ch)
    target_seg.append(copy.deepcopy(ref_cp))
    for k in ("bezier", "hideConLines"):
        if ref_seg.get(k) is not None:
            target_seg.set(k, ref_seg.get(k))


def _ensure_drawing_options(layout, ref_layout, factor=DRAW_SCALE):
    has_opts = any(_local_tag(ch) == "layouttrackdrawingoptions" for ch in layout)
    if has_opts:
        return
    for ch in ref_layout:
        if _local_tag(ch) == "layouttrackdrawingoptions":
            clone = copy.deepcopy(ch)
            layout.insert(0, clone)
            break


def _apply_draw_scale(layout, factor=DRAW_SCALE):
    """Enlarge track/turnout drawing without changing window dimensions."""
    def scale_float_str(val, f):
        try:
            return str(float(val) * f)
        except (TypeError, ValueError):
            return val

    def scale_int_str(val, f):
        try:
            return str(int(round(float(val) * f)))
        except (TypeError, ValueError):
            return val

    for attr in ("xscale", "yscale"):
        cur = layout.get(attr) or "1.0"
        try:
            nv = float(cur) * factor
            layout.set(attr, str(nv))
        except ValueError:
            layout.set(attr, str(factor))

    for attr in ("mainlinetrackwidth", "sidetrackwidth"):
        if layout.get(attr) is not None:
            layout.set(attr, scale_int_str(layout.get(attr), factor))

    for attr in (
        "turnoutbx",
        "turnoutcx",
        "turnoutwid",
        "turnoutcirclesize",
        "xoverlong",
        "xoverhwid",
        "xovershort",
    ):
        if layout.get(attr):
            layout.set(attr, scale_float_str(layout.get(attr), factor))

    for child in layout:
        t = _local_tag(child)
        if t != "layouttrackdrawingoptions":
            continue
        for sub in child.iter():
            tag = _local_tag(sub)
            if sub.text is None or not str(sub.text).strip():
                continue
            if "width" in tag:
                sub.text = scale_int_str(sub.text.strip(), factor)
            # Keep colors / counts as-is


def _segment_adjacency(linear_layout):
    """segment ident -> set of adjacent segment idents (via shared node names)."""
    node_to_segs = {}
    for elem in linear_layout:
        if _local_tag(elem) != "tracksegment":
            continue
        sid = elem.get("ident")
        if not sid:
            continue
        for end in (elem.get("connect1name"), elem.get("connect2name")):
            if not end:
                continue
            node_to_segs.setdefault(end, set()).add(sid)
    adj = {}
    for seg_list in node_to_segs.values():
        for s in seg_list:
            adj.setdefault(s, set()).update(seg_list - {s})
    return adj


def _infer_segment_blocks(linear_layout, ref_seg_index):
    """
    blockname per segment ident: exact topology match to ref preferred, else ref by ident,
    else propagate from neighbors that already have a block.
    """
    assign = {}
    for elem in linear_layout:
        if _local_tag(elem) != "tracksegment":
            continue
        ident = elem.get("ident")
        if not ident or ident not in ref_seg_index:
            continue
        _rc1, _rc2, rb = ref_seg_index[ident]
        if rb:
            # Same segment ident as reference panel → keep its block (topology may differ slightly).
            assign[ident] = rb

    changed = True
    adj = _segment_adjacency(linear_layout)
    while changed:
        changed = False
        for elem in linear_layout:
            if _local_tag(elem) != "tracksegment":
                continue
            ident = elem.get("ident")
            if not ident or ident in assign:
                continue
            neigh_blocks = []
            for nb in adj.get(ident, ()):
                if nb in assign:
                    neigh_blocks.append(assign[nb])
            if not neigh_blocks:
                continue
            first = neigh_blocks[0]
            if all(b == first for b in neigh_blocks):
                assign[ident] = first
                changed = True

    return assign


def _block_name_to_number(excel_path):
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    m = {}
    if "Blocks" in wb.sheetnames:
        ws = wb["Blocks"]
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            if row and row[0] is not None and row[1] is not None:
                m[str(row[1]).strip()] = row[0]
    wb.close()
    return m


def _update_segment_sheet(excel_path, seg_to_block_name):
    """Upsert Segment_to_Block rows from seg ident -> Block_Name."""
    if not excel_path or not os.path.isfile(excel_path):
        return 0
    name_to_num = _block_name_to_number(excel_path)
    wb = openpyxl.load_workbook(excel_path, read_only=False, data_only=False)
    if "Segment_to_Block" not in wb.sheetnames:
        wb.close()
        return 0
    ws = wb["Segment_to_Block"]
    existing = {}
    max_row = 1
    for r in range(2, ws.max_row + 1):
        seg = ws.cell(row=r, column=1).value
        if seg:
            max_row = max(max_row, r)
            existing[str(seg).strip()] = r

    n_write = 0
    for seg_id, bname in sorted(seg_to_block_name.items()):
        bname = str(bname).strip()
        num = name_to_num.get(bname)
        if num is None:
            continue
        row = existing.get(seg_id)
        if row is None:
            max_row += 1
            row = max_row
            existing[seg_id] = row
        ws.cell(row=row, column=1, value=seg_id)
        ws.cell(row=row, column=2, value=num)
        ws.cell(row=row, column=3, value=bname)
        n_write += 1
    wb.save(excel_path)
    wb.close()
    return n_write


def sync_linear(
    linear_path,
    ref_path,
    excel_path=None,
    draw_scale=DRAW_SCALE,
):
    lt = ET.parse(linear_path)
    rt = ET.parse(ref_path)
    lroot, rroot = lt.getroot(), rt.getroot()
    lin = _find_layout(lroot)
    ref = _find_layout(rroot)
    if lin is None or ref is None:
        raise SystemExit("LayoutEditor missing in linear or reference file")

    ref_by_ident = _index_ref_geometry(ref)
    ref_seg = _index_ref_segments(ref)
    n_snap = 0
    for elem in lin:
        t = _local_tag(elem)
        if "positionablepoint" in t or "layoutturnout" in t:
            ident = elem.get("ident")
            if ident and ident in ref_by_ident:
                _snap_coords(elem, ref_by_ident[ident])
                n_snap += 1
        elif t == "tracksegment":
            ident = elem.get("ident")
            if ident and ident in ref_seg:
                ref_el = None
                for e in ref:
                    if _local_tag(e) == "tracksegment" and e.get("ident") == ident:
                        ref_el = e
                        break
                if ref_el is not None:
                    _sync_bezier_controlpoints(elem, ref_el)

    seg_blocks = _infer_segment_blocks(lin, ref_seg)

    _ensure_drawing_options(lin, ref, factor=draw_scale)
    _apply_draw_scale(lin, factor=draw_scale)

    ET.indent(lt, space="  ", level=0)
    with open(linear_path, "w", encoding="utf-8") as f:
        f.write('<?xml version="1.0" encoding="UTF-8"?>\n')
        lt.write(f, encoding="unicode", default_namespace=None, xml_declaration=False)

    n_ex = 0
    if excel_path and seg_blocks:
        n_ex = _update_segment_sheet(excel_path, seg_blocks)

    print(f"Synced {os.path.basename(linear_path)} from {os.path.basename(ref_path)}")
    print(f"  Geometry snap: {n_snap} points/turnouts matched by ident")
    print(f"  Segment block mappings written: {len(seg_blocks)} (Excel rows touched: {n_ex})")
    print(
        f"  Draw scale x{draw_scale} on layout (xscale/yscale/line widths); window attrs unchanged: "
        f"{', '.join(sorted(a for a in WINDOW_ATTRS if lin.get(a)))}"
    )


def main():
    linear = sys.argv[1] if len(sys.argv) > 1 else os.path.join(_PATHS["working"], "linear.xml")
    ref = sys.argv[2] if len(sys.argv) > 2 else _PATHS["authoritative"]
    excel = sys.argv[3] if len(sys.argv) > 3 else _PATHS["excel"]
    if not os.path.isfile(linear):
        print(f"Not found: {linear}")
        sys.exit(1)
    if not os.path.isfile(ref):
        print(f"Not found: {ref}")
        sys.exit(1)
    sync_linear(linear, ref, excel_path=excel if os.path.isfile(excel) else None)


if __name__ == "__main__":
    main()
