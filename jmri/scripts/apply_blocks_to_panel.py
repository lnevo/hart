#!/usr/bin/env python3
"""
Apply block names from layout_blocks.xlsx to a JMRI panel XML file.
Sets blockname on every tracksegment and layoutturnout that has a mapping.
Output includes jmriversion, sensors, turnouts (from mac_jmri2.xml), blocks,
layoutblocks, and LayoutEditor so the layout has everything it needs.

Source: mac_jmri2.xml (authoritative for turnouts, defaults, labels). Track geometry
refresh: upper_both4.xml (AnyRail export). Do not use tables.xml or other legacy files for current data.

Refreshing from AnyRail (new track geometry): (1) Export updated track plan to upper_both4.xml.
(2) Run: apply_blocks_to_panel.py upper_both4.xml mac_jmri_blocked.xml mac_jmri2.xml use-panel-layout
Layout comes from upper_both4; defaults, labels, hidden tracks from mac_jmri2.xml.

Extras after use-panel-layout: no-nx (no Entry/Exit sensors/attrs; blockrouting=no), minimal-blocks
(emit only blocks that appear on segments or turnouts).
"""
import copy
import os
import re
import sys
import xml.etree.ElementTree as ET
import openpyxl

JMRI_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if JMRI_ROOT not in sys.path:
    sys.path.insert(0, JMRI_ROOT)
from layout_paths import layout_paths  # noqa: E402

_PATHS = layout_paths()
EXCEL_FILE = _PATHS["excel"]
MERGE_FILE = _PATHS["merge"]
# Authoritative panel (preferences, comments, defaults). Track refresh: upper_both4.xml.
DEFAULT_PANEL = _PATHS["authoritative"]
DEFAULT_OUTPUT = _PATHS["output"]
# Optional: if panel_path is an AnyRail export, use this file's layout instead (JMRI-cleaned).
LAYOUT_OVERRIDE_PANEL = _PATHS["authoritative"]
# Source for turnouts table (authoritative panel). Not tables.xml — use mac_jmri2.xml to avoid stale data.
REFERENCE_FILE = _PATHS["authoritative"]
STYLE_DEFAULTS = _PATHS["style_defaults"]
# If True, output turnouts as T1, T2, ... and rename layout idents to match. Disabled for now.
SIMPLIFY_TURNOUT_NAMES = False


def load_mappings(excel_path):
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    seg_to_block = {}
    if "Segment_to_Block" in wb.sheetnames:
        ws = wb["Segment_to_Block"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] is not None:
                seg_id, blk_num, blk_name = row[0], row[1], row[2]
                if blk_name:
                    seg_to_block[str(seg_id).strip()] = str(blk_name).strip()
    turnout_to_block = {}
    if "Turnout_to_Block" in wb.sheetnames:
        ws = wb["Turnout_to_Block"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] is not None:
                to_id, blk_num, blk_name = row[0], row[1], row[2]
                if blk_name:
                    turnout_to_block[str(to_id).strip()] = str(blk_name).strip()
    wb.close()
    return seg_to_block, turnout_to_block


def load_block_names(excel_path):
    """Load ordered list of block names (Block_1, Block_2, ...) from Blocks sheet."""
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    names = []
    if "Blocks" in wb.sheetnames:
        ws = wb["Blocks"]
        # Blocks sheet: col A = Block #, col B = Block Name
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            if row and row[0] is not None and row[1] is not None:
                names.append(str(row[1]).strip())
    wb.close()
    return names


def load_merge_pairs(merge_path):
    """
    Load block merge pairs from a text file. Each line: "KEEP MERGE_AWAY" (merge second into first).
    Example: "24 29" means merge Block_29 into Block_24 so that segment set becomes one block.
    Returns list of (keep_num, merge_away_num) 1-based block numbers.
    """
    if not merge_path or not os.path.isfile(merge_path):
        return []
    pairs = []
    with open(merge_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            parts = line.split()
            if len(parts) >= 2:
                try:
                    keep, away = int(parts[0]), int(parts[1])
                    if keep != away and keep >= 1 and away >= 1:
                        pairs.append((keep, away))
                except ValueError:
                    pass
    return pairs


def apply_merges(block_names, seg_to_block, turnout_to_block, merge_pairs):
    """
    Apply merge pairs: segments/turnouts in the 'merge away' block get the 'keep' block name;
    block list is reduced by skipping the merged-away block numbers.
    Modifies seg_to_block and turnout_to_block in place. Returns new block_names list (shorter).
    """
    if not merge_pairs:
        return block_names
    # Map: Block_29 -> Block_24, Block_114 -> Block_101, etc.
    merged_name = {}
    merge_away = set()
    for keep, away in merge_pairs:
        if away <= len(block_names):
            merged_name[block_names[away - 1]] = block_names[keep - 1]
            merge_away.add(away)
    for ident, blk in list(seg_to_block.items()):
        if blk in merged_name:
            seg_to_block[ident] = merged_name[blk]
    for ident, blk in list(turnout_to_block.items()):
        if blk in merged_name:
            turnout_to_block[ident] = merged_name[blk]
    # Build reduced block list: skip indices that are merged away (1-based).
    new_names = [block_names[i - 1] for i in range(1, len(block_names) + 1) if i not in merge_away]
    return new_names


def _block_number(block_name):
    """Extract block number from name: Block_24 -> 24. Returns None if not Block_N."""
    if not block_name or not block_name.startswith("Block_"):
        return None
    try:
        return int(block_name.split("_", 1)[1])
    except (ValueError, IndexError):
        return None


def _block_sensor_user_names(block_names, turnout_to_block):
    """
    One userName per block. For blocks that have at least one turnout, use "BS <turnout_ident>"
    (e.g. BS TOL35287); otherwise "Block Sensor N". Returns list of length len(block_names).
    """
    block_to_turnouts = {}
    for ident, block_name in (turnout_to_block or {}).items():
        block_to_turnouts.setdefault(block_name, []).append(ident)
    for blk in block_to_turnouts:
        block_to_turnouts[blk] = sorted(set(block_to_turnouts[blk]))
    names = []
    for i, name in enumerate(block_names, 1):
        num = _block_number(name)
        if name in block_to_turnouts and block_to_turnouts[name]:
            names.append("BS " + block_to_turnouts[name][0])
        else:
            names.append(f"Block Sensor {num}" if num is not None else f"Block Sensor {i}")
    return names


def build_sensors_element(block_names, turnout_to_block=None):
    """Create JMRI <sensors> element: one sensor per block. userName = BS <turnout> for turnout blocks, else Block Sensor N."""
    sensor_user_names = _block_sensor_user_names(block_names, turnout_to_block)
    sensors = ET.Element("sensors", attrib={"class": "jmri.jmrix.internal.configurexml.InternalSensorManagerXml"})
    ET.SubElement(sensors, "defaultInitialState").text = "unknown"
    for i, name in enumerate(block_names, 1):
        sen = ET.SubElement(sensors, "sensor", attrib={"inverted": "false"})
        ET.SubElement(sen, "systemName").text = f"ISIS{i}"
        ET.SubElement(sen, "userName").text = sensor_user_names[i - 1]
    return sensors


def get_input_block_comments(root):
    """Extract block userName -> comment text from input panel's <blocks> so we can preserve user comments."""
    out = {}
    blocks_elem = root.find("blocks")
    if blocks_elem is None:
        blocks_elem = root.find("{*}blocks")
    if blocks_elem is None:
        for elem in root.iter():
            if elem.tag and "blocks" in elem.tag and "BlockManager" in str(elem.get("class", "")):
                blocks_elem = elem
                break
    if blocks_elem is None:
        return out
    for b in blocks_elem.findall("block") or blocks_elem.findall("{*}block") or []:
        uname_elem = b.find("userName")
        if uname_elem is None:
            uname_elem = b.find("{*}userName")
        comment_elem = b.find("comment")
        if comment_elem is None:
            comment_elem = b.find("{*}comment")
        if uname_elem is not None and uname_elem.text and comment_elem is not None and comment_elem.text:
            out[uname_elem.text.strip()] = comment_elem.text.strip()
    return out


def get_generated_block_comments(layout_elem, turnout_to_block, block_names):
    """Same logic as Excel comments: turnouts + NX entry/exit names, no labels. Returns dict block_name -> comment string."""
    block_to_turnouts = {}
    for ident, block_name in turnout_to_block.items():
        block_to_turnouts.setdefault(block_name, []).append(ident)
    for block_name in block_to_turnouts:
        block_to_turnouts[block_name] = sorted(set(block_to_turnouts[block_name]))
    block_to_nx = get_block_to_nx_sensors(layout_elem)
    out = {}
    for block_name in block_names:
        parts = []
        if block_name in block_to_turnouts:
            parts.append(", ".join(block_to_turnouts[block_name]))
        if block_name in block_to_nx:
            parts.append(", ".join(block_to_nx[block_name]))
        if parts:
            out[block_name] = ". ".join(parts)
    return out


def add_sensor_comments(sensors_elem, block_names, generated_block_comments, block_to_nx, turnout_to_block=None):
    """
    Add or replace <comment> on each sensor so comments show in JMRI.
    Block Sensor N / BS <turnout> get the same comment as their block. NX sensors get the block names that use them.
    Replaces any existing test comments.
    """
    nx_to_blocks = {}
    for block_name, nx_list in block_to_nx.items():
        for nx_name in nx_list:
            nx_to_blocks.setdefault(nx_name, []).append(block_name)
    for nx_name in nx_to_blocks:
        nx_to_blocks[nx_name] = sorted(set(nx_to_blocks[nx_name]))
    for sen in sensors_elem:
        tag = (sen.tag or "").strip().lower()
        if tag != "sensor":
            continue
        uname_elem = sen.find("userName")
        if uname_elem is None:
            uname_elem = sen.find("{*}userName")
        if uname_elem is None or not uname_elem.text:
            continue
        user_name = uname_elem.text.strip()
        comment_text = None
        if user_name.startswith("BS ") and turnout_to_block:
            ident = user_name[3:].strip()
            block_name = turnout_to_block.get(ident)
            if block_name:
                comment_text = generated_block_comments.get(block_name)
        elif user_name.startswith("Block Sensor "):
            num_str = user_name.replace("Block Sensor ", "").strip()
            try:
                n = int(num_str)
                block_name = f"Block_{n}"
                if block_name in block_names:
                    comment_text = generated_block_comments.get(block_name)
            except ValueError:
                pass
        elif user_name.startswith("NX "):
            blocks = nx_to_blocks.get(user_name, [])
            if blocks:
                comment_text = ", ".join(blocks)
        if comment_text:
            existing = sen.find("comment")
            if existing is None:
                existing = sen.find("{*}comment")
            if existing is not None:
                existing.text = str(comment_text)
            else:
                ET.SubElement(sen, "comment").text = str(comment_text)


def build_blocks_element(block_names, comments=None, turnout_to_block=None):
    """Create JMRI <blocks> element: each block's occupancysensor = BS <turnout> or Block Sensor N.
    If comments dict is provided (block userName -> comment text), add <comment> child so comments show in JMRI."""
    blocks = ET.Element("blocks", attrib={"class": "jmri.configurexml.BlockManagerXml"})
    ET.SubElement(blocks, "defaultspeed").text = "Normal"
    comments = comments or {}
    sensor_user_names = _block_sensor_user_names(block_names, turnout_to_block)
    for i, user_name in enumerate(block_names, 1):
        sensor_name = sensor_user_names[i - 1]
        sys_name = f"IB:AUTO:{i:04d}"
        block = ET.SubElement(
            blocks, "block",
            attrib={"systemName": sys_name, "length": "0.0", "curve": "0"}
        )
        ET.SubElement(block, "systemName").text = sys_name
        ET.SubElement(block, "userName").text = user_name
        comment_text = comments.get(user_name)
        if comment_text:
            ET.SubElement(block, "comment").text = str(comment_text)
        ET.SubElement(block, "permissive").text = "no"
        ET.SubElement(block, "occupancysensor").text = sensor_name
    return blocks


def build_layoutblocks_element(block_names, turnout_to_block=None, blockrouting="yes"):
    """Create JMRI <layoutblocks> element: each layoutblock's occupancysensor = BS <turnout> or Block Sensor N.
    blockrouting='yes' enables advanced block routing so Entry/Exit Auto Generate Pairs can find valid paths."""
    layoutblocks = ET.Element(
        "layoutblocks",
        attrib={
            "class": "jmri.jmrit.display.layoutEditor.configurexml.LayoutBlockManagerXml",
            "blockrouting": blockrouting,
        }
    )
    sensor_user_names = _block_sensor_user_names(block_names, turnout_to_block)
    for i, user_name in enumerate(block_names, 1):
        sensor_name = sensor_user_names[i - 1]
        lb = ET.SubElement(
            layoutblocks, "layoutblock",
            attrib={
                "systemName": f"ILB{i}",
                "occupancysensor": sensor_name,
                "occupiedsense": "2",
                "trackcolor": "gray",
                "occupiedcolor": "red",
                "extracolor": "white",
            }
        )
        ET.SubElement(lb, "systemName").text = f"ILB{i}"
        ET.SubElement(lb, "userName").text = user_name
    return layoutblocks


def copy_turnouts_from_reference(ref_path, simplify_turnout_names=False):
    """
    Copy <turnouts> element from authoritative panel (mac_jmri2.xml). Do not use tables.xml for current data.
    If simplify_turnout_names is True, assign T1, T2, ... to userName and return (elem, old_to_new)
    for renaming layout idents. Otherwise return (elem, None).
    """
    if not ref_path or not os.path.isfile(ref_path):
        return None, None
    tree = ET.parse(ref_path)
    root = tree.getroot()
    turnouts = root.find("turnouts")
    if turnouts is None:
        turnouts = next((c for c in root if c.tag == "turnouts"), None)
    if turnouts is None:
        return None, None
    turnouts = copy.deepcopy(turnouts)
    # Strip <comment> from each turnout so reference-file comments (e.g. "Switch 4-3") don't
    # reappear in output when the user has removed them from their data source.
    for t in turnouts.findall("turnout") or turnouts.findall("{*}turnout") or []:
        to_remove = [c for c in t if (c.tag or "").strip().lower() == "comment"]
        for c in to_remove:
            t.remove(c)
    old_to_new = None
    if simplify_turnout_names:
        # Reference order (IT1, IT2, ...) -> T1, T2, ...
        old_to_new = {}
        for i, t in enumerate(turnouts.findall("turnout") or turnouts.findall("{*}turnout") or []):
            uname = t.find("userName")
            if uname is None:
                uname = t.find("{*}userName")
            if uname is not None and uname.text:
                old = uname.text.strip()
                new = f"T{i + 1}"
                old_to_new[old] = new
                uname.text = new
                # 3-way turnouts: reference has TO_3W35179-1, TO_3W35179-2; layout uses base in segment idents.
                # Map base -> first sub-turnout for T-I, T-E1, -CS; map T-O and T-E2 -> second sub-turnout.
                m = re.match(r"^(TO_3W\d+)-\d+$", old)
                if m:
                    base = m.group(1)
                    if old.endswith("-1"):
                        old_to_new[base] = new  # T-I-*, T-E1-*, *-CS -> T19
                    else:
                        # Second part (T20): segment idents for through and E2 leg
                        old_to_new[f"T-O-{base}"] = f"T-O-{new}"
                        old_to_new[f"T-E2-{base}"] = f"T-E2-{new}"
        if not old_to_new:
            old_to_new = None
    return turnouts, old_to_new


def add_turnout_feedback_sensors(sensors_elem, turnouts_elem, system_name_base):
    """
    For each turnout in the turnouts table: create two sensors (normal/closed + reverse/thrown) and set
    feedback="TWOSENSOR", sensor1=<userName> FB_R (reverse), sensor2=<userName> FB_N (normal).
    System names: normal ISIS (e.g. ISIS176, ISIS177, ...) starting at system_name_base; two per turnout.
    userName: "<turnout> FB_N", "<turnout> FB_R".
    Returns the number of turnouts that got feedback sensors.
    """
    turnouts_list = turnouts_elem.findall("turnout") or turnouts_elem.findall("{*}turnout") or []
    if not turnouts_list:
        return 0
    idx = 0
    for t in turnouts_list:
        uname_el = t.find("userName")
        if uname_el is None:
            uname_el = t.find("{*}userName")
        user_name = (uname_el.text or "").strip() if uname_el is not None else (t.get("userName") or "").strip()
        if not user_name:
            continue
        closed_user = f"{user_name} FB_N"
        thrown_user = f"{user_name} FB_R"
        sys_closed = system_name_base + 2 * idx
        sys_thrown = system_name_base + 2 * idx + 1
        for sys_name, u_name, comment_text in (
            (f"ISIS{sys_closed}", closed_user, f"{user_name} Closed"),
            (f"ISIS{sys_thrown}", thrown_user, f"{user_name} Thrown"),
        ):
            sen = ET.SubElement(sensors_elem, "sensor", attrib={"inverted": "false"})
            ET.SubElement(sen, "systemName").text = sys_name
            ET.SubElement(sen, "userName").text = u_name
            ET.SubElement(sen, "comment").text = comment_text
        t.set("feedback", "TWOSENSOR")
        t.set("sensor1", thrown_user)   # sensor1 = thrown (reverse)
        t.set("sensor2", closed_user)   # sensor2 = closed (normal)
        idx += 1
    return idx


def set_turnoutnames_in_layout(layout_elem):
    """Set turnoutname=ident on every layoutturnout so turnouts are 'set' (linked to roster) like in export."""
    for elem in layout_elem.iter():
        tag = elem.tag if hasattr(elem, "tag") else None
        if tag is None:
            continue
        if "layoutturnout" in (tag or "").lower() or (tag or "").endswith("LayoutTurnout"):
            ident = elem.get("ident") or elem.get("turnoutname")
            if ident:
                elem.set("turnoutname", ident)


# Link segment to connect Block_14 legs at duplicate end bumpers (EB265 / EB279) so they display connected.
BLOCK14_LINK_ENDS = ("EB265", "EB279")
BLOCK14_LINK_IDENT = "LINK-EB265-EB279"


# --- Sensor order in output: (1) block ISIS1-N, (2) turnout feedback FB-Nn/FB-Rn, (3) NX Entry/Exit ISIS200+ ---
NX_SENSOR_BASE = 200  # ISIS200, ISIS201, ... for NX boundary sensors (after block + feedback in document order)


def get_segment_block_map(layout_elem):
    """Build map segment ident -> block name from tracksegment blockname in layout."""
    seg_block = {}
    for elem in layout_elem.iter():
        tag = (elem.tag or "").strip().lower()
        if tag != "tracksegment":
            continue
        ident = elem.get("ident")
        block = elem.get("blockname")
        if ident and block:
            seg_block[ident] = block
    return seg_block


def collect_boundary_sensors(layout_elem, seg_block_map):
    """
    Find block boundaries and assign NX sensor user names.
    - END_BUMPER: one boundary -> one sensor (eastboundsensor), userName "NX <ident>".
    - ANCHOR where the two connected segments have different block names: two sensors (east + west), "NX <ident>-E" / "NX <ident>-W".
    Returns: list of (point_elem, east_user_name_or_None, west_user_name_or_None), and set of all NX sensor user names to create.
    """
    assignments = []
    sensor_names = set()
    for elem in layout_elem.iter():
        tag = (elem.tag or "").strip().lower()
        if "positionablepoint" not in tag:
            continue
        pt_type = (elem.get("type") or "").strip().upper()
        ident = elem.get("ident") or ""
        if not ident:
            continue
        c1 = elem.get("connect1name")
        c2 = elem.get("connect2name")
        if pt_type == "END_BUMPER":
            # One boundary: use the track side (ignore link segment like LINK-EB265-EB279 for naming)
            track_seg = c1 if c1 and (not c2 or "LINK-" in (c2 or "")) else c2
            if not track_seg or (track_seg and "LINK-" in track_seg):
                track_seg = c1 or c2
            if track_seg:
                name = f"NX {ident}"
                sensor_names.add(name)
                assignments.append((elem, name, None))
        elif pt_type == "ANCHOR" and c1 and c2:
            b1 = seg_block_map.get(c1)
            b2 = seg_block_map.get(c2)
            if b1 and b2 and b1 != b2:
                east_name = f"NX {ident}-E"
                west_name = f"NX {ident}-W"
                sensor_names.add(east_name)
                sensor_names.add(west_name)
                assignments.append((elem, east_name, west_name))
    return assignments, sensor_names


def collect_turnout_boundary_sensors(layout_elem, seg_block_map):
    """
    For each layout turnout, at each leg (A,B,C,D) where the connected segment's block
    differs from the turnout's block, we need an NX sensor (JMRI "Set Sensors..." at turnout).
    Returns: set of NX sensor user names to create, and list of (turnout_elem, {leg: sensor_name}).
    """
    sensor_names = set()
    turnout_assignments = []
    for elem in layout_elem.iter():
        tag = (elem.tag or "").strip().lower()
        if "layoutturnout" not in tag:
            continue
        ident = elem.get("ident") or elem.get("turnoutname") or ""
        if not ident:
            continue
        to_block = elem.get("blockname")
        if not to_block:
            continue
        legs = {}
        for leg, attr in (("A", "connectaname"), ("B", "connectbname"), ("C", "connectcname"), ("D", "connectdname")):
            seg_ident = elem.get(attr)
            if not seg_ident:
                continue
            seg_block = seg_block_map.get(seg_ident)
            if seg_block and seg_block != to_block:
                name = f"NX {ident}-{leg}"
                sensor_names.add(name)
                legs[leg] = name
        if legs:
            turnout_assignments.append((elem, legs))
    return sensor_names, turnout_assignments


def build_nx_sensor_elements(nx_user_names, start=None):
    """Create sensor elements for NX boundary sensors. systemName ISIS<start>, ISIS<start+1>, ...; userName as given."""
    if start is None:
        start = NX_SENSOR_BASE
    elements = []
    for i, user_name in enumerate(sorted(nx_user_names), start=start):
        sen = ET.Element("sensor", attrib={"inverted": "false"})
        ET.SubElement(sen, "systemName").text = f"ISIS{i}"
        ET.SubElement(sen, "userName").text = user_name
        elements.append(sen)
    return elements


def add_boundary_sensors_to_layout(layout_elem, seg_block_map, sensors_parent, nx_sensor_start=None):
    """
    Create NX sensors, add them to sensors_parent; set eastboundsensor/westboundsensor on
    positionable points; set sensorA/sensorB/sensorC/sensorD on layout turnouts (same logic
    as "Set Sensors..." in Layout Editor so Auto Generate Entry-Exit Pairs can find paths).
    nx_sensor_start: first ISIS index for NX sensors (default NX_SENSOR_BASE); use after block + feedback so ranges don't overlap.
    """
    point_assignments, point_nx_names = collect_boundary_sensors(layout_elem, seg_block_map)
    turnout_nx_names, turnout_assignments = collect_turnout_boundary_sensors(layout_elem, seg_block_map)
    nx_names = point_nx_names | turnout_nx_names
    if not nx_names:
        return 0
    for sen_elem in build_nx_sensor_elements(nx_names, start=nx_sensor_start):
        sensors_parent.append(sen_elem)
    for point_elem, east_name, west_name in point_assignments:
        if east_name:
            point_elem.set("eastboundsensor", east_name)
        if west_name:
            point_elem.set("westboundsensor", west_name)
    for to_elem, legs in turnout_assignments:
        for leg, sensor_name in legs.items():
            child = to_elem.find(f"sensor{leg}")
            if child is not None:
                child.text = sensor_name
            else:
                sub = ET.SubElement(to_elem, f"sensor{leg}")
                sub.text = sensor_name
    return len(point_assignments) + len(turnout_assignments)


def remove_sensoricons_from_layout(layout_elem):
    """
    Remove all sensoricon elements from the layout so no sensors are shown on the panel.
    Entry/exit logic (sensors in table, sensorA/B/C on turnouts, eastboundsensor/westboundsensor
    on positionable points) is unchanged; only the on-panel icons are stripped.
    """
    to_remove = [
        c for c in layout_elem
        if (c.tag or "").strip().lower() == "sensoricon"
    ]
    for elem in to_remove:
        layout_elem.remove(elem)
    return len(to_remove)


# LayoutEditor attributes to copy when applying defaults from another panel (name, colors, sizes).
LAYOUT_DEFAULT_ATTRS = (
    "name", "mainlinetrackwidth", "sidetrackwidth", "xscale", "yscale",
    "defaulttrackcolor", "defaultoccupiedtrackcolor", "defaultalternativetrackcolor", "defaulttextcolor",
    "drawgrid", "snaponadd", "snaponmove", "antialiasing", "turnoutcircles", "turnoutcirclecolor",
    "turnoutcirclethrowncolor", "turnoutfillcontrolcircles", "turnoutcirclesize", "turnoutdrawunselectedleg",
    "turnoutbx", "turnoutcx", "turnoutwid", "xoverlong", "xoverhwid", "xovershort",
    "redBackground", "greenBackground", "blueBackground", "gridSize", "gridSize2nd",
    "sliders", "scrollable", "tooltipsnotedit", "tooltipsinedit", "autoblkgenerate",
    "openDispatcher", "useDirectTurnoutControl",
)
# Standard label appearance (match Bridgeville, Scully Yard, etc.)
LABEL_STANDARD_FONT = "Lucida Grande"
LABEL_STANDARD_FONTNAME = "LucidaGrande-Bold"
LABEL_STANDARD_SIZE = "12"
LABEL_STANDARD_STYLE = "1"
BRIDGEVILLE_Y = "580"  # align "lower Duffs Junction" vertically with Bridgeville


def apply_layout_defaults(layout_elem, defaults_panel_path, defaults_root=None):
    """
    Copy LayoutEditor display defaults (name, colors, track options, layoutTrackDrawingOptions)
    from another panel file onto layout_elem. Use when the layout source is authoritative (e.g.
    upper_both4.xml) but you want visual defaults from your authoritative panel (e.g. mac_jmri2.xml).
    If defaults_root is provided (pre-parsed), it is used and defaults_panel_path is not read.
    """
    if defaults_root is None:
        if not defaults_panel_path or not os.path.isfile(defaults_panel_path):
            return False
        try:
            defaults_root = ET.parse(defaults_panel_path).getroot()
        except Exception:
            return False
    def_root = defaults_root
    def_layout = def_root.find(".//LayoutEditor")
    if def_layout is None:
        def_layout = def_root.find(".//{*}LayoutEditor")
    if def_layout is None:
        for elem in def_root.iter():
            if elem.tag and "LayoutEditor" in elem.tag:
                def_layout = elem
                break
    if def_layout is None:
        return False
    for attr in LAYOUT_DEFAULT_ATTRS:
        val = def_layout.get(attr)
        if val is not None:
            layout_elem.set(attr, val)
    # Replace layoutTrackDrawingOptions if present in defaults
    def_opts = def_layout.find("layoutTrackDrawingOptions")
    if def_opts is None:
        def_opts = def_layout.find("{*}layoutTrackDrawingOptions")
    if def_opts is not None:
        old_opts = layout_elem.find("layoutTrackDrawingOptions")
        if old_opts is None:
            old_opts = layout_elem.find("{*}layoutTrackDrawingOptions")
        if old_opts is not None:
            layout_elem.remove(old_opts)
        layout_elem.insert(0, copy.deepcopy(def_opts))
    return True


def remove_icon_positionable_labels(layout_elem):
    """
    Remove positionablelabel elements that are icon-type (background image). These can cause
    PositionableLabel import errors in JMRI when the layout comes from a different source
    (e.g. AnyRail export). Returns the number removed.

    IMPORTANT: Do not remove this step when updating for new AnyRail exports. AnyRail (or
    JMRI-saved panels that once came from AnyRail) can include a positionablelabel with
    icon="yes" and a child <icon> (e.g. background image). JMRI fails to load that element
    in our output; stripping it here keeps geometry correct and avoids the import error.
    Text labels (Washington, Bridgeville, etc.) are added later from defaults and are not
    affected.
    """
    removed = 0
    to_remove = []
    for c in layout_elem:
        tag = (c.tag or "").strip().lower()
        if tag != "positionablelabel":
            continue
        if c.get("icon") == "yes":
            to_remove.append(c)
            continue
        if c.find("icon") is not None or c.find("{*}icon") is not None:
            to_remove.append(c)
    for elem in to_remove:
        layout_elem.remove(elem)
        removed += 1
    return removed


def remove_all_positionable_labels(layout_elem):
    """
    Remove every positionablelabel from the layout. Use when the layout is from a different
    schema (e.g. AnyRail layout-4-19-2) so that labels from a 5.x defaults file would cause
    PositionableLabel import errors. Returns the number removed.
    """
    to_remove = [
        c for c in layout_elem
        if (c.tag or "").strip().lower() == "positionablelabel"
    ]
    for elem in to_remove:
        layout_elem.remove(elem)
    return len(to_remove)


def add_labels_from_defaults(layout_elem, defaults_panel_path, defaults_root=None):
    """
    Copy all positionablelabel elements from the defaults panel's LayoutEditor into the layout
    (e.g. Washington, Bridgeville, Scully Yard, etc.). Removes any existing labels in the layout
    that have the same text so there are no duplicates. Returns number of labels added.
    If defaults_root is provided (pre-parsed), it is used and defaults_panel_path is not read.
    """
    if defaults_root is None:
        if not defaults_panel_path or not os.path.isfile(defaults_panel_path):
            return 0
        try:
            defaults_root = ET.parse(defaults_panel_path).getroot()
        except Exception:
            return 0
    def_root = defaults_root
    def_layout = def_root.find(".//LayoutEditor")
    if def_layout is None:
        def_layout = def_root.find(".//{*}LayoutEditor")
    if def_layout is None:
        for elem in def_root.iter():
            if elem.tag and "LayoutEditor" in elem.tag:
                def_layout = elem
                break
    if def_layout is None:
        return 0
    default_labels = []
    for child in def_layout:
        tag = (child.tag or "").strip().lower()
        if tag == "positionablelabel":
            default_labels.append(child)
    if not default_labels:
        return 0
    default_texts = {lb.get("text") for lb in default_labels if lb.get("text")}
    to_remove = [
        c for c in layout_elem
        if (c.tag or "").strip().lower() == "positionablelabel" and c.get("text") in default_texts
    ]
    for elem in to_remove:
        layout_elem.remove(elem)
    insert_index = 1
    for i, opts in enumerate(layout_elem):
        tag = (opts.tag or "").strip().lower()
        if tag == "layouttrackdrawingoptions":
            insert_index = i + 1
            break
    for lb in default_labels:
        clone = copy.deepcopy(lb)
        # Standardize appearance to match Bridgeville/Scully Yard style
        clone.set("fontFamily", LABEL_STANDARD_FONT)
        clone.set("fontname", LABEL_STANDARD_FONTNAME)
        clone.set("size", LABEL_STANDARD_SIZE)
        clone.set("style", LABEL_STANDARD_STYLE)
        # Align lower "Duffs Junction" vertically with Bridgeville (same y)
        if (clone.get("text") or "").strip() == "Duffs Junction":
            try:
                y_val = float(clone.get("y") or 0)
                if y_val > 400:  # lower one (upper is ~180)
                    clone.set("y", BRIDGEVILLE_Y)
            except (TypeError, ValueError):
                pass
        layout_elem.insert(insert_index, clone)
        insert_index += 1
    return len(default_labels)


def apply_hidden_tracks_from_defaults(layout_elem, defaults_panel_path, seg_to_block, defaults_root=None):
    """
    Apply hidden track state from the defaults panel: (1) segment idents that have hidden="yes"
    in defaults get hidden="yes" in the layout; (2) if a hidden segment exists in defaults but
    not in the layout (e.g. missing from AnyRail export), copy it in with blockname applied.
    Returns (number set hidden, number added).
    If defaults_root is provided (pre-parsed), it is used and defaults_panel_path is not read.
    """
    if defaults_root is None:
        if not defaults_panel_path or not os.path.isfile(defaults_panel_path):
            return 0, 0
        try:
            defaults_root = ET.parse(defaults_panel_path).getroot()
        except Exception:
            return 0, 0
    def_root = defaults_root
    def_layout = def_root.find(".//LayoutEditor")
    if def_layout is None:
        def_layout = def_root.find(".//{*}LayoutEditor")
    if def_layout is None:
        for elem in def_root.iter():
            if elem.tag and "LayoutEditor" in elem.tag:
                def_layout = elem
                break
    if def_layout is None:
        return 0, 0
    # Idents that should be hidden (from defaults), and the default element for each (for adding if missing)
    hidden_idents = {}
    for seg in def_layout.findall(".//tracksegment") or def_layout.findall(".//{*}tracksegment"):
        if seg.get("hidden") != "yes":
            continue
        ident = seg.get("ident")
        if ident:
            hidden_idents[ident] = seg
    if not hidden_idents:
        return 0, 0
    existing_idents = set()
    for seg in layout_elem.findall(".//tracksegment") or layout_elem.findall(".//{*}tracksegment"):
        ident = seg.get("ident")
        if ident:
            existing_idents.add(ident)
    set_hidden = 0
    for seg in layout_elem.findall(".//tracksegment") or layout_elem.findall(".//{*}tracksegment"):
        if seg.get("ident") in hidden_idents:
            seg.set("hidden", "yes")
            set_hidden += 1
    added = 0
    for ident, def_seg in hidden_idents.items():
        if ident in existing_idents:
            continue
        new_seg = copy.deepcopy(def_seg)
        if ident in seg_to_block:
            new_seg.set("blockname", seg_to_block[ident])
        layout_elem.append(new_seg)
        existing_idents.add(ident)
        added += 1
    return set_hidden, added


def get_block_to_nx_sensors(layout_elem):
    """
    Build map block_name -> list of NX sensor user names (entry/exit) at that block's boundaries.
    Uses positionable point eastboundsensor/westboundsensor and layout turnout sensorA/B/C/D.
    """
    seg_block = get_segment_block_map(layout_elem)
    block_to_nx = {}
    def add(block, name):
        if block and name:
            block_to_nx.setdefault(block, []).append(name)
    for elem in layout_elem.iter():
        tag = (elem.tag or "").strip().lower()
        if "positionablepoint" not in tag:
            continue
        pt_type = (elem.get("type") or "").strip().upper()
        c1, c2 = elem.get("connect1name"), elem.get("connect2name")
        east = elem.get("eastboundsensor")
        west = elem.get("westboundsensor")
        if pt_type == "END_BUMPER":
            track_seg = c1 if c1 and (not c2 or "LINK-" in (c2 or "")) else c2
            if not track_seg or "LINK-" in (track_seg or ""):
                track_seg = c1 or c2
            block = seg_block.get(track_seg) if track_seg else None
            if block and east:
                add(block, east)
        elif pt_type == "ANCHOR" and c1 and c2:
            b1, b2 = seg_block.get(c1), seg_block.get(c2)
            if b1 and b2 and b1 != b2:
                if east:
                    add(b1, east)
                if west:
                    add(b2, west)
    for elem in layout_elem.iter():
        tag = (elem.tag or "").strip().lower()
        if "layoutturnout" not in tag:
            continue
        to_block = elem.get("blockname")
        if not to_block:
            continue
        leg_attr = {"A": "connectaname", "B": "connectbname", "C": "connectcname", "D": "connectdname"}
        for leg in ("A", "B", "C", "D"):
            child = elem.find(f"sensor{leg}")
            if child is None or not (child.text or "").strip():
                continue
            sensor_name = (child.text or "").strip()
            seg_ident = elem.get(leg_attr[leg])
            if seg_ident:
                seg_block_name = seg_block.get(seg_ident)
                if seg_block_name and seg_block_name != to_block:
                    add(to_block, sensor_name)
                    add(seg_block_name, sensor_name)
    for block in block_to_nx:
        block_to_nx[block] = sorted(set(block_to_nx[block]))
    return block_to_nx


def update_blocks_sheet_comments(excel_path, layout_elem, turnout_to_block, block_names):
    """
    Add or update the Comments column in the Blocks sheet: turnout idents for turnout blocks,
    and entry/exit (NX) sensor names for track segment blocks.
    """
    if not excel_path or not os.path.isfile(excel_path):
        return
    block_to_turnouts = {}
    for ident, block_name in turnout_to_block.items():
        block_to_turnouts.setdefault(block_name, []).append(ident)
    for block_name in block_to_turnouts:
        block_to_turnouts[block_name] = sorted(set(block_to_turnouts[block_name]))
    block_to_nx = get_block_to_nx_sensors(layout_elem)
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=False, data_only=False)
    except Exception:
        return
    if "Blocks" not in wb.sheetnames:
        wb.close()
        return
    ws = wb["Blocks"]
    headers = [c.value for c in list(ws.iter_rows(min_row=1, max_row=1))[0]]
    if "Comments" not in headers:
        col = len(headers) + 1
        ws.cell(row=1, column=col, value="Comments")
    else:
        col = headers.index("Comments") + 1
    for row_idx in range(2, ws.max_row + 1):
        block_name = ws.cell(row=row_idx, column=2).value
        if not block_name or block_name not in block_names:
            continue
        parts = []
        turnouts = block_to_turnouts.get(block_name, [])
        if turnouts:
            parts.append(", ".join(turnouts))
        nx_list = block_to_nx.get(block_name, [])
        if nx_list:
            parts.append(", ".join(nx_list))
        comment = ". ".join(parts) if parts else ""
        ws.cell(row=row_idx, column=col, value=comment)
    try:
        wb.save(excel_path)
    except Exception:
        pass
    wb.close()


def add_block14_link_segment(layout_elem, block_14_name="Block_14"):
    """Add a track segment linking EB265–EB279 so Block_14 (F35261-S-0 + T-I-TOL35290) appears connected."""
    # Find positionable points and check they exist and only have one connection
    bumpers = {}
    for elem in layout_elem.iter():
        tag = elem.tag if hasattr(elem, "tag") else None
        if tag is None:
            continue
        if "positionablepoint" not in (tag or "").lower():
            continue
        ident = elem.get("ident")
        if ident in BLOCK14_LINK_ENDS:
            bumpers[ident] = elem
    if len(bumpers) != 2:
        return
    # If link segment already present, skip
    for elem in layout_elem.iter():
        if (elem.tag or "").lower() == "tracksegment" and elem.get("ident") == BLOCK14_LINK_IDENT:
            return
    e1, e2 = BLOCK14_LINK_ENDS[0], BLOCK14_LINK_ENDS[1]
    # Create link segment (same structure as other straight segments)
    seg = ET.Element(
        "tracksegment",
        attrib={
            "ident": BLOCK14_LINK_IDENT,
            "connect1name": e1,
            "type1": "POS_POINT",
            "connect2name": e2,
            "type2": "POS_POINT",
            "dashed": "no",
            "mainline": "yes",
            "hidden": "no",
            "blockname": block_14_name,
            "class": "jmri.jmrit.display.layoutEditor.configurexml.TrackSegmentXml",
        },
    )
    # Append to layout (tracksegments are direct children of LayoutEditor)
    layout_elem.append(seg)
    # Give both bumpers a second connection to the link
    for ident in BLOCK14_LINK_ENDS:
        if ident in bumpers:
            bumpers[ident].set("connect2name", BLOCK14_LINK_IDENT)


def rename_turnout_idents_in_layout(layout_elem, old_to_new):
    """Replace every old turnout ident with T1, T2, ... in layout (ident, turnoutname, connect*)."""
    if not old_to_new:
        return
    # Longest first so e.g. TOL35007 is not partially replaced by TOL3500
    attrs = ("ident", "turnoutname", "connectaname", "connectbname", "connectcname", "connectdname", "connect1name", "connect2name", "connect3name", "connect4name")
    for old_ident in sorted(old_to_new.keys(), key=len, reverse=True):
        new_ident = old_to_new[old_ident]
        for elem in layout_elem.iter():
            if elem.tag is None:
                continue
            for attr in attrs:
                val = elem.get(attr)
                if val and (val == old_ident or old_ident in val):
                    elem.set(attr, val.replace(old_ident, new_ident))
    return


def _find_layout_in_root(root):
    layout = root.find(".//LayoutEditor")
    if layout is None:
        layout = root.find(".//{*}LayoutEditor")
    if layout is None:
        for elem in root.iter():
            if elem.tag and "LayoutEditor" in elem.tag:
                return elem
    return layout


def collect_blocknames_used_on_layout(layout_elem):
    """Block user names assigned on track segments and layout turnouts (after blockname apply)."""
    used = set()
    for seg in layout_elem.findall(".//tracksegment") or layout_elem.findall(".//{*}tracksegment"):
        bn = seg.get("blockname")
        if bn and str(bn).strip():
            used.add(str(bn).strip())
    for to in layout_elem.findall(".//layoutturnout") or layout_elem.findall(".//{*}layoutturnout"):
        bn = to.get("blockname")
        if bn and str(bn).strip():
            used.add(str(bn).strip())
    return used


def strip_nx_attributes_from_layout(layout_elem):
    """Remove Entry/Exit (NX) prerequisites: point boundary attrs and turnout sensorA–D children."""
    n = 0
    for elem in layout_elem.iter():
        tag = (elem.tag or "").strip().lower()
        if "positionablepoint" in tag:
            for attr in ("eastboundsensor", "westboundsensor"):
                if elem.get(attr):
                    del elem.attrib[attr]
                    n += 1
        if "layoutturnout" in tag or "layoutslip" in tag or "layoutxing" in tag:
            for child in list(elem):
                ct = (child.tag or "").strip().lower()
                if ct in ("sensora", "sensorb", "sensorc", "sensord"):
                    elem.remove(child)
                    n += 1
    return n


def _find_entryexitpairs_and_signalmastlogics(root):
    """
    Find entryexitpairs and signalmastlogics in root (direct children of layout-config).
    Returns (signalmastlogics_elem or None, entryexitpairs_elem or None).
    """
    signalmastlogics = None
    entryexitpairs = None
    for child in root:
        tag = (child.tag or "").strip().lower()
        if tag == "signalmastlogics":
            signalmastlogics = child
        elif tag == "entryexitpairs":
            entryexitpairs = child
    return signalmastlogics, entryexitpairs


def apply_blocks(
    panel_path,
    output_path,
    seg_to_block,
    turnout_to_block,
    block_names,
    reference_path=None,
    defaults_path=None,
    layout_override_path=None,
    use_panel_layout=False,
    no_nx=False,
    minimal_blocks=False,
):
    tree = ET.parse(panel_path)
    root = tree.getroot()
    layout = _find_layout_in_root(root)
    if layout is None:
        raise SystemExit("LayoutEditor not found in panel XML")
    # When use_panel_layout: keep the panel file's layout (e.g. upper_both4 track geometry). May show import error in JMRI; open and save once to clear.
    if use_panel_layout:
        print(f"  Using layout from panel file {os.path.basename(panel_path)} (track plan geometry)")
    # When NOT use_panel_layout: use layout override (mac_jmri2) when set so we don't carry AnyRail import errors.
    elif layout_override_path and os.path.isfile(layout_override_path) and os.path.abspath(layout_override_path) != os.path.abspath(panel_path):
        override_tree = ET.parse(layout_override_path)
        override_root = override_tree.getroot()
        override_layout = _find_layout_in_root(override_root)
        if override_layout is not None:
            layout = override_layout
            if defaults_path is None:
                defaults_path = layout_override_path
            print(f"  Using layout from {os.path.basename(layout_override_path)} (JMRI-cleaned)")

    # Apply blockname to segments and turnouts (in place)
    seg_count = 0
    to_count = 0
    for seg in layout.findall(".//tracksegment") or layout.findall(".//{*}tracksegment"):
        ident = seg.get("ident")
        if ident and ident in seg_to_block:
            seg.set("blockname", seg_to_block[ident])
            seg_count += 1
    for to in layout.findall(".//layoutturnout") or layout.findall(".//{*}layoutturnout"):
        ident = to.get("ident") or to.get("turnoutname")
        if ident and ident in turnout_to_block:
            to.set("blockname", turnout_to_block[ident])
            to_count += 1

    # Prepare layout copy first so we can optionally drop unused blocks from tables before building sensors.
    defaults_root = None
    if defaults_path and os.path.isfile(defaults_path):
        try:
            defaults_root = ET.parse(defaults_path).getroot()
        except Exception:
            pass
    out_root = defaults_root if (use_panel_layout and defaults_root is not None) else root
    jmriversion = out_root.find("jmriversion")
    if jmriversion is None:
        jmriversion = next((c for c in out_root if c.tag == "jmriversion" or "jmriversion" in str(c.tag)), None)

    turnouts_elem, old_to_new = copy_turnouts_from_reference(
        reference_path, simplify_turnout_names=SIMPLIFY_TURNOUT_NAMES
    )
    layout_copy = copy.deepcopy(layout)
    if old_to_new:
        rename_turnout_idents_in_layout(layout_copy, old_to_new)
    set_turnoutnames_in_layout(layout_copy)  # link each layout turnout to roster (turnoutname=ident)
    add_block14_link_segment(layout_copy, block_14_name="Block_14")  # connect F35261-S-0 + T-I-TOL35290
    if use_panel_layout:
        # Layout is from AnyRail (e.g. layout-4-19-2). Remove all positionablelabels so we don't
        # get PositionableLabel import error (icon label from AnyRail or 5.x-format labels).
        n_removed = remove_all_positionable_labels(layout_copy)
        if n_removed:
            print(f"  Positionable labels removed: {n_removed} (use-panel-layout: avoids import error)")
    else:
        n_icon_labels = remove_icon_positionable_labels(layout_copy)
        if n_icon_labels:
            print(f"  Icon positionable labels removed: {n_icon_labels} (avoids PositionableLabel import error)")
    if defaults_path:
        if apply_layout_defaults(layout_copy, defaults_path, defaults_root=defaults_root):
            print(f"  Layout defaults applied from {os.path.basename(defaults_path)}")
        # With use-panel-layout, geometry is from AnyRail/scaled export; do not copy mac labels or hidden segments.
        if not use_panel_layout:
            n_labels = add_labels_from_defaults(layout_copy, defaults_path, defaults_root=defaults_root)
            if n_labels:
                print(f"  Labels from defaults: {n_labels} added")
            set_hidden, added = apply_hidden_tracks_from_defaults(
                layout_copy, defaults_path, seg_to_block, defaults_root=defaults_root
            )
            if set_hidden or added:
                print(f"  Hidden tracks from defaults: {set_hidden} set hidden, {added} added")

    if minimal_blocks:
        used = collect_blocknames_used_on_layout(layout_copy)
        before_ct = len(block_names)
        retained = [b for b in block_names if b in used]
        block_names[:] = retained  # update caller list for main() summary
        if len(block_names) < before_ct:
            print(f"  Minimal blocks: {before_ct} -> {len(block_names)} (only names on segments/turnouts)")

    blockrouting = "no" if no_nx else "yes"
    if no_nx:
        nx_removed = strip_nx_attributes_from_layout(layout_copy)
        if nx_removed:
            print(f"  NX routing stripped from layout ({nx_removed} attrs / turnout sensor elements)")

    # Build output root: jmriversion, sensors, turnouts, blocks, layoutblocks, LayoutEditor
    new_root = ET.Element(out_root.tag, dict(out_root.attrib))
    if use_panel_layout and defaults_root is not None:
        new_root.attrib["{http://www.w3.org/2001/XMLSchema-instance}noNamespaceSchemaLocation"] = "http://jmri.org/xml/schema/layout-5-5-5.xsd"
    if jmriversion is not None:
        new_root.append(copy.deepcopy(jmriversion))
    sensors_elem = None
    n_fb = 0
    if block_names:
        sensors_elem = build_sensors_element(block_names, turnout_to_block)
        new_root.append(sensors_elem)
    if turnouts_elem is not None:
        new_root.append(turnouts_elem)
        if sensors_elem is not None:
            fb_base = len(block_names) + 1
            n_fb = add_turnout_feedback_sensors(sensors_elem, turnouts_elem, system_name_base=fb_base)
            if n_fb:
                print(f"  Turnout feedback sensors: {n_fb} turnouts (userName FB_N/FB_R, sensor1=Thrown, sensor2=Closed)")
    if sensors_elem is not None and not no_nx:
        seg_block_map = get_segment_block_map(layout_copy)
        nx_start = len(block_names) + 1 + 2 * n_fb
        n_boundaries = add_boundary_sensors_to_layout(layout_copy, seg_block_map, sensors_elem, nx_sensor_start=nx_start)
        if n_boundaries:
            print(f"  Boundary sensors: {n_boundaries} points with NX sensors (Entry/Exit)")
    if sensors_elem is not None:
        n_rm_icons = remove_sensoricons_from_layout(layout_copy)
        if n_rm_icons:
            print(f"  Sensor icons removed from layout: {n_rm_icons} (place icons manually in JMRI if needed)")
    update_blocks_sheet_comments(EXCEL_FILE, layout_copy, turnout_to_block, block_names)
    if block_names:
        generated_comments = get_generated_block_comments(layout_copy, turnout_to_block, block_names)
        block_to_nx = get_block_to_nx_sensors(layout_copy)
        if sensors_elem is not None:
            add_sensor_comments(sensors_elem, block_names, generated_comments, block_to_nx, turnout_to_block=turnout_to_block)
        merged_comments = {k: v for k, v in generated_comments.items() if v}
        new_root.append(build_blocks_element(block_names, comments=merged_comments, turnout_to_block=turnout_to_block))
        new_root.append(
            build_layoutblocks_element(block_names, turnout_to_block=turnout_to_block, blockrouting=blockrouting)
        )
    # Do not copy entry-exit pairs into output so the panel can be deleted/reloaded without pairs persisting.
    # You can recreate pairs in JMRI (Layout Editor → Tools → Entry Exit → Add Pair or Auto Generate).
    copy_source = defaults_root if defaults_root is not None else root
    sig_logics, _entry_exit = _find_entryexitpairs_and_signalmastlogics(copy_source)
    if sig_logics is not None:
        new_root.append(copy.deepcopy(sig_logics))
    new_root.append(layout_copy)

    out_tree = ET.ElementTree(new_root)
    ET.indent(out_tree, space="  ", level=0)

    with open(output_path, "w", encoding="utf-8") as f:
        f.write('<?xml version="1.0" encoding="UTF-8"?>\n')
        f.write('<?xml-stylesheet href="/xml/XSLT/panelfile-5-5-5.xsl" type="text/xsl"?>\n')
        out_tree.write(
            f,
            encoding="unicode",
            default_namespace=None,
            method="xml",
            xml_declaration=False,
        )
    return seg_count, to_count


def main():
    panel = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_PANEL
    output = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_OUTPUT
    defaults = sys.argv[3] if len(sys.argv) > 3 else STYLE_DEFAULTS
    use_panel_layout = len(sys.argv) > 4 and (sys.argv[4].lower() in ("use-panel-layout", "no-override", "1", "yes"))
    extra = [a.lower() for a in sys.argv[5:]]
    no_nx = any(x in ("no-nx", "no-nx-routing") for x in extra)
    minimal_blocks = any(x in ("minimal-blocks", "used-blocks-only") for x in extra)
    if not os.path.isfile(panel):
        print(f"Panel file not found: {panel}")
        sys.exit(1)
    if not os.path.isfile(EXCEL_FILE):
        print(f"Excel file not found: {EXCEL_FILE}")
        sys.exit(1)

    seg_to_block, turnout_to_block = load_mappings(EXCEL_FILE)
    block_names = load_block_names(EXCEL_FILE)
    merge_pairs = load_merge_pairs(MERGE_FILE)
    if merge_pairs:
        block_names = apply_merges(block_names, seg_to_block, turnout_to_block, merge_pairs)
        print(f"Applied {len(merge_pairs)} block merge(s); blocks count now {len(block_names)}")
    print(f"Loaded {len(seg_to_block)} segment->block, {len(turnout_to_block)} turnout->block, {len(block_names)} blocks")

    ref = REFERENCE_FILE
    if not os.path.isfile(ref):
        ref = defaults if (defaults and os.path.isfile(defaults)) else panel
    layout_override = None if use_panel_layout else (LAYOUT_OVERRIDE_PANEL if os.path.isfile(LAYOUT_OVERRIDE_PANEL) else None)
    if use_panel_layout:
        print("  Refreshing from AnyRail: using panel file layout (track geometry); if JMRI shows import error, open and save once to clear")
    if defaults and os.path.isfile(defaults):
        print(f"  Style defaults: {os.path.basename(defaults)}")
    seg_count, to_count = apply_blocks(
        panel, output, seg_to_block, turnout_to_block, block_names,
        reference_path=ref, defaults_path=defaults, layout_override_path=layout_override,
        use_panel_layout=use_panel_layout,
        no_nx=no_nx,
        minimal_blocks=minimal_blocks,
    )
    print(f"Wrote {output}")
    print(
        f"  Tables: jmriversion, sensors, turnouts (from {os.path.basename(ref)}), blocks, layoutblocks, LayoutEditor"
        + ("; no NX routing" if no_nx else "")
        + ("; minimal blocks only" if minimal_blocks else "")
    )
    print(f"  Sensors: {len(block_names)}; Blocks: {len(block_names)} with occupancysensor; blockname on {seg_count} segments, {to_count} turnouts")


if __name__ == "__main__":
    main()
