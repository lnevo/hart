#!/usr/bin/env python3
"""
Build a block map from JMRI layout XML: blocks = track between two turnouts (or turnout and endpoint).
Output: Excel with Block Name, track segments, and end labels (Throat/Normal/Diverging) per turnout.
"""
import xml.etree.ElementTree as ET
from collections import defaultdict, deque
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import re
import os
import sys

JMRI_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, JMRI_ROOT)
from layout_paths import layout_paths

_PATHS = layout_paths()
LAYOUT_FILE = _PATHS["authoritative"]
EXCEL_FILE = _PATHS["excel"]

# JMRI: TURNOUT_A = throat (inlet), TURNOUT_B = one leg, TURNOUT_C = other leg.
# We label A=Throat, B=Normal, C=Diverging (convention; user can swap in Excel if needed).
CONN_LABEL = {"TURNOUT_A": "Throat", "TURNOUT_B": "Normal", "TURNOUT_C": "Diverging"}

# Double crossover: A,B,C,D - we label as A/B/C/D or Throat/Left/Right/Throat2 etc.
def conn_label_dxover(conn_name):
    if "SL-A-" in conn_name: return "A (Throat1)"
    if "SL-B-" in conn_name: return "B"
    if "SL-C-" in conn_name: return "C"
    if "SL-D-" in conn_name: return "D (Throat2)"
    return conn_name

# 3-way: T-I- = inlet, TO_3W35179-CS = common, T-E1-/T-E2- = legs, T-O- = other
def conn_label_3way(conn_name):
    if "T-I-" in conn_name: return "Throat"
    if "TO_3W" in conn_name and "-CS" in conn_name: return "Common"
    if "T-E1-" in conn_name: return "Leg1"
    if "T-E2-" in conn_name: return "Leg2"
    if "T-O-" in conn_name: return "Other"
    return conn_name


def parse_layout(path):
    tree = ET.parse(path)
    root = tree.getroot()
    ns = {"": "http://www.w3.org/2001/XMLSchema-instance"}  # layout-config has no ns
    # JMRI uses no namespace for layout elements
    layout = root.find(".//LayoutEditor")
    if layout is None:
        layout = root.find(".//{*}LayoutEditor")
    if layout is None:
        # Single-line layout: root might be layout-config and children direct
        for elem in root.iter():
            if "LayoutEditor" in elem.tag:
                layout = elem
                break
    if layout is None:
        raise SystemExit("LayoutEditor not found in XML")

    turnouts = {}  # ident -> {type, connectaname, connectbname, connectcname, connectdname?}
    conn_to_turnout = {}  # connection point name -> (turnout_ident, leg_label)

    for to in layout.findall(".//layoutturnout") or layout.findall(".//{*}layoutturnout"):
        ident = to.get("ident") or to.get("turnoutname")
        if not ident:
            continue
        typ = to.get("type", "")
        conn_a = to.get("connectaname") or ""
        conn_b = to.get("connectbname") or ""
        conn_c = to.get("connectcname") or ""
        conn_d = to.get("connectdname") or ""
        turnouts[ident] = {"type": typ, "conn_a": conn_a, "conn_b": conn_b, "conn_c": conn_c, "conn_d": conn_d}
        if typ == "DOUBLE_XOVER":
            conn_to_turnout[conn_a] = (ident, conn_label_dxover(conn_a))
            conn_to_turnout[conn_b] = (ident, conn_label_dxover(conn_b))
            conn_to_turnout[conn_c] = (ident, conn_label_dxover(conn_c))
            conn_to_turnout[conn_d] = (ident, conn_label_dxover(conn_d))
        elif "3W" in ident or "TO_3W" in conn_a:
            conn_to_turnout[conn_a] = (ident, conn_label_3way(conn_a))
            conn_to_turnout[conn_b] = (ident, conn_label_3way(conn_b))
            conn_to_turnout[conn_c] = (ident, conn_label_3way(conn_c))
        else:
            conn_to_turnout[conn_a] = (ident, "Throat")
            conn_to_turnout[conn_b] = (ident, "Normal")
            conn_to_turnout[conn_c] = (ident, "Diverging")

    # Map (turnout_ident, type1 string) -> connection point name
    def resolve_vertex(name, typ):
        if typ == "POS_POINT":
            return name
        # typ is TURNOUT_A, TURNOUT_B, TURNOUT_C
        t = turnouts.get(name, {})
        if typ == "TURNOUT_A": return t.get("conn_a") or name
        if typ == "TURNOUT_B": return t.get("conn_b") or name
        if typ == "TURNOUT_C": return t.get("conn_c") or name
        if typ == "TURNOUT_D": return t.get("conn_d") or name
        return name

    segments = []  # (ident, vertex1, vertex2)
    for seg in layout.findall(".//tracksegment") or layout.findall(".//{*}tracksegment"):
        ident = seg.get("ident")
        if not ident:
            continue
        c1, t1 = seg.get("connect1name"), seg.get("type1", "")
        c2, t2 = seg.get("connect2name"), seg.get("type2", "")
        v1 = resolve_vertex(c1, t1)
        v2 = resolve_vertex(c2, t2)
        segments.append((ident, v1, v2))

    return turnouts, conn_to_turnout, segments


def build_graph(segments, conn_to_turnout):
    """Vertices = anchors (POS_POINT) and turnout connection point names. Edges = segments."""
    adj = defaultdict(list)  # vertex -> [(segment_id, other_vertex), ...]
    for ident, v1, v2 in segments:
        adj[v1].append((ident, v2))
        adj[v2].append((ident, v1))
    return adj


def is_terminal(vertex, adj, conn_to_turnout):
    """Block boundary: turnout connection point, or anchor with only one segment (bumper/end)."""
    if vertex in conn_to_turnout:
        return True
    return len(adj.get(vertex, [])) == 1


def find_block_terminals(segment_id, segments, adj, conn_to_turnout):
    """For the block containing segment_id, return (set of segment idents, terminal1, terminal2)."""
    seg_by_id = {s[0]: (s[1], s[2]) for s in segments}
    if segment_id not in seg_by_id:
        return set(), None, None
    # Two segments are in the same block if they share a non-terminal (anchor) vertex.
    # BFS over segments: start with segment_id, add neighbors that share an anchor.
    visited_segs = {segment_id}
    q = deque([segment_id])
    terminals = set()

    while q:
        seg_id = q.popleft()
        v1, v2 = seg_by_id[seg_id]
        for v in (v1, v2):
            if is_terminal(v, adj, conn_to_turnout):
                terminals.add(v)
            else:
                for next_seg_id, _ in adj.get(v, []):
                    if next_seg_id not in visited_segs:
                        visited_segs.add(next_seg_id)
                        q.append(next_seg_id)

    term_list = sorted(terminals)
    t1 = term_list[0] if len(term_list) >= 1 else None
    t2 = term_list[1] if len(term_list) >= 2 else None
    return visited_segs, t1, t2


def compute_all_blocks(segments, adj, conn_to_turnout):
    """Group segments into blocks by their two terminals. Return list of (segment_set, terminal1, terminal2)."""
    done = set()
    blocks = []
    for ident, _v1, _v2 in segments:
        if ident in done:
            continue
        seg_set, t1, t2 = find_block_terminals(ident, segments, adj, conn_to_turnout)
        done |= seg_set
        blocks.append((seg_set, t1, t2))
    return blocks


def format_end(vertex, conn_to_turnout):
    """Format terminal for display: 'TurnoutName (Throat)' or 'Endpoint A123'."""
    if vertex in conn_to_turnout:
        to_name, leg = conn_to_turnout[vertex]
        return f"{to_name} ({leg})"
    return f"Endpoint {vertex}"


def main():
    turnouts, conn_to_turnout, segments = parse_layout(LAYOUT_FILE)
    adj = build_graph(segments, conn_to_turnout)
    blocks = compute_all_blocks(segments, adj, conn_to_turnout)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Blocks"

    thin = Side(style="thin")
    header_font = Font(bold=True)
    headers = [
        "Block #",
        "Block Name",
        "Block Name (suggested long)",
        "End 1 (Turnout or Endpoint)",
        "End 1 Leg",
        "End 2 (Turnout or Endpoint)",
        "End 2 Leg",
        "Track Segments (comma-separated)",
        "Segment Count",
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.alignment = Alignment(wrap_text=True, horizontal="center")
    ws.row_dimensions[1].height = 24

    for i, (seg_set, t1, t2) in enumerate(blocks, 1):
        seg_list = sorted(seg_set)
        end1_str = format_end(t1, conn_to_turnout) if t1 else ""
        end2_str = format_end(t2, conn_to_turnout) if t2 else ""
        leg1 = conn_to_turnout.get(t1, (None, ""))[1] if t1 and t1 in conn_to_turnout else ("Endpoint" if t1 else "")
        leg2 = conn_to_turnout.get(t2, (None, ""))[1] if t2 and t2 in conn_to_turnout else ("Endpoint" if t2 else "")
        to1 = conn_to_turnout.get(t1, (None, ""))[0] if t1 and t1 in conn_to_turnout else None
        to2 = conn_to_turnout.get(t2, (None, ""))[0] if t2 and t2 in conn_to_turnout else None
        suggested_name = f"Block_{i}"
        if to1 and to2:
            suggested_name = f"{to1}-{leg1}_to_{to2}-{leg2}"
        elif to1:
            suggested_name = f"{to1}-{leg1}_to_{t2 or 'End'}"
        elif to2:
            suggested_name = f"{t1 or 'End'}_to_{to2}-{leg2}"
        block_name = f"Block_{i}"

        row = [
            i,
            block_name,
            suggested_name,
            to1 or t1 or "",
            leg1 if leg1 else "",
            to2 or t2 or "",
            leg2 if leg2 else "",
            ", ".join(seg_list),
            len(seg_list),
        ]
        for col, val in enumerate(row, 1):
            ws.cell(row=i + 1, column=col, value=val)
            ws.cell(row=i + 1, column=col).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i + 1].height = max(18, min(60, 15 + 12 * (seg_list.__len__() // 8)))

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max(14, {1: 8, 2: 12, 3: 38, 4: 22, 5: 12, 6: 22, 7: 12, 8: 72, 9: 12}.get(col, 14))

    # Build block terminals for turnout assignment: block index -> (t1, t2)
    block_terminals = [(t1, t2) for (_, t1, t2) in blocks]

    # Sheet 2: Segment -> Block mapping (Block Name = Block_1, Block_2, ... for XML)
    seg_to_block = {}
    for i, (seg_set, t1, t2) in enumerate(blocks, 1):
        block_name = f"Block_{i}"
        for seg_id in seg_set:
            seg_to_block[seg_id] = (i, block_name)

    ws2 = wb.create_sheet("Segment_to_Block", 1)
    for col, h in enumerate(["Track Segment", "Block #", "Block Name"], 1):
        ws2.cell(row=1, column=col, value=h).font = header_font
    for row_idx, seg_id in enumerate(sorted(seg_to_block.keys()), 2):
        blk_num, blk_name = seg_to_block[seg_id]
        ws2.cell(row=row_idx, column=1, value=seg_id)
        ws2.cell(row=row_idx, column=2, value=blk_num)
        ws2.cell(row=row_idx, column=3, value=blk_name)
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 12

    # Sheet 3: Turnout -> Block (each turnout gets its OWN block: Block_118 .. Block_182)
    # Sort turnouts so they sit near the track blocks they touch (by min adjacent track block #)
    num_track_blocks = len(blocks)
    block_terminals = [(t1, t2) for (_, t1, t2) in blocks]
    adjacent_track_blocks = defaultdict(set)  # turnout_ident -> set of track block numbers (1-based)
    for i, (t1, t2) in enumerate(block_terminals):
        for term in (t1, t2):
            if term and term in conn_to_turnout:
                to_ident = conn_to_turnout[term][0]
                adjacent_track_blocks[to_ident].add(i + 1)
    def sort_key(to_ident):
        adj = adjacent_track_blocks.get(to_ident) or set()
        return (min(adj) if adj else 999, to_ident)
    turnout_idents_sorted = sorted(turnouts.keys(), key=sort_key)
    turnout_to_block = {}  # turnout_ident -> (block_num, block_name)
    for idx, to_ident in enumerate(turnout_idents_sorted):
        blk_num = num_track_blocks + idx + 1
        block_name = f"Block_{blk_num}"
        turnout_to_block[to_ident] = (blk_num, block_name)

    # Append turnout blocks to Blocks sheet (Block_118, Block_119, ... Block_182), sorted by adjacent track
    for idx, to_ident in enumerate(turnout_idents_sorted):
        blk_num = num_track_blocks + idx + 1
        block_name = f"Block_{blk_num}"
        row = [
            blk_num,
            block_name,
            f"Turnout block: {to_ident}",
            to_ident,
            "Turnout",
            "",
            "",
            "",
            0,
        ]
        for col, val in enumerate(row, 1):
            ws.cell(row=num_track_blocks + 2 + idx, column=col, value=val)
            ws.cell(row=num_track_blocks + 2 + idx, column=col).alignment = Alignment(wrap_text=True, vertical="top")

    ws_t2b = wb.create_sheet("Turnout_to_Block", 2)
    for col, h in enumerate(["Turnout", "Block #", "Block Name"], 1):
        ws_t2b.cell(row=1, column=col, value=h).font = header_font
    for row_idx, to_ident in enumerate(turnout_idents_sorted, 2):
        blk_num, blk_name = turnout_to_block[to_ident]
        ws_t2b.cell(row=row_idx, column=1, value=to_ident)
        ws_t2b.cell(row=row_idx, column=2, value=blk_num)
        ws_t2b.cell(row=row_idx, column=3, value=blk_name)
    ws_t2b.column_dimensions["A"].width = 22
    ws_t2b.column_dimensions["B"].width = 10
    ws_t2b.column_dimensions["C"].width = 12

    # Sheet 4: Turnouts with Throat / Normal / Diverging
    ws3 = wb.create_sheet("Turnouts", 3)
    for col, h in enumerate(["Turnout", "Type", "Throat (A)", "Normal (B)", "Diverging (C)", "Note"], 1):
        ws3.cell(row=1, column=col, value=h).font = header_font
    row_idx = 2
    for ident, t in sorted(turnouts.items()):
        typ = t["type"]
        ws3.cell(row=row_idx, column=1, value=ident)
        ws3.cell(row=row_idx, column=2, value=typ)
        ws3.cell(row=row_idx, column=3, value=t.get("conn_a", ""))
        ws3.cell(row=row_idx, column=4, value=t.get("conn_b", ""))
        ws3.cell(row=row_idx, column=5, value=t.get("conn_c", ""))
        note = ""
        if typ == "DOUBLE_XOVER":
            note = "A/B/C/D = 4 connection points"
        elif "3W" in ident:
            note = "3-way: Throat, Common, Leg1, Leg2, Other"
        ws3.cell(row=row_idx, column=6, value=note)
        row_idx += 1
    ws3.column_dimensions["A"].width = 18
    ws3.column_dimensions["B"].width = 14
    for c in "CDEF": ws3.column_dimensions[c].width = 20

    wb.save(EXCEL_FILE)
    print(f"Wrote {len(blocks)} blocks to {EXCEL_FILE}")
    print(f"  Sheets: Blocks, Segment_to_Block, Turnout_to_Block, Turnouts")
    print(f"  Turnouts assigned to blocks: {len(turnout_to_block)}")
    print(f"  Segments total: {len(segments)}, Turnout connection points: {len(conn_to_turnout)}")


if __name__ == "__main__":
    main()
